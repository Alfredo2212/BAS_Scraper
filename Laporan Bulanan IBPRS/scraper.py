"""
IBPRS Scraper
Scrapes data from IBPRS DataKeuangan page
"""

import time
import logging
import shutil
from pathlib import Path
from datetime import datetime
from selenium.webdriver.remote.webdriver import WebDriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from bs4 import BeautifulSoup

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[WARNING] openpyxl not installed. Excel export will not work. Install with: pip install openpyxl")
    Workbook = None

# Handle imports for both package and direct execution
try:
    from helper import ExtJSHelper
    from selenium_setup import SeleniumSetup
except ImportError:
    # If relative imports fail, try absolute imports
    import sys
    from pathlib import Path
    module_dir = Path(__file__).parent.parent / "Laporan Publikasi BPR Konvensional"
    if str(module_dir) not in sys.path:
        sys.path.insert(0, str(module_dir))
    from helper import ExtJSHelper
    from selenium_setup import SeleniumSetup

from config.settings import OJKConfig


class IBPRSScraper:
    """Scraper for IBPRS DataKeuangan page"""
    
    # URL for IBPRS DataKeuangan page
    URL_DATAKEUANGAN = "https://ibpr-s.ojk.go.id/DataKeuangan"
    
    def __init__(self, headless: bool = False):
        """
        Initialize the scraper
        
        Args:
            headless: Whether to run browser in headless mode (default: False for visibility)
        """
        self.driver: WebDriver = None
        self.wait: WebDriverWait = None
        self.extjs: ExtJSHelper = None
        self.headless = headless
        self.all_pages_content = []  # Store all scraped page content
        self.all_extracted_data = []  # Store structured extracted data
        
        # Setup logging
        log_dir = Path(__file__).parent.parent / "logs"
        log_dir.mkdir(exist_ok=True)
        log_file = log_dir / f"ibprs_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
        
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(log_file, encoding='utf-8'),
                logging.StreamHandler()
            ]
        )
        self.logger = logging.getLogger(__name__)
        self.logger.info(f"Log file: {log_file}")
    
    def initialize(self):
        """Initialize WebDriver and ExtJS helper"""
        if self.driver is None:
            self.driver = SeleniumSetup.create_driver(headless=self.headless)
            self.wait = SeleniumSetup.create_wait(self.driver)
            self.extjs = ExtJSHelper(self.driver, self.wait)
            
            # Don't minimize - keep window visible for monitoring
            self.logger.info("Chrome browser initialized")
    
    def cleanup(self, kill_processes: bool = True):
        """Cleanup resources"""
        if self.driver:
            try:
                # Suppress urllib3 warnings during cleanup
                import warnings
                try:
                    import urllib3
                    urllib3.disable_warnings()
                except ImportError:
                    pass
                
                # Suppress warnings during cleanup
                with warnings.catch_warnings():
                    warnings.simplefilter("ignore")
                    # Small delay to let connections close naturally
                    time.sleep(0.5)
                    self.driver.quit()
                    self.logger.info("Browser closed")
            except Exception as e:
                # Ignore connection errors during cleanup (browser is already closing)
                error_str = str(e).lower()
                if "connection" not in error_str and "refused" not in error_str and "session" not in error_str:
                    self.logger.error(f"Error closing browser: {e}")
                # Connection errors are expected during cleanup, so we silently ignore them
        
        if kill_processes:
            try:
                import sys
                from pathlib import Path
                module_dir = Path(__file__).parent.parent / "Laporan Publikasi BPR Konvensional"
                if str(module_dir) not in sys.path:
                    sys.path.insert(0, str(module_dir))
                from utils import kill_chrome_processes
                kill_chrome_processes()
            except Exception as e:
                self.logger.debug(f"Error killing Chrome processes: {e}")
    
    def navigate_to_page(self):
        """Navigate to the IBPRS DataKeuangan page and wait for it to load with retry logic"""
        if self.driver is None:
            self.initialize()
        
        # Switch to default content in case we're in an iframe
        try:
            self.driver.switch_to.default_content()
        except:
            pass
        
        self.logger.info(f"Navigating to: {self.URL_DATAKEUANGAN}")
        
        # Set page load timeout to handle slow pages
        self.driver.set_page_load_timeout(30)
        
        max_retries = 2
        for attempt in range(max_retries + 1):  # 0, 1, 2 = 3 attempts total
            try:
                if attempt > 0:
                    self.logger.info(f"Retry attempt {attempt}/{max_retries}: Refreshing page...")
                    self.driver.refresh()
                else:
                    self.driver.get(self.URL_DATAKEUANGAN)
            except TimeoutException:
                self.logger.warning("Page load timeout, but continuing...")
                # Try to stop loading
                try:
                    self.driver.execute_script("window.stop();")
                except:
                    pass
            
            # Wait 10 seconds for page to load
            self.logger.info(f"Waiting 10 seconds for page to load (attempt {attempt + 1}/{max_retries + 1})...")
            time.sleep(10.0)
            
            # Check if page loaded successfully
            page_loaded = False
            try:
                # Check if page is ready
                ready_state = self.driver.execute_script("return document.readyState")
                if ready_state == "complete":
                    page_loaded = True
                    self.logger.info("Page readyState is complete")
            except:
                pass
            
            # Try to detect if input field is present (indicates page is loaded)
            input_detected = False
            try:
                # Check for the input field with id="search_4" as indicator that page loaded
                element = self.driver.find_element(By.ID, "search_4")
                if element.is_displayed():
                    input_detected = True
                    self.logger.info("Input field detected using id='search_4'")
            except:
                pass
            
            # If page is loaded and input is detected, we're good
            if page_loaded and input_detected:
                self.logger.info("Page loaded successfully and input field detected")
                
                # Check if ExtJS is available (like other OJK pages)
                try:
                    if self.extjs and self.extjs.check_extjs_available():
                        self.logger.info("ExtJS detected on page")
                        # Wait a bit more for ExtJS to initialize
                        time.sleep(2.0)
                except:
                    self.logger.debug("ExtJS check failed or not available")
                
                # Additional wait for any dynamic content
                time.sleep(2.0)
                return  # Success, exit retry loop
            
            # If we get here, page might not be fully loaded
            if attempt < max_retries:
                self.logger.warning(f"Page not fully detected on attempt {attempt + 1}, will retry...")
            else:
                self.logger.warning("Page detection failed after all retries, continuing anyway...")
        
        # Final wait even if detection failed
        time.sleep(2.0)
    
    def input_province(self, province_name: str = "Kep. Riau"):
        """
        Find the input field and type the province name
        
        Args:
            province_name: Name of the province to input (default: "Kep. Riau")
        """
        if self.driver is None:
            self.logger.error("Driver not initialized. Call initialize() or navigate_to_page() first.")
            return False
        
        try:
            self.logger.info(f"Looking for input field with id='search_4' to type: {province_name}")
            
            # Wait for the input field with id="search_4" to be present and visible
            try:
                input_element = self.wait.until(
                    EC.presence_of_element_located((By.ID, "search_4"))
                )
                # Wait for it to be visible
                WebDriverWait(self.driver, 10).until(
                    EC.visibility_of(input_element)
                )
                self.logger.info("Found input field with id='search_4'")
            except TimeoutException:
                self.logger.error("Could not find input field with id='search_4'")
                return False
            
            # Scroll to element if needed
            self.driver.execute_script("arguments[0].scrollIntoView(true);", input_element)
            time.sleep(0.5)
            
            # Clear the field first
            input_element.clear()
            time.sleep(0.5)
            
            # Type the province name
            input_element.send_keys(province_name)
            time.sleep(1.0)
            
            self.logger.info(f"Successfully typed '{province_name}' into input field (id='search_4')")
            return True
            
        except Exception as e:
            self.logger.error(f"Error inputting province name: {e}")
            import traceback
            self.logger.debug(traceback.format_exc())
            return False
    
    def click_search_button(self):
        """
        Find and click the button with id="Cari"
        """
        if self.driver is None:
            self.logger.error("Driver not initialized. Call initialize() or navigate_to_page() first.")
            return False
        
        try:
            self.logger.info("Looking for search button with id='Cari'")
            
            # Wait for button to be present and clickable
            search_button = self.wait.until(
                EC.element_to_be_clickable((By.ID, "Cari"))
            )
            
            # Scroll to button if needed
            self.driver.execute_script("arguments[0].scrollIntoView(true);", search_button)
            time.sleep(0.5)
            
            # Click the button
            search_button.click()
            self.logger.info("Successfully clicked search button")
            
            # Wait for results to load (give some time for page to respond)
            time.sleep(2.0)
            
            return True
            
        except TimeoutException:
            self.logger.error("Timeout: Could not find or click button with id='Cari'")
            return False
        except Exception as e:
            self.logger.error(f"Error clicking search button: {e}")
            import traceback
            self.logger.debug(traceback.format_exc())
            return False
    
    def test_navigation(self, province_name: str = "Kep. Riau"):
        """
        Test the complete navigation flow:
        1. Navigate to page
        2. Input province name
        3. Click search button
        
        Args:
            province_name: Name of the province to search for
        """
        try:
            self.logger.info("=" * 70)
            self.logger.info("Starting IBPRS navigation test")
            self.logger.info("=" * 70)
            
            # Initialize browser
            self.initialize()
            
            # Navigate to page
            self.navigate_to_page()
            
            # Input province
            if not self.input_province(province_name):
                self.logger.error("Failed to input province name")
                return False
            
            # Click search button
            if not self.click_search_button():
                self.logger.error("Failed to click search button")
                return False
            
            self.logger.info("=" * 70)
            self.logger.info("Navigation test completed successfully")
            self.logger.info("=" * 70)
            
            # Keep browser open for a few seconds to see results
            self.logger.info("Keeping browser open for 5 seconds to view results...")
            self.logger.info("You can now check the search results in the browser window.")
            time.sleep(5.0)
            self.logger.info("Wait time completed.")
            
            return True
            
        except Exception as e:
            self.logger.error(f"Error during navigation test: {e}")
            import traceback
            self.logger.error(traceback.format_exc())
            return False
        finally:
            # Note: cleanup is commented out so user can see the results
            # Uncomment if you want to auto-close browser
            # self.cleanup()
            pass
    
    def extract_page_content(self) -> BeautifulSoup:
        """
        Extract page content using BeautifulSoup
        
        Returns:
            BeautifulSoup object of the current page
        """
        try:
            # Get page source
            page_source = self.driver.page_source
            soup = BeautifulSoup(page_source, 'html.parser')
            self.logger.info("Page content extracted with BeautifulSoup")
            return soup
        except Exception as e:
            self.logger.error(f"Error extracting page content: {e}")
            import traceback
            self.logger.debug(traceback.format_exc())
            return None
    
    def extract_table_data(self, soup: BeautifulSoup) -> list:
        """
        Extract structured data from <tbody class="fs-7"> table rows
        
        Args:
            soup: BeautifulSoup object of the page
            
        Returns:
            List of dictionaries, each containing:
            - nama_bpr: NAMA BPR/S (from <a> tag in first <td>)
            - jenis: JENIS (from second <td>)
            - kab_kota: KAB/KOTA (from third <td>)
            - provinsi: PROVINSI (from fourth <td>, removing <mark> tags)
            - aset: ASET (from fifth <td>)
            - dana_pihak_ketiga: DANA PIHAK KETIGA (from sixth <td>)
            - kredit_pembiayaan: KREDIT/PEMBIAYAAN (from seventh <td>)
        """
        extracted_data = []
        
        try:
            # Find tbody with class="fs-7"
            tbody = soup.find('tbody', class_='fs-7')
            if not tbody:
                self.logger.warning("Could not find tbody with class='fs-7'")
                return extracted_data
            
            # Find all rows
            rows = tbody.find_all('tr')
            self.logger.info(f"Found {len(rows)} rows in tbody.fs-7")
            
            for row in rows:
                try:
                    # Get all <td> elements
                    tds = row.find_all('td')
                    if len(tds) < 7:
                        self.logger.debug(f"Row has less than 7 columns, skipping")
                        continue
                    
                    # Extract NAMA BPR/S (first <td>, get text from <a> tag if exists)
                    first_td = tds[0]
                    a_tag = first_td.find('a')
                    nama_bpr = a_tag.get_text(strip=True) if a_tag else first_td.get_text(strip=True)
                    
                    # Extract JENIS (second <td>)
                    jenis = tds[1].get_text(strip=True)
                    
                    # Extract KAB/KOTA (third <td>)
                    kab_kota = tds[2].get_text(strip=True)
                    
                    # Extract PROVINSI (fourth <td>, remove <mark> tags)
                    provinsi_td = tds[3]
                    # Remove all <mark> tags and get text
                    for mark in provinsi_td.find_all('mark'):
                        mark.unwrap()  # Remove mark tag but keep text
                    provinsi = provinsi_td.get_text(strip=True)
                    
                    # Extract ASET (fifth <td>)
                    aset = tds[4].get_text(strip=True)
                    
                    # Extract DANA PIHAK KETIGA (sixth <td>)
                    dana_pihak_ketiga = tds[5].get_text(strip=True)
                    
                    # Extract KREDIT/PEMBIAYAAN (seventh <td>)
                    kredit_pembiayaan = tds[6].get_text(strip=True)
                    
                    # Create record dictionary
                    record = {
                        'nama_bpr': nama_bpr,
                        'jenis': jenis,
                        'kab_kota': kab_kota,
                        'provinsi': provinsi,
                        'aset': aset,
                        'dana_pihak_ketiga': dana_pihak_ketiga,
                        'kredit_pembiayaan': kredit_pembiayaan
                    }
                    
                    extracted_data.append(record)
                    
                except Exception as e:
                    self.logger.debug(f"Error extracting row data: {e}")
                    continue
            
            self.logger.info(f"Extracted {len(extracted_data)} records from table")
            return extracted_data
            
        except Exception as e:
            self.logger.error(f"Error extracting table data: {e}")
            import traceback
            self.logger.debug(traceback.format_exc())
            return extracted_data
    
    def has_next_button(self) -> bool:
        """
        Check if there is a next button (<i class="next"></i>)
        Checks parent and parent's parent for "disabled" class
        
        Returns:
            True if next button exists and is clickable, False otherwise
        """
        try:
            # Look for <i class="next"></i> element
            next_button = self.driver.find_element(By.CSS_SELECTOR, "i.next")
            if not next_button or not next_button.is_displayed():
                return False
            
            # Check parent for disabled class
            try:
                parent = next_button.find_element(By.XPATH, "./..")
                if parent:
                    classes = parent.get_attribute("class") or ""
                    if "disabled" in classes.lower():
                        return False
                    
                    # Check parent's parent for disabled class
                    try:
                        parent_parent = parent.find_element(By.XPATH, "./..")
                        if parent_parent:
                            parent_classes = parent_parent.get_attribute("class") or ""
                            if "disabled" in parent_classes.lower():
                                return False
                    except:
                        pass
            except:
                pass
            
            return True
        except NoSuchElementException:
            return False
        except Exception as e:
            self.logger.debug(f"Error checking for next button: {e}")
            return False
    
    def click_next_button(self) -> bool:
        """
        Find and click the next button (<i class="next"></i>)
        
        Returns:
            True if clicked successfully, False otherwise
        """
        try:
            self.logger.info("Looking for next button (<i class='next'>)...")
            
            # Wait for next button to be present
            next_button = self.wait.until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "i.next"))
            )
            
            # Check if it's disabled
            try:
                parent = next_button.find_element(By.XPATH, "./..")
                if parent:
                    classes = parent.get_attribute("class") or ""
                    if "disabled" in classes.lower():
                        self.logger.info("Next button is disabled, no more pages")
                        return False
            except:
                pass
            
            # Scroll to button if needed
            self.driver.execute_script("arguments[0].scrollIntoView(true);", next_button)
            time.sleep(0.5)
            
            # Click the button (try clicking parent if button itself doesn't work)
            try:
                next_button.click()
            except:
                # Try clicking parent element
                try:
                    parent = next_button.find_element(By.XPATH, "./..")
                    parent.click()
                except:
                    # Try JavaScript click
                    self.driver.execute_script("arguments[0].click();", next_button)
            
            self.logger.info("Successfully clicked next button")
            
            # Wait for page to load after clicking
            time.sleep(3.0)
            
            return True
            
        except TimeoutException:
            self.logger.info("Next button not found or not clickable")
            return False
        except Exception as e:
            self.logger.error(f"Error clicking next button: {e}")
            import traceback
            self.logger.debug(traceback.format_exc())
            return False
    
    def scrape_all_pages(self, province_name: str = "Kep. Riau"):
        """
        Scrape all pages by:
        1. Navigate and search
        2. Extract page content with BeautifulSoup
        3. Click next button
        4. Repeat until no more next button
        
        Args:
            province_name: Name of the province to search for
        """
        try:
            self.logger.info("=" * 70)
            self.logger.info("Starting IBPRS page scraping")
            self.logger.info("=" * 70)
            
            # Clear previous data
            self.all_pages_content = []
            self.all_extracted_data = []
            
            # Initialize browser
            self.initialize()
            
            # Navigate to page
            self.navigate_to_page()
            
            # Input province
            if not self.input_province(province_name):
                self.logger.error("Failed to input province name")
                return False
            
            # Click search button
            if not self.click_search_button():
                self.logger.error("Failed to click search button")
                return False
            
            # Wait for results to load
            time.sleep(3.0)
            
            page_number = 1
            
            # Loop through all pages
            while True:
                self.logger.info(f"Scraping page {page_number}...")
                
                # Extract page content
                soup = self.extract_page_content()
                if soup:
                    self.all_pages_content.append({
                        'page': page_number,
                        'soup': soup,
                        'html': str(soup)
                    })
                    self.logger.info(f"Page {page_number} content extracted")
                    
                    # Extract structured table data
                    table_data = self.extract_table_data(soup)
                    if table_data:
                        # Add page number to each record for tracking
                        for record in table_data:
                            record['page_number'] = page_number
                        self.all_extracted_data.extend(table_data)
                        self.logger.info(f"Extracted {len(table_data)} records from page {page_number}")
                
                # Check if there's a next button
                if not self.has_next_button():
                    self.logger.info(f"No more pages. Total pages scraped: {page_number}")
                    break
                
                # Click next button
                if not self.click_next_button():
                    self.logger.info(f"Could not click next button. Total pages scraped: {page_number}")
                    break
                
                page_number += 1
            
            self.logger.info("=" * 70)
            self.logger.info(f"Scraping completed. Total pages: {page_number}")
            self.logger.info("=" * 70)
            
            return True
            
        except Exception as e:
            self.logger.error(f"Error during page scraping: {e}")
            import traceback
            self.logger.error(traceback.format_exc())
            return False
    
    def save_to_txt(self, filename: str = None):
        """
        Save all scraped page content to a .txt file
        
        Args:
            filename: Optional filename. If None, generates automatic filename
        """
        if not self.all_pages_content:
            self.logger.warning("No content to save to .txt file")
            return None
        
        try:
            # Save to destination directory: D:\APP\OSS\client\assets-no_backup\ibprs
            destination_dir = Path(r"D:\APP\OSS\client\assets-no_backup\ibprs")
            destination_dir.mkdir(parents=True, exist_ok=True)
            
            # Use fixed filename that will overwrite previous files
            filename = "IBPRS.txt"
            
            filepath = destination_dir / filename
            
            # Write all pages to file
            with open(filepath, 'w', encoding='utf-8') as f:
                for page_data in self.all_pages_content:
                    f.write(f"{'=' * 70}\n")
                    f.write(f"PAGE {page_data['page']}\n")
                    f.write(f"{'=' * 70}\n\n")
                    f.write(page_data['html'])
                    f.write("\n\n")
            
            self.logger.info(f"Saved content to .txt file: {filepath}")
            return filepath
            
        except Exception as e:
            self.logger.error(f"Error saving to .txt file: {e}")
            import traceback
            self.logger.debug(traceback.format_exc())
            return None
    
    def save_to_excel(self, filename: str = None):
        """
        Save all extracted structured data to an .xlsx file
        Each record is saved as a row with proper column headers
        
        Args:
            filename: Optional filename. If None, generates automatic filename
        """
        if not Workbook:
            self.logger.error("openpyxl not installed. Cannot create Excel file.")
            return None
        
        if not self.all_extracted_data:
            self.logger.warning("No extracted data to save to Excel file")
            return None
        
        try:
            # Save to destination directory: D:\APP\OSS\client\assets-no_backup\ibprs
            destination_dir = Path(r"D:\APP\OSS\client\assets-no_backup\ibprs")
            destination_dir.mkdir(parents=True, exist_ok=True)
            
            # Use fixed filename that will overwrite previous files
            filename = "IBPRS.xlsx"
            
            filepath = destination_dir / filename
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "IBPRS Data"
            
            # Define styles
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            center_align = Alignment(horizontal="center", vertical="center")
            left_align = Alignment(horizontal="left", vertical="center")
            right_align = Alignment(horizontal="right", vertical="center")
            
            # Define column headers
            headers = [
                "NAMA BPR/S",
                "JENIS",
                "KAB/KOTA",
                "PROVINSI",
                "ASET",
                "DANA PIHAK KETIGA",
                "KREDIT/PEMBIAYAAN"
            ]
            
            # Write headers
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = border
            
            # Write data
            for row_idx, record in enumerate(self.all_extracted_data, start=2):
                # NAMA BPR/S
                cell = ws.cell(row=row_idx, column=1, value=record.get('nama_bpr', ''))
                cell.border = border
                cell.alignment = left_align
                
                # JENIS
                cell = ws.cell(row=row_idx, column=2, value=record.get('jenis', ''))
                cell.border = border
                cell.alignment = center_align
                
                # KAB/KOTA
                cell = ws.cell(row=row_idx, column=3, value=record.get('kab_kota', ''))
                cell.border = border
                cell.alignment = left_align
                
                # PROVINSI
                cell = ws.cell(row=row_idx, column=4, value=record.get('provinsi', ''))
                cell.border = border
                cell.alignment = left_align
                
                # ASET (as text to preserve format)
                cell = ws.cell(row=row_idx, column=5, value=record.get('aset', ''))
                cell.border = border
                cell.alignment = right_align
                cell.number_format = '@'  # Text format
                
                # DANA PIHAK KETIGA (as text to preserve format)
                cell = ws.cell(row=row_idx, column=6, value=record.get('dana_pihak_ketiga', ''))
                cell.border = border
                cell.alignment = right_align
                cell.number_format = '@'  # Text format
                
                # KREDIT/PEMBIAYAAN (as text to preserve format)
                cell = ws.cell(row=row_idx, column=7, value=record.get('kredit_pembiayaan', ''))
                cell.border = border
                cell.alignment = right_align
                cell.number_format = '@'  # Text format
            
            # Auto-adjust column widths
            column_widths = {
                'A': 40,  # NAMA BPR/S
                'B': 15,  # JENIS
                'C': 20,  # KAB/KOTA
                'D': 20,  # PROVINSI
                'E': 20,  # ASET
                'F': 20,  # DANA PIHAK KETIGA
                'G': 25   # KREDIT/PEMBIAYAAN
            }
            for col_letter, width in column_widths.items():
                ws.column_dimensions[col_letter].width = width
            
            # Save workbook
            wb.save(filepath)
            self.logger.info(f"Saved content to Excel file: {filepath}")
            return filepath
            
        except Exception as e:
            self.logger.error(f"Error saving to Excel file: {e}")
            import traceback
            self.logger.debug(traceback.format_exc())
            return None
    
    def scrape_and_save(self, province_name: str = "Kep. Riau"):
        """
        Complete workflow: scrape all pages and save to both .txt and .xlsx files
        
        Args:
            province_name: Name of the province to search for
        """
        try:
            # Scrape all pages
            if not self.scrape_all_pages(province_name):
                self.logger.error("Failed to scrape pages")
                return False
            
            # Save to .txt
            txt_filepath = self.save_to_txt()
            
            # Save to .xlsx
            excel_filepath = self.save_to_excel()
            
            self.logger.info("=" * 70)
            self.logger.info("Scraping and saving completed!")
            if txt_filepath:
                self.logger.info(f"Text file: {txt_filepath}")
            if excel_filepath:
                self.logger.info(f"Excel file: {excel_filepath}")
            self.logger.info("=" * 70)
            
            return True
            
        except Exception as e:
            self.logger.error(f"Error during scrape and save: {e}")
            import traceback
            self.logger.error(traceback.format_exc())
            return False
        finally:
            # Cleanup
            self.cleanup()

