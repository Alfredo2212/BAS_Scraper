"""
Excel export functionality
Converts scraped data to Excel format
"""

from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from config.settings import Settings


class ExcelExporter:
    """Handles Excel file creation and data export"""
    
    @staticmethod
    def export_to_excel(data: list, headers: list = None, filename: str = None) -> Path:
        """
        Export data to Excel file
        
        Args:
            data: List of dictionaries containing the data
            headers: Optional list of headers. If None, uses keys from first data row
            filename: Optional filename. If None, generates automatic filename
            
        Returns:
            Path to the created Excel file
        """
        if not data:
            raise ValueError("No data to export")
        
        # Generate filename if not provided
        if filename is None:
            timestamp = datetime.now().strftime(Settings.EXCEL_FILENAME_DATE_FORMAT)
            filename = f"{Settings.EXCEL_FILENAME_PREFIX}_{timestamp}.xlsx"
        
        # Ensure .xlsx extension
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
        
        # Create full path
        filepath = Settings.OUTPUT_DIR / filename
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "OJK Report"
        
        # Determine headers
        if headers is None:
            headers = list(data[0].keys())
        
        # Write headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Write data
        for row_idx, row_data in enumerate(data, start=2):
            for col_idx, header in enumerate(headers, start=1):
                value = row_data.get(header, "")
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-adjust column widths
        for col_idx, header in enumerate(headers, start=1):
            column_letter = get_column_letter(col_idx)
            max_length = len(str(header))
            
            # Check data in column
            for row in ws[column_letter]:
                if row.value:
                    max_length = max(max_length, len(str(row.value)))
            
            # Set column width (with some padding)
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # Freeze header row
        ws.freeze_panes = "A2"
        
        # Save workbook
        wb.save(filepath)
        
        return filepath

