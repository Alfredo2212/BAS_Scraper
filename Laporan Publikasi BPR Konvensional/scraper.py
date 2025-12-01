"""
OJK ExtJS Scraper
Main scraping logic using ExtJS API exclusively
No DOM clicking, pure ExtJS ComponentQuery
"""

import time
import csv
import re
from pathlib import Path
from selenium.webdriver.remote.webdriver import WebDriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
except ImportError:
    print("[WARNING] openpyxl not installed. Excel export will not work. Install with: pip install openpyxl")
    Workbook = None

# Handle imports for both package and direct execution
try:
    from .helper import ExtJSHelper
    from .selenium_setup import SeleniumSetup
except ImportError:
    # If relative imports fail, try absolute imports
    import sys
    from pathlib import Path
    module_dir = Path(__file__).parent
    if str(module_dir) not in sys.path:
        sys.path.insert(0, str(module_dir))
    from helper import ExtJSHelper
    from selenium_setup import SeleniumSetup

from config.settings import OJKConfig, Settings


class OJKExtJSScraper:
    """Main scraper class using ExtJS API exclusively"""
    
    # Direct URL to BPR Konvensional report page
    REPORT_URL = "https://cfs.ojk.go.id/cfs/Report.aspx?BankTypeCode=BPK&BankTypeName=BPR%20Konvensional"
    
    def __init__(self, headless: bool = None):
        """
        Initialize the scraper
        
        Args:
            headless: Whether to run browser in headless mode
        """
        self.base_url = self.REPORT_URL  # Use direct report URL
        self.driver: WebDriver = None
        self.wait: WebDriverWait = None
        self.extjs: ExtJSHelper = None
        self.headless = headless if headless is not None else OJKConfig.HEADLESS_MODE
        self.output_dir = Path(Settings.OUTPUT_DIR)
        self.output_dir.mkdir(exist_ok=True)
        self.excel_wb = None  # Workbook for appending data
        self.excel_ws = None  # Worksheet for appending data
        self.excel_row = 1  # Current row in Excel
        self.all_data = []  # Store all extracted data for final Excel generation
        self.sheets_1_3_data = []  # Store data for Sheets 1-3 (ASET, Kredit, DPK)
        self.sheets_4_5_data = []  # Store data for Sheets 4-5 (Laba Kotor, Rasio)
    
    def initialize(self):
        """Initialize WebDriver and ExtJS helper"""
        if self.driver is None:
            self.driver = SeleniumSetup.create_driver(headless=self.headless)
            self.wait = SeleniumSetup.create_wait(self.driver)
            self.extjs = ExtJSHelper(self.driver, self.wait)
            
            # Minimize Chrome window (but keep it visible, not headless)
            if not self.headless:
                try:
                    self.driver.minimize_window()
                    print("[INFO] Chrome window minimized")
                except Exception as e:
                    print(f"[WARNING] Could not minimize Chrome window: {e}")
    
    def navigate_to_page(self):
        """Navigate directly to BPR Konvensional report page"""
        if self.driver is None:
            self.initialize()
        
        print(f"[INFO] Navigating directly to report page: {self.base_url}")
        self.driver.get(self.base_url)
        
        # Wait for page to load
        print("[INFO] Waiting for page to fully load...")
        time.sleep(0.75)
        
        # Check if page is in iframe or main page
        # Try to find ExtJS in main page first
        print("[INFO] Checking for ExtJS in main page...")
        max_attempts = 10
        for attempt in range(max_attempts):
            try:
                if self.extjs.check_extjs_available():
                    print("[OK] ExtJS is available in main page")
                    return
            except:
                pass
            time.sleep(0.75)
        
        # If not in main page, check for iframes
        print("[INFO] ExtJS not in main page, checking for iframes...")
        iframes = self.driver.find_elements(By.TAG_NAME, "iframe")
        if iframes:
            print(f"[INFO] Found {len(iframes)} iframe(s), checking inside...")
            for i, iframe in enumerate(iframes):
                try:
                    self.driver.switch_to.frame(iframe)
                    print(f"[INFO] Switched to iframe {i+1}")
                    time.sleep(1.125)  # Reduced from 2s to 0.4s (80% faster)
                    
                    # Check for ExtJS in this iframe
                    if self.extjs.check_extjs_available():
                        print(f"[OK] ExtJS is available in iframe {i+1}")
                        return
                    
                    # Switch back to try next iframe
                    self.driver.switch_to.default_content()
                except:
                    self.driver.switch_to.default_content()
                    continue
        
        # Wait a bit more and check again (reduced by 80%)
        print("[INFO] Waiting for ExtJS to load...")
        time.sleep(0.75)
        
        # Final check
        if self.extjs.check_extjs_available():
            print("[OK] ExtJS is now available")
            return
        
        # Debug: Check JavaScript context
        print("[WARNING] ExtJS not immediately available, checking JavaScript context...")
        try:
            debug_js = """
            (function() {
                try {
                    return {
                        hasExt: typeof Ext !== 'undefined',
                        hasComponentQuery: typeof Ext !== 'undefined' && typeof Ext.ComponentQuery !== 'undefined',
                        hasWindow: typeof window !== 'undefined',
                        documentReady: document.readyState,
                        hasJQuery: typeof jQuery !== 'undefined',
                        url: window.location.href
                    };
                } catch (e) {
                    return {error: e.toString()};
                }
            })();
            """
            debug_result = self.driver.execute_script(debug_js)
            print(f"[DEBUG] JavaScript context: {debug_result}")
        except Exception as debug_error:
            print(f"[DEBUG] Could not execute debug script: {debug_error}")
        
        print("[WARNING] ExtJS not available yet, but will continue (it may load after page fully loads)")
    
    def _get_target_month_year(self) -> tuple[str, str]:
        """
        Determine target month and year based on current date.
        Uses quarterly mapping logic:
        - Jan, Feb, Mar → use December (YYYY-1)
        - Apr, May, Jun → use March YYYY
        - Jul, Aug, Sep → use June YYYY
        - Oct, Nov, Dec → use September YYYY
        Where YYYY is the current server year.
        
        Returns:
            Tuple of (month_name, year) e.g., ("September", "2025")
        """
        from datetime import datetime
        
        now = datetime.now()
        current_month = now.month
        current_year = now.year
        
        # Quarterly mapping logic
        month_names = {
            3: "Maret",      # 03
            6: "Juni",       # 06
            9: "September",  # 09
            12: "Desember"   # 12
        }
        
        # Map current month to quarterly report month
        if current_month in [1, 2, 3]:  # Jan, Feb, Mar
            target_month_num = 12  # Desember
            target_year = current_year - 1
        elif current_month in [4, 5, 6]:  # Apr, May, Jun
            target_month_num = 3  # Maret
            target_year = current_year
        elif current_month in [7, 8, 9]:  # Jul, Aug, Sep
            target_month_num = 6  # Juni
            target_year = current_year
        else:  # Oct, Nov, Dec
            target_month_num = 9  # September
            target_year = current_year
        
        month_name = month_names[target_month_num]
        
        print(f"[INFO] Current date: {now.strftime('%B %Y')}")
        print(f"[INFO] Target month/year: {month_name} {target_year}")
        
        return month_name, str(target_year)
    
    def _check_and_handle_period_error(self, current_month: str, current_year: str) -> tuple[bool, str, str]:
        """
        Check for error by looking for span with id="ReportStatus". 
        Only detects errors - does NOT update selections (that should only happen during initial setup).
        
        Args:
            current_month: Current month being used (e.g., "September")
            current_year: Current year being used (e.g., "2025")
            
        Returns:
            Tuple of (error_found: bool, new_month: str, new_year: str)
            If error found, returns (True, new_month, new_year) with suggested period (but doesn't change selections)
            If no error, returns (False, current_month, current_year)
        """
        try:
            # Wait a bit for any error messages to appear
            time.sleep(1.5)
            
            # Check for span with id="ReportStatus"
            self.driver.switch_to.default_content()
            
            max_retries = 2
            new_month = current_month
            new_year = current_year
            
            for retry_attempt in range(max_retries + 1):
                try:
                    # Just check if ReportStatus span exists - if it exists, automatically skip
                    report_status_span = self.driver.find_element(By.ID, "ReportStatus")
                    
                    # ReportStatus span exists - this means there's an error
                    if retry_attempt == 0:
                        print(f"[WARNING] Found ReportStatus span, automatically skipping this period")
                        print(f"[INFO] Detected period error - suggested period update...")
                        
                        # Get the current available period using _get_target_month_year
                        new_month, new_year = self._get_target_month_year()
                        print(f"[INFO] Suggested period: {new_month} {new_year}")
                        print(f"[WARNING] Note: Period selection should be updated during initial setup, not during bank iterations")
                    
                    # Re-click Tampilkan button if we haven't exceeded max retries
                    if retry_attempt < max_retries:
                        print(f"[INFO] Re-clicking Tampilkan button (percobaan {retry_attempt + 1}/{max_retries})...")
                        try:
                            tampilkan_button = None
                            try:
                                tampilkan_button = self.driver.find_element(By.ID, "ShowReportButton-btnInnerEl")
                            except:
                                tampilkan_button = self.driver.find_element(By.XPATH, "//span[contains(text(), 'Tampilkan')]")
                            
                            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", tampilkan_button)
                            time.sleep(1.125)
                            self.driver.execute_script("arguments[0].click();", tampilkan_button)
                            print(f"[OK] Clicked 'Tampilkan' button again (retry {retry_attempt + 1})")
                            
                            # Wait a bit for the page to update
                            time.sleep(3.0)
                            
                            # Check again in next iteration
                            continue
                        except Exception as e:
                            print(f"[ERROR] Error re-clicking Tampilkan: {e}")
                            return True, new_month, new_year
                    else:
                        # Max retries reached, ReportStatus still exists
                        print(f"[WARNING] ReportStatus masih ada setelah {max_retries} percobaan, melewati periode ini")
                        return True, new_month, new_year
                except:
                    # ReportStatus span not found - no error, return success
                    if retry_attempt > 0:
                        print(f"[OK] ReportStatus tidak ditemukan setelah {retry_attempt} percobaan (error sudah hilang)")
                    return False, current_month, current_year
            
            # If we get here, max retries reached
            return True, new_month, new_year
            
        except Exception as e:
            print(f"[WARNING] Error checking for period error: {e}")
            return False, current_month, current_year
    
    def _setup_month_year_province(self, month: str, year: str):
        """
        Setup month, year, and province selections.
        This should only be called twice: once before sheets 1-3, and once before sheets 4-5.
        
        Args:
            month: Month to select (e.g., "September")
            year: Year to select (e.g., "2025")
        """
        print("\n[INFO] Setting up month, year, and province...")
        
        # Step 1: Select month
        print(f"[Step 1] Selecting month: {month}")
        self._select_month(month)
        time.sleep(0.75)
        
        # Step 2: Select year
        print(f"[Step 2] Selecting year: {year}")
        self._select_year(year)
        time.sleep(0.75)
        
        # Step 3: Select province
        print("[Step 3] Selecting province...")
        province_name = "Provinsi Kep. Riau"
        self._select_province(province_name)
        time.sleep(0.75)
        
        print("[OK] Month, year, and province setup completed")
    
    def _select_month(self, month: str):
        """Select month in the dropdown"""
        try:
            # Click month dropdown trigger
            trigger = self.driver.find_element(By.ID, "ext-gen1050")
            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", trigger)
            time.sleep(0.75)
            self.driver.execute_script("arguments[0].click();", trigger)
            time.sleep(0.75)
            
            # Wait for dropdown and find month
            from selenium.webdriver.support.ui import WebDriverWait
            from selenium.webdriver.support import expected_conditions as EC
            wait = WebDriverWait(self.driver, 5)
            wait.until(EC.presence_of_element_located((By.XPATH, "//ul[contains(@class, 'x-list-plain')]")))
            
            li_elements = self.driver.find_elements(By.XPATH, "//li[@role='option' or contains(@class, 'x-boundlist-item')]")
            if not li_elements:
                li_elements = self.driver.find_elements(By.XPATH, "//ul[contains(@class, 'x-list-plain')]//li")
            
            for li in li_elements:
                if li.text.strip().lower() == month.lower():
                    self.driver.execute_script("arguments[0].click();", li)
                    time.sleep(1.125)
                    print(f"[OK] Selected month: {month}")
                    return
        except Exception as e:
            print(f"[WARNING] Error selecting month {month}: {e}")
    
    def _select_year(self, year: str):
        """Select year in the input field"""
        try:
            year_input = self.driver.find_element(By.ID, "Year-inputEl")
            year_input.clear()
            year_input.send_keys(year)
            from selenium.webdriver.common.keys import Keys
            year_input.send_keys(Keys.TAB)
            time.sleep(0.75)
            print(f"[OK] Selected year: {year}")
        except Exception as e:
            print(f"[WARNING] Error selecting year {year}: {e}")
    
    def _select_province(self, province_name: str):
        """Select province in the dropdown"""
        try:
            # Try to find and click the trigger arrow with ID ext-gen1059 (static ID for province dropdown)
            print(f"  [INFO] Looking for province trigger arrow (id='ext-gen1059')...")
            province_trigger_found = False
            max_attempts = 10
            wait_interval = 0.75
            
            for attempt in range(max_attempts):
                try:
                    # Try to find by ID
                    province_trigger = self.driver.find_element(By.ID, "ext-gen1059")
                    print(f"  [OK] Found province trigger arrow (attempt {attempt + 1})")
                    
                    # Click the trigger to open dropdown
                    print("  [INFO] Clicking province trigger arrow to open dropdown...")
                    self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", province_trigger)
                    time.sleep(1.125)
                    self.driver.execute_script("arguments[0].click();", province_trigger)
                    print("  [OK] Province trigger arrow clicked")
                    province_trigger_found = True
                    break
                except:
                    if attempt < max_attempts - 1:
                        print(f"  [DEBUG] Province trigger not found, waiting {wait_interval} seconds...")
                        time.sleep(wait_interval)
                    else:
                        print("  [WARNING] Could not find province trigger arrow")
            
            # Wait for dropdown to appear and find <li> element
            if province_trigger_found:
                print("  [INFO] Waiting for province dropdown menu to appear...")
                time.sleep(1.125)
                
                print(f"  [INFO] Looking for <li> element with text '{province_name}'...")
                try:
                    from selenium.webdriver.support.ui import WebDriverWait
                    from selenium.webdriver.support import expected_conditions as EC
                    
                    wait = WebDriverWait(self.driver, 5)
                    dropdown_list = wait.until(
                        EC.presence_of_element_located((By.XPATH, "//ul[contains(@class, 'x-list-plain')]"))
                    )
                    print("  [OK] Province dropdown menu appeared")
                    
                    # Find all <li> elements
                    li_elements = self.driver.find_elements(By.XPATH, "//li[@role='option' or contains(@class, 'x-boundlist-item')]")
                    if not li_elements:
                        li_elements = self.driver.find_elements(By.XPATH, "//ul[contains(@class, 'x-list-plain')]//li")
                    
                    print(f"  [DEBUG] Found {len(li_elements)} <li> elements in province dropdown")
                    
                    # Find and click the matching <li>
                    target_li = None
                    for li in li_elements:
                        try:
                            li_text = li.text.strip()
                            if not li_text:  # Skip empty elements
                                continue
                            print(f"  [DEBUG] Checking <li>: '{li_text}'")
                            # Match if province name is contained in li_text or vice versa (but not empty)
                            if (province_name.lower() in li_text.lower() or li_text.lower() in province_name.lower()) and li_text:
                                target_li = li
                                print(f"  [OK] Found matching <li> element: '{li_text}'")
                                break
                        except:
                            continue
                    
                    if target_li:
                        print(f"  [INFO] Clicking <li> element with text '{province_name}'...")
                        self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", target_li)
                        time.sleep(1.125)
                        self.driver.execute_script("arguments[0].click();", target_li)
                        print(f"  [OK] Clicked <li> element with text '{province_name}'")
                        time.sleep(1.125)  # Wait for PostBack
                        print("  [OK] Selected province")
                        return
                    else:
                        available_options = [li.text.strip() for li in li_elements if li.text.strip()]
                        print(f"  [WARNING] Could not find <li> with text '{province_name}'. Available: {available_options[:10]}...")
                except Exception as e:
                    print(f"  [WARNING] Could not click province <li> element: {e}")
        except Exception as e:
            print(f"  [WARNING] Error selecting province: {e}")
    
    def scrape_all_data(self, month: str = None, year: str = None, phase: str = 'all'):
        """
        Main scraping loop
        Iterates through all provinces, cities, and banks
        
        Flow:
        - phase='all': Run all phases sequentially (original behavior)
        - phase='001': Run only Phase 1 (checkbox 001 → Sheets 1-3), then close Chrome and update Excel
        - phase='002': Run only Phase 2 (checkbox 002 → Sheet 4), then close Chrome and update Excel
        - phase='003': Run only Phase 3 (checkbox 003 → Sheet 5), then close Chrome and update Excel
        
        Args:
            month: Month to select (e.g., "Desember"). If None, auto-detects based on current date.
            year: Year to select (e.g., "2024"). If None, auto-detects based on current date.
            phase: Phase to run ('all', '001', '002', '003'). Default is 'all'.
        """
        # Auto-detect month and year if not provided
        if month is None or year is None:
            detected_month, detected_year = self._get_target_month_year()
            month = month or detected_month
            year = year or detected_year
        
        # Store current month/year for error handling
        self.current_month = month
        self.current_year = year
        
        # Initialize browser if needed
        if self.driver is None:
            self.initialize()
        
        # Wait for page to fully load
        print("[INFO] Waiting for page to fully load...")
        time.sleep(0.75)
        
        # Check for ExtJS availability
        print("[INFO] Checking for ExtJS availability...")
        max_attempts = 10
        for attempt in range(max_attempts):
            if self.extjs.check_extjs_available():
                print("[OK] ExtJS is available")
                break
            time.sleep(0.75)
        else:
            print("[WARNING] ExtJS not available, but will try to continue...")
        
        # Navigate to page if not already there
        if self.driver.current_url != self.base_url:
            self.navigate_to_page()
        
        # Store the year from input field for Excel labeling
        excel_year = year  # Default to provided year
        try:
            year_input = self.driver.find_element(By.ID, "Year-inputEl")
            excel_year = year_input.get_attribute('value') or year_input.get_property('value') or year
            if excel_year:
                print(f"[OK] Read year from input field for Excel: {excel_year}")
        except:
            pass
        
        # Phase 001: Sheets 1-3 (ASET, Kredit, DPK)
        if phase == '001' or phase == 'all':
            print("\n" + "="*60)
            print("[PHASE 001] Starting Phase 1: Sheets 1-3 (ASET, Kredit, DPK)")
            print("="*60)
            
            # Setup month, year, and province
            self._setup_month_year_province(month, year)
            
            # Select initial dropdowns and checkboxes (checkbox 001)
            print("\n[Step 4] Starting initial dropdown and checkbox selection...")
            print("[INFO] Setting up for Sheets 1-3 (ASET, Kredit, DPK)...")
            self._select_initial_dropdowns_and_checkboxes()
            
            # Wait a bit after checkbox is ticked to ensure dropdowns are ready
            print("\n[INFO] Waiting for dropdowns to be ready after checkbox selection...")
            time.sleep(1.5)
            
            # Initialize Excel file
            self._initialize_excel(year)
            # Initialize separate data lists for Sheets 1-3 and Sheets 4-5
            self.sheets_1_3_data = []
            self.sheets_4_5_data = []
            
            # Iterate through all cities and banks for Sheets 1-3
            print("\n[Step 5] Starting sequential iteration through all cities and all banks for Sheets 1-3...")
            print("[INFO] Iteration starts after checkbox is ticked - will select city, then bank, then click Tampilkan for each combination")
            
            city_index = 0
            while True:
                print(f"\n{'='*60}")
                print(f"[CITY] Processing city at index {city_index}...")
                print(f"{'='*60}")
                
                current_city = self._get_city_by_index(city_index)
                if not current_city:
                    print(f"  [INFO] No city found at index {city_index}. Reached end of cities list.")
                    break
                
                is_first_bank_in_city = True
                time.sleep(0.75)
                print(f"\n[INFO] Processing: {current_city}")
                bank_names = self._get_all_bank_names(city_index, city_already_selected=True)
                print(f"  [INFO] Found {len(bank_names)} banks in {current_city}")
                
                if not bank_names:
                    print(f"  [WARNING] No banks found on first attempt, waiting and retrying...")
                    time.sleep(1.125)
                    bank_names = self._get_all_bank_names(city_index, city_already_selected=True)
                    print(f"  [INFO] Found {len(bank_names)} banks in {current_city} (retry)")
                
                if not bank_names:
                    print(f"  [WARNING] No banks found in {current_city}, moving to next city...")
                    city_index += 1
                    continue
                
                for bank_index, current_bank in enumerate(bank_names):
                    print(f"\n  [BANK ({bank_index+1}/{len(bank_names)})] Processing: {current_bank}")
                    
                    max_retries = 3 if bank_index == 0 else 1
                    selected_bank_name = None
                    
                    for retry in range(max_retries):
                        selected_bank_name = self._select_bank_by_index(bank_index, city_index, city_already_selected=True)
                        if selected_bank_name:
                            break
                        elif retry < max_retries - 1:
                            print(f"  [WARNING] Could not select bank at index {bank_index}, retrying ({retry+1}/{max_retries})...")
                            time.sleep(1.125)
                    
                    if not selected_bank_name:
                        print(f"  [WARNING] Could not select bank at index {bank_index} after {max_retries} attempts")
                        continue
                    
                    time.sleep(1.125)
                    
                    print(f"  [INFO] Clicking Tampilkan and waiting for data extraction to complete...")
                    extracted_data = self._click_tampilkan_and_extract_data(year, current_city, selected_bank_name, extract_mode='sheets_1_3')
                    
                    if extracted_data:
                        print(f"  [INFO] Storing data...")
                        self._append_to_excel(extracted_data, year, current_city, selected_bank_name, is_first_bank_in_city, data_list='sheets_1_3')
                        print(f"  [OK] Data successfully extracted and saved to Excel for {current_city} - {selected_bank_name}")
                        is_first_bank_in_city = False
                        time.sleep(1.125)
                    else:
                        print(f"  [WARNING] Failed to extract data for {current_city} - {selected_bank_name}")
                    
                    if bank_index == len(bank_names) - 1:
                        print(f"  [INFO] This is the last bank ({bank_index+1}/{len(bank_names)}) in {current_city}")
                        time.sleep(1.125)
                
                print(f"  [INFO] Finished processing all {len(bank_names)} banks in {current_city}")
                print(f"  [INFO] Moving to next city...")
                time.sleep(0.75)
                city_index += 1
            
            # If phase is '001' only, close Chrome and update Excel
            if phase == '001':
                print("\n[OK] Phase 001 data collection completed!")
                print("[INFO] Closing browser...")
                self.cleanup()
                
                print("\n" + "="*60)
                print("[INFO] Updating Excel file with Phase 001 data...")
                print("="*60)
                self._finalize_excel(month, excel_year)
                print("\n[OK] Phase 001 completed successfully!")
                return
        
        # Phase 002: Sheet 4 (Laba Kotor)
        if phase == '002' or phase == 'all':
            if phase == '002':
                print("\n" + "="*60)
                print("[PHASE 002] Starting Phase 2: Sheet 4 (Laba Kotor)")
                print("="*60)
                
                # Setup month, year, and province
                self._setup_month_year_province(month, year)
                
                # Select checkbox 002 only
                print("\n[Step 4] Setting up checkbox 002 only (Laba Kotor)...")
                self._select_checkbox_002_only()
                
                # Wait a bit after checkbox is ticked
                print("\n[INFO] Waiting for dropdowns to be ready after checkbox selection...")
                time.sleep(1.5)
                
                # Initialize data storage
                self.sheets_4_5_data = []
            else:
                # phase == 'all': Continue from Phase 001
                print("\n" + "="*60)
                print("[INFO] Starting Sheet 4: Laba Kotor")
                print("="*60)
                
                # Verify browser session is still alive
                print("[INFO] Verifying browser session is still active...")
                try:
                    current_url = self.driver.current_url
                    print(f"[OK] Browser session is active (current URL: {current_url[:80]}...)")
                except Exception as e:
                    print(f"[ERROR] Browser session is invalid: {e}")
                    print("[ERROR] Cannot proceed with Laba Kotor/Rasio extraction - browser session ended")
                    print("[INFO] Finalizing Excel with collected data from Sheets 1-3...")
                    self._finalize_excel(month, excel_year)
                    self._finalize_excel_laba_kotor(month, excel_year)
                    self._finalize_excel_rasio(month, excel_year)
                    
                    # Retry banks with zero values using direct URL method
                    print("")
                    print("=" * 70)
                    print("Starting retry for banks with zero values...")
                    print("=" * 70)
                    self._retry_zero_value_banks(month, excel_year)
                    
                    print("[INFO] Closing browser...")
                    self.cleanup()
                    return
                
                # Refresh the page to get a clean state
                print("\n[INFO] Refreshing page to get clean state for Sheets 4-5...")
                self.driver.refresh()
                time.sleep(3.0)
                
                print("[INFO] Waiting for page to fully load after refresh...")
                time.sleep(0.75)
                
                # Re-do the full setup: month, year, province, and checkboxes (002 and 003)
                print("\n[INFO] Re-setting up month, year, province, and checkboxes for Sheets 4-5...")
                self._setup_for_sheets_4_5(month, year)
                
                # Wait a bit after checkbox is ticked
                print("\n[INFO] Waiting for dropdowns to be ready after checkbox selection...")
                time.sleep(1.5)
                
                # Clear data storage for Laba Kotor
                self.sheets_4_5_data = []
            
            # Iterate through all cities and banks for Laba Kotor
            print("\n[INFO] Starting iteration through all cities for Laba Kotor...")
            
            city_index = 0
            while True:
                print(f"\n{'='*60}")
                print(f"[CITY] Processing city at index {city_index}...")
                print(f"{'='*60}")
                
                current_city = self._get_city_by_index(city_index)
                if not current_city:
                    print(f"  [INFO] No city found at index {city_index}. Reached end of cities list.")
                    break
                
                is_first_bank_in_city = True
                time.sleep(0.75)
                print(f"\n[INFO] Processing: {current_city}")
                bank_names = self._get_all_bank_names(city_index, city_already_selected=True)
                print(f"  [INFO] Found {len(bank_names)} banks in {current_city}")
                
                if not bank_names:
                    print(f"  [WARNING] No banks found on first attempt, waiting and retrying...")
                    time.sleep(1.125)
                    bank_names = self._get_all_bank_names(city_index, city_already_selected=True)
                    print(f"  [INFO] Found {len(bank_names)} banks in {current_city} (retry)")
                
                if not bank_names:
                    print(f"  [WARNING] No banks found in {current_city}, moving to next city...")
                    city_index += 1
                    continue
                
                for bank_index, current_bank in enumerate(bank_names):
                    print(f"\n  [BANK ({bank_index+1}/{len(bank_names)})] Processing: {current_bank}")
                    
                    max_retries = 3 if bank_index == 0 else 1
                    selected_bank_name = None
                    
                    for retry in range(max_retries):
                        selected_bank_name = self._select_bank_by_index(bank_index, city_index, city_already_selected=True)
                        if selected_bank_name:
                            break
                        elif retry < max_retries - 1:
                            print(f"  [WARNING] Could not select bank at index {bank_index}, retrying ({retry+1}/{max_retries})...")
                            time.sleep(1.125)
                    
                    if not selected_bank_name:
                        print(f"  [WARNING] Could not select bank at index {bank_index} after {max_retries} attempts")
                        continue
                    
                    time.sleep(1.125)
                    
                    print(f"  [INFO] Clicking Tampilkan and waiting for data extraction to complete...")
                    # For phase 002, extract only Laba Kotor (skip Rasio)
                    extracted_data = self._click_tampilkan_and_extract_data(year, current_city, selected_bank_name, extract_mode='sheets_4_5', skip_rasio=True)
                    
                    if extracted_data:
                        print(f"  [INFO] Storing data...")
                        self._append_to_excel(extracted_data, year, current_city, selected_bank_name, is_first_bank_in_city, data_list='sheets_4_5')
                        print(f"  [OK] Data successfully extracted and saved to Excel for {current_city} - {selected_bank_name}")
                        is_first_bank_in_city = False
                        time.sleep(1.125)
                    else:
                        print(f"  [WARNING] Failed to extract data for {current_city} - {selected_bank_name}")
                    
                    if bank_index == len(bank_names) - 1:
                        print(f"  [INFO] This is the last bank ({bank_index+1}/{len(bank_names)}) in {current_city}")
                        time.sleep(1.125)
                
                print(f"  [INFO] Finished processing all {len(bank_names)} banks in {current_city}")
                print(f"  [INFO] Moving to next city...")
                time.sleep(0.75)
                city_index += 1
            
            # If phase is '002' only, close Chrome and update Excel
            if phase == '002':
                print("\n[OK] Phase 002 data collection completed!")
                print("[INFO] Closing browser...")
                self.cleanup()
                
                print("\n" + "="*60)
                print("[INFO] Updating Excel file with Phase 002 data...")
                print("="*60)
                self._finalize_excel_laba_kotor(month, excel_year)
                print("\n[OK] Phase 002 completed successfully!")
                return
        
        # Phase 003: Sheet 5 (Rasio)
        if phase == '003' or phase == 'all':
            if phase == '003':
                print("\n" + "="*60)
                print("[PHASE 003] Starting Phase 3: Sheet 5 (Rasio)")
                print("="*60)
                
                # Setup month, year, and province
                self._setup_month_year_province(month, year)
                
                # Select checkbox 003 only
                print("\n[Step 4] Setting up checkbox 003 only (Rasio)...")
                self._select_checkbox_003_only()
                
                # Wait a bit after checkbox is ticked
                print("\n[INFO] Waiting for dropdowns to be ready after checkbox selection...")
                time.sleep(1.5)
                
                # Initialize data storage
                self.sheets_4_5_data = []
            # else: phase == 'all' continues from Phase 002 (no refresh needed, just continue)
            
            # Iterate through all cities and banks for Rasio
            print("\n[INFO] Starting iteration through all cities for Rasio...")
            
            city_index = 0
            while True:
                print(f"\n{'='*60}")
                print(f"[CITY] Processing city at index {city_index}...")
                print(f"{'='*60}")
                
                current_city = self._get_city_by_index(city_index)
                if not current_city:
                    print(f"  [INFO] No city found at index {city_index}. Reached end of cities list.")
                    break
                
                is_first_bank_in_city = True
                time.sleep(0.75)
                print(f"\n[INFO] Processing: {current_city}")
                bank_names = self._get_all_bank_names(city_index, city_already_selected=True)
                print(f"  [INFO] Found {len(bank_names)} banks in {current_city}")
                
                if not bank_names:
                    print(f"  [WARNING] No banks found on first attempt, waiting and retrying...")
                    time.sleep(1.125)
                    bank_names = self._get_all_bank_names(city_index, city_already_selected=True)
                    print(f"  [INFO] Found {len(bank_names)} banks in {current_city} (retry)")
                
                if not bank_names:
                    print(f"  [WARNING] No banks found in {current_city}, moving to next city...")
                    city_index += 1
                    continue
                
                for bank_index, current_bank in enumerate(bank_names):
                    print(f"\n  [BANK ({bank_index+1}/{len(bank_names)})] Processing: {current_bank}")
                    
                    max_retries = 3 if bank_index == 0 else 1
                    selected_bank_name = None
                    
                    for retry in range(max_retries):
                        selected_bank_name = self._select_bank_by_index(bank_index, city_index, city_already_selected=True)
                        if selected_bank_name:
                            break
                        elif retry < max_retries - 1:
                            print(f"  [WARNING] Could not select bank at index {bank_index}, retrying ({retry+1}/{max_retries})...")
                            time.sleep(1.125)
                    
                    if not selected_bank_name:
                        print(f"  [WARNING] Could not select bank at index {bank_index} after {max_retries} attempts")
                        continue
                    
                    time.sleep(1.125)
                    
                    print(f"  [INFO] Clicking Tampilkan and waiting for data extraction to complete...")
                    # For phase 003, extract only Rasio (skip Laba Kotor)
                    extracted_data = self._click_tampilkan_and_extract_data(year, current_city, selected_bank_name, extract_mode='sheets_4_5', skip_laba_kotor=True)
                    
                    if extracted_data:
                        print(f"  [INFO] Storing data...")
                        self._append_to_excel(extracted_data, year, current_city, selected_bank_name, is_first_bank_in_city, data_list='sheets_4_5')
                        print(f"  [OK] Data successfully extracted and saved to Excel for {current_city} - {selected_bank_name}")
                        is_first_bank_in_city = False
                        time.sleep(1.125)
                    else:
                        print(f"  [WARNING] Failed to extract data for {current_city} - {selected_bank_name}")
                    
                    if bank_index == len(bank_names) - 1:
                        print(f"  [INFO] This is the last bank ({bank_index+1}/{len(bank_names)}) in {current_city}")
                        time.sleep(1.125)
                
                print(f"  [INFO] Finished processing all {len(bank_names)} banks in {current_city}")
                print(f"  [INFO] Moving to next city...")
                time.sleep(0.75)
                city_index += 1
            
            # If phase is '003' only, close Chrome and update Excel
            if phase == '003':
                print("\n[OK] Phase 003 data collection completed!")
                print("[INFO] Closing browser...")
                self.cleanup()
                
                print("\n" + "="*60)
                print("[INFO] Updating Excel file with Phase 003 data...")
                print("="*60)
                print("[INFO] Starting Sheet 5: Rasio (9 tables)")
                self._finalize_excel_rasio(month, excel_year)
                
                # Note: Retry for zero values only runs after phase='all' completes
                # (not for individual phases)
                
                print("\n[OK] Phase 003 completed successfully!")
                return
        
        # phase == 'all': Finalize all Excel sheets
        print("\n[OK] All data collection completed!")
        print("[INFO] Closing browser...")
        self.cleanup()
        
        print("\n" + "="*60)
        print("[INFO] Creating Excel file with all sheets...")
        print("="*60)
        
        self._finalize_excel(month, excel_year)
        self._finalize_excel_laba_kotor(month, excel_year)
        print("\n" + "="*60)
        print("[INFO] Starting Sheet 5: Rasio (9 tables)")
        print("="*60)
        self._finalize_excel_rasio(month, excel_year)
        
        # Retry banks with zero values using direct URL method
        print("")
        print("=" * 70)
        print("Starting retry for banks with zero values...")
        print("=" * 70)
        self._retry_zero_value_banks(month, excel_year)
        
        print("\n[OK] Excel file created successfully!")
    
    def run_all_phases(self, month: str = None, year: str = None):
        """
        Run all 3 phases sequentially, each with its own Chrome session.
        This is the main entry point for the 3-phase scraping approach.
        
        Args:
            month: Month to select (e.g., "Desember"). If None, auto-detects based on current date.
            year: Year to select (e.g., "2024"). If None, auto-detects based on current date.
        """
        # Auto-detect month and year if not provided
        if month is None or year is None:
            detected_month, detected_year = self._get_target_month_year()
            month = month or detected_month
            year = year or detected_year
        
        print("\n" + "="*60)
        print("[ORCHESTRATOR] Starting 3-phase scraping process")
        print(f"[ORCHESTRATOR] Target: {month} {year}")
        print("="*60)
        
        # Phase 001: Sheets 1-3 (ASET, Kredit, DPK)
        print("\n" + "="*60)
        print("[ORCHESTRATOR] Phase 001: Starting...")
        print("="*60)
        self.scrape_all_data(month=month, year=year, phase='001')
        print("\n[ORCHESTRATOR] Phase 001: Completed")
        
        # Phase 002: Sheet 4 (Laba Kotor)
        print("\n" + "="*60)
        print("[ORCHESTRATOR] Phase 002: Starting...")
        print("="*60)
        self.scrape_all_data(month=month, year=year, phase='002')
        print("\n[ORCHESTRATOR] Phase 002: Completed")
        
        # Phase 003: Sheet 5 (Rasio)
        print("\n" + "="*60)
        print("[ORCHESTRATOR] Phase 003: Starting...")
        print("="*60)
        self.scrape_all_data(month=month, year=year, phase='003')
        print("\n[ORCHESTRATOR] Phase 003: Completed")
        
        print("\n" + "="*60)
        print("[ORCHESTRATOR] All phases completed successfully!")
        print("="*60)
    
    def _select_initial_dropdowns_and_checkboxes(self):
        """
        3-step sequential process:
        1. Click dropdown arrow ext-gen1064 and select topmost <li>
        2. Click dropdown arrow ext-gen1069 and select topmost <li>
        3. Find treeview elements and check checkboxes inside them
        4. Click Tampilkan button and extract data
        
        Args:
            year: Selected year for data extraction
        """
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        
        # Step 1: Skip dropdown selections (handled in main loop)
        print("\n  [Step 4.1] Skipping city dropdown selection (handled in main loop)...")
        time.sleep(0.3)
        
        # Step 2: Skip bank dropdown selection (handled in main loop)
        print("\n  [Step 4.2] Skipping bank dropdown selection (handled in main loop)...")
        time.sleep(0.3)
        
        # Step 3: Find treeview element and check only the first checkbox
        # All three data types (Kredit, Total Aset, DPK) use the same checkbox
        print("\n  [Step 4.3] Finding treeview element and checking checkbox...")
        time.sleep(0.75)  # Wait a bit longer for treeview to be ready
        
        treeview_id = "treeview-1012-record-BPK-901-000001"
        
        try:
            print(f"    [INFO] Looking for treeview element: {treeview_id}")
            
            # Wait for treeview element to be present
            wait = WebDriverWait(self.driver, 10)
            treeview_element = wait.until(
                EC.presence_of_element_located((By.ID, treeview_id))
            )
            print(f"    [OK] Found treeview element: {treeview_id}")
            
            # Find nested divs with role="checkbox" inside this treeview element
            # Try multiple XPath patterns to find checkboxes
            checkboxes = []
            
            # Pattern 1: div with role="checkbox"
            checkboxes = treeview_element.find_elements(By.XPATH, ".//div[@role='checkbox']")
            if checkboxes:
                print(f"    [DEBUG] Found {len(checkboxes)} checkbox(es) using pattern: div[@role='checkbox']")
            else:
                # Pattern 2: input type="checkbox"
                checkboxes = treeview_element.find_elements(By.XPATH, ".//input[@type='checkbox']")
                if checkboxes:
                    print(f"    [DEBUG] Found {len(checkboxes)} checkbox(es) using pattern: input[@type='checkbox']")
                else:
                    # Pattern 3: elements with checkbox in class
                    checkboxes = treeview_element.find_elements(By.XPATH, ".//div[contains(@class, 'checkbox')]")
                    if checkboxes:
                        print(f"    [DEBUG] Found {len(checkboxes)} checkbox(es) using pattern: div[contains(@class, 'checkbox')]")
                    else:
                        # Pattern 4: any element with aria-checked attribute
                        checkboxes = treeview_element.find_elements(By.XPATH, ".//*[@aria-checked]")
                        if checkboxes:
                            print(f"    [DEBUG] Found {len(checkboxes)} checkbox(es) using pattern: *[@aria-checked]")
                        else:
                            # Pattern 5: look for elements with checkbox-related attributes
                            checkboxes = treeview_element.find_elements(By.XPATH, ".//*[contains(@class, 'x-tree-checkbox') or contains(@class, 'tree-checkbox')]")
                            if checkboxes:
                                print(f"    [DEBUG] Found {len(checkboxes)} checkbox(es) using pattern: *[contains(@class, 'tree-checkbox')]")
                            else:
                                # Pattern 6: look for any clickable element that might be a checkbox
                                checkboxes = treeview_element.find_elements(By.XPATH, ".//*[@role='checkbox' or @type='checkbox' or contains(@class, 'checkbox')]")
                                if checkboxes:
                                    print(f"    [DEBUG] Found {len(checkboxes)} checkbox(es) using pattern: *[@role='checkbox' or @type='checkbox' or contains(@class, 'checkbox')]")
            
            if checkboxes:
                print(f"    [OK] Found {len(checkboxes)} checkbox(es) in {treeview_id}")
                # Only check the first checkbox (all three data types use the same checkbox)
                checkbox = checkboxes[0]
                try:
                    # Check if element is visible
                    if not checkbox.is_displayed():
                        print(f"    [DEBUG] Checkbox is not visible, skipping")
                    else:
                        # Get checkbox attributes for debugging
                        aria_checked = checkbox.get_attribute("aria-checked")
                        checkbox_type = checkbox.get_attribute("type")
                        
                        # Check if checkbox is already checked
                        if aria_checked == "true" or (checkbox_type == "checkbox" and checkbox.is_selected()):
                            print(f"    [INFO] Checkbox already checked in {treeview_id}")
                        else:
                            # Scroll to checkbox and click
                            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center', behavior: 'smooth'});", checkbox)
                            time.sleep(1.125)  # Wait for scroll
                            
                            # Try clicking with JavaScript first
                            try:
                                self.driver.execute_script("arguments[0].click();", checkbox)
                                print(f"    [OK] Checked checkbox in {treeview_id} (JavaScript click)")
                            except:
                                # Fallback to regular click
                                checkbox.click()
                                print(f"    [OK] Checked checkbox in {treeview_id} (regular click)")
                            
                            time.sleep(1.125)  # Wait after click
                except Exception as e:
                    print(f"    [WARNING] Could not check checkbox in {treeview_id}: {e}")
            else:
                print(f"    [WARNING] No checkboxes found in {treeview_id}")
        except Exception as e:
            print(f"    [WARNING] Could not find treeview element {treeview_id}: {e}")
        
        print("  [OK] Completed initial setup (dropdowns and checkbox)")
    
    def _setup_for_sheets_4_5(self, month: str, year: str):
        """
        Full setup for Sheets 4-5: Select month, year, province, and checkboxes (002 and 003).
        This is called after refreshing the page to get a clean state.
        
        Args:
            month: Month to select (e.g., "September")
            year: Year to select (e.g., "2025")
        """
        # Setup month, year, and province (second time - after page refresh for sheets 4-5)
        self._setup_month_year_province(month, year)
        
        # Step 4: Select checkboxes (002 and 003) - uncheck 001 first, then check 002 and 003
        print("\n  [Step 4] Setting up checkboxes for Sheets 4-5 (002 and 003)...")
        self._change_checkboxes_for_laba_kotor()
        
        print("  [OK] Completed setup for Sheets 4-5 (month, year, province, checkboxes)")
    
    def _change_checkboxes_for_laba_kotor(self):
        """
        Change checkboxes: uncheck treeview-1012-record-BPK-901-000001, check the two new ones.
        
        NOTE: This function ONLY changes checkboxes. It does NOT touch month/year selection.
        Month and year should remain set from the initial selection at the start of scrape_all_data.
        """
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        
        print("\n[INFO] Changing checkboxes for Laba Kotor...")
        print("[INFO] NOTE: Month and year remain unchanged - only checkboxes are being modified")
        
        # Uncheck first checkbox
        treeview_id_uncheck = "treeview-1012-record-BPK-901-000001"
        print(f"  [INFO] Unchecking: {treeview_id_uncheck}")
        try:
            wait = WebDriverWait(self.driver, 10)
            treeview_element = wait.until(
                EC.presence_of_element_located((By.ID, treeview_id_uncheck))
            )
            
            # Find checkboxes
            checkboxes = treeview_element.find_elements(By.XPATH, ".//*[contains(@class, 'x-tree-checkbox') or contains(@class, 'tree-checkbox') or @role='checkbox' or @type='checkbox']")
            if not checkboxes:
                checkboxes = treeview_element.find_elements(By.XPATH, ".//*[@aria-checked]")
            
            if checkboxes:
                checkbox = checkboxes[0]
                aria_checked = checkbox.get_attribute("aria-checked")
                if aria_checked == "true":
                    self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", checkbox)
                    time.sleep(1.125)
                    self.driver.execute_script("arguments[0].click();", checkbox)
                    print(f"  [OK] Unchecked: {treeview_id_uncheck}")
                    time.sleep(1.125)
        except Exception as e:
            print(f"  [WARNING] Could not uncheck {treeview_id_uncheck}: {e}")
        
        # Check the two new checkboxes
        treeview_ids_check = [
            "treeview-1012-record-BPK-901-000002",
            "treeview-1012-record-BPK-901-000003"
        ]
        
        for treeview_id in treeview_ids_check:
            print(f"  [INFO] Checking: {treeview_id}")
            try:
                wait = WebDriverWait(self.driver, 10)
                treeview_element = wait.until(
                    EC.presence_of_element_located((By.ID, treeview_id))
                )
                
                # Find checkboxes
                checkboxes = treeview_element.find_elements(By.XPATH, ".//*[contains(@class, 'x-tree-checkbox') or contains(@class, 'tree-checkbox') or @role='checkbox' or @type='checkbox']")
                if not checkboxes:
                    checkboxes = treeview_element.find_elements(By.XPATH, ".//*[@aria-checked]")
                
                if checkboxes:
                    checkbox = checkboxes[0]
                    aria_checked = checkbox.get_attribute("aria-checked")
                    if aria_checked != "true":
                        self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", checkbox)
                        time.sleep(1.125)
                        self.driver.execute_script("arguments[0].click();", checkbox)
                        print(f"  [OK] Checked: {treeview_id}")
                        time.sleep(1.125)
                    else:
                        print(f"  [INFO] Already checked: {treeview_id}")
            except Exception as e:
                print(f"  [WARNING] Could not check {treeview_id}: {e}")
        
        print("  [OK] Checkbox changes completed")
    
    def _select_checkbox_002_only(self):
        """
        Select only checkbox 002 (Laba Kotor).
        Uncheck checkboxes 001 and 003 to ensure only 002 is checked.
        
        NOTE: This function ONLY changes checkboxes. It does NOT touch month/year selection.
        """
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        
        print("\n[INFO] Setting up checkbox 002 only (Laba Kotor)...")
        print("[INFO] NOTE: Month and year remain unchanged - only checkboxes are being modified")
        
        # List of all checkboxes to manage
        treeview_ids = {
            "001": "treeview-1012-record-BPK-901-000001",
            "002": "treeview-1012-record-BPK-901-000002",
            "003": "treeview-1012-record-BPK-901-000003"
        }
        
        # Uncheck 001 and 003, ensure 002 is checked
        for checkbox_id, treeview_id in treeview_ids.items():
            try:
                wait = WebDriverWait(self.driver, 10)
                treeview_element = wait.until(
                    EC.presence_of_element_located((By.ID, treeview_id))
                )
                
                # Find checkboxes
                checkboxes = treeview_element.find_elements(By.XPATH, ".//*[contains(@class, 'x-tree-checkbox') or contains(@class, 'tree-checkbox') or @role='checkbox' or @type='checkbox']")
                if not checkboxes:
                    checkboxes = treeview_element.find_elements(By.XPATH, ".//*[@aria-checked]")
                
                if checkboxes:
                    checkbox = checkboxes[0]
                    aria_checked = checkbox.get_attribute("aria-checked")
                    
                    if checkbox_id == "002":
                        # Ensure 002 is checked
                        if aria_checked != "true":
                            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", checkbox)
                            time.sleep(1.125)
                            self.driver.execute_script("arguments[0].click();", checkbox)
                            print(f"  [OK] Checked: {treeview_id}")
                            time.sleep(1.125)
                        else:
                            print(f"  [INFO] Already checked: {treeview_id}")
                    else:
                        # Uncheck 001 and 003
                        if aria_checked == "true":
                            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", checkbox)
                            time.sleep(1.125)
                            self.driver.execute_script("arguments[0].click();", checkbox)
                            print(f"  [OK] Unchecked: {treeview_id}")
                            time.sleep(1.125)
                        else:
                            print(f"  [INFO] Already unchecked: {treeview_id}")
            except Exception as e:
                print(f"  [WARNING] Could not manage {treeview_id}: {e}")
        
        print("  [OK] Checkbox 002 only setup completed")
    
    def _select_checkbox_003_only(self):
        """
        Select only checkbox 003 (Rasio).
        Uncheck checkboxes 001 and 002 to ensure only 003 is checked.
        
        NOTE: This function ONLY changes checkboxes. It does NOT touch month/year selection.
        """
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        
        print("\n[INFO] Setting up checkbox 003 only (Rasio)...")
        print("[INFO] NOTE: Month and year remain unchanged - only checkboxes are being modified")
        
        # List of all checkboxes to manage
        treeview_ids = {
            "001": "treeview-1012-record-BPK-901-000001",
            "002": "treeview-1012-record-BPK-901-000002",
            "003": "treeview-1012-record-BPK-901-000003"
        }
        
        # Uncheck 001 and 002, ensure 003 is checked
        for checkbox_id, treeview_id in treeview_ids.items():
            try:
                wait = WebDriverWait(self.driver, 10)
                treeview_element = wait.until(
                    EC.presence_of_element_located((By.ID, treeview_id))
                )
                
                # Find checkboxes
                checkboxes = treeview_element.find_elements(By.XPATH, ".//*[contains(@class, 'x-tree-checkbox') or contains(@class, 'tree-checkbox') or @role='checkbox' or @type='checkbox']")
                if not checkboxes:
                    checkboxes = treeview_element.find_elements(By.XPATH, ".//*[@aria-checked]")
                
                if checkboxes:
                    checkbox = checkboxes[0]
                    aria_checked = checkbox.get_attribute("aria-checked")
                    
                    if checkbox_id == "003":
                        # Ensure 003 is checked
                        if aria_checked != "true":
                            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", checkbox)
                            time.sleep(1.125)
                            self.driver.execute_script("arguments[0].click();", checkbox)
                            print(f"  [OK] Checked: {treeview_id}")
                            time.sleep(1.125)
                        else:
                            print(f"  [INFO] Already checked: {treeview_id}")
                    else:
                        # Uncheck 001 and 002
                        if aria_checked == "true":
                            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", checkbox)
                            time.sleep(1.125)
                            self.driver.execute_script("arguments[0].click();", checkbox)
                            print(f"  [OK] Unchecked: {treeview_id}")
                            time.sleep(1.125)
                        else:
                            print(f"  [INFO] Already unchecked: {treeview_id}")
            except Exception as e:
                print(f"  [WARNING] Could not manage {treeview_id}: {e}")
        
        print("  [OK] Checkbox 003 only setup completed")
    
    def _get_city_by_index(self, index: int) -> str:
        """Get city name by index from dropdown ext-gen1064. Returns city name or None if not found."""
        try:
            print(f"    [DEBUG] Attempting to get city at index {index}...")
            
            # First, close any open dropdowns to avoid confusion
            try:
                from selenium.webdriver.common.keys import Keys
                self.driver.switch_to.default_content()
                body = self.driver.find_element(By.TAG_NAME, "body")
                body.send_keys(Keys.ESCAPE)
                time.sleep(1.125)  # MAX(0.5, 50% of 0.3) = 0.5
            except:
                pass
            
            # Click dropdown trigger to open
            dropdown_trigger = self.driver.find_element(By.ID, "ext-gen1064")
            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", dropdown_trigger)
            time.sleep(0.75)  # MAX(0.5, 50% of 0.5) = 0.5
            self.driver.execute_script("arguments[0].click();", dropdown_trigger)
            print(f"    [DEBUG] Clicked city dropdown trigger")
            
            # Wait for dropdown to appear - find the specific dropdown boundlist for city
            time.sleep(0.75)  # MAX(0.5, 50% of 0.5) = 0.5
            wait = WebDriverWait(self.driver, 10)
            # Wait for the boundlist to appear (city dropdown should have a specific boundlist)
            wait.until(EC.presence_of_element_located((By.XPATH, "//div[contains(@class, 'x-boundlist') and contains(@class, 'x-boundlist-floating')]")))
            print(f"    [DEBUG] City dropdown appeared")
            
            # Find all visible boundlists and use the one that's visible (not hidden)
            # Get all boundlists and filter for visible ones
            all_boundlists = self.driver.find_elements(By.XPATH, "//div[contains(@class, 'x-boundlist') and contains(@class, 'x-boundlist-floating')]")
            boundlist = None
            for bl in all_boundlists:
                try:
                    # Check if boundlist is visible
                    if bl.is_displayed():
                        boundlist = bl
                        break
                except:
                    continue
            
            # If no visible boundlist found, use the first one
            if boundlist is None and all_boundlists:
                boundlist = all_boundlists[0]
            
            # Get all city options from within this specific boundlist only
            if boundlist:
                li_elements = boundlist.find_elements(By.XPATH, ".//li[@role='option' or contains(@class, 'x-boundlist-item')]")
                if not li_elements:
                    li_elements = boundlist.find_elements(By.XPATH, ".//ul[contains(@class, 'x-list-plain')]//li")
            else:
                # Fallback to global search if boundlist not found
                li_elements = self.driver.find_elements(By.XPATH, "//li[@role='option' or contains(@class, 'x-boundlist-item')]")
                if not li_elements:
                    li_elements = self.driver.find_elements(By.XPATH, "//ul[contains(@class, 'x-list-plain')]//li")
            
            print(f"    [DEBUG] Found {len(li_elements)} <li> elements in city dropdown")
            
            # Retry mechanism with JavaScript fallback to get text content
            # Sometimes .text is empty even though elements exist, so we use JavaScript to get textContent/innerText
            max_retries = 5
            valid_cities = []
            
            for retry in range(max_retries):
                valid_cities = []
                for li in li_elements:
                    # Try to get text using Selenium's .text property
                    text = li.text.strip()
                    if not text:
                        # If .text is empty, try JavaScript to get textContent or innerText
                        try:
                            text = self.driver.execute_script(
                                "return arguments[0].textContent || arguments[0].innerText || '';", li
                            ).strip()
                        except:
                            text = ""
                    
                    if text:
                        valid_cities.append((li, text))
                
                if valid_cities:
                    break  # Found valid cities, exit retry loop
                
                if retry < max_retries - 1:
                    print(f"    [DEBUG] No valid cities found (retry {retry + 1}/{max_retries}), waiting and retrying...")
                    time.sleep(1.125 if retry == 0 else 0.75)  # MAX(0.5, 50% of 1.5) = 0.75, MAX(0.5, 50% of 1.0) = 0.5
                    # Re-fetch li_elements from the city dropdown boundlist
                    try:
                        all_boundlists = self.driver.find_elements(By.XPATH, "//div[contains(@class, 'x-boundlist') and contains(@class, 'x-boundlist-floating')]")
                        boundlist = None
                        for bl in all_boundlists:
                            try:
                                if bl.is_displayed():
                                    boundlist = bl
                                    break
                            except:
                                continue
                        
                        if boundlist:
                            li_elements = boundlist.find_elements(By.XPATH, ".//li[@role='option' or contains(@class, 'x-boundlist-item')]")
                            if not li_elements:
                                li_elements = boundlist.find_elements(By.XPATH, ".//ul[contains(@class, 'x-list-plain')]//li")
                        else:
                            # Fallback to global search if boundlist not found
                            li_elements = self.driver.find_elements(By.XPATH, "//li[@role='option' or contains(@class, 'x-boundlist-item')]")
                            if not li_elements:
                                li_elements = self.driver.find_elements(By.XPATH, "//ul[contains(@class, 'x-list-plain')]//li")
                    except:
                        # Fallback to global search if boundlist not found
                        li_elements = self.driver.find_elements(By.XPATH, "//li[@role='option' or contains(@class, 'x-boundlist-item')]")
                        if not li_elements:
                            li_elements = self.driver.find_elements(By.XPATH, "//ul[contains(@class, 'x-list-plain')]//li")
            
            print(f"    [DEBUG] Found {len(valid_cities)} valid cities (non-empty)")
            
            if valid_cities:
                for i, (city_li, city_text) in enumerate(valid_cities[:5]):  # Print first 5 for debugging
                    print(f"    [DEBUG] City {i}: '{city_text[:50]}...'")
            
            if index < len(valid_cities):
                city_li, city_name = valid_cities[index]
                print(f"    [DEBUG] Selecting city at index {index}: '{city_name}'")
                # Select it
                self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", city_li)
                time.sleep(1.125)  # MAX(0.5, 50% of 0.5) = 0.5
                self.driver.execute_script("arguments[0].click();", city_li)
                time.sleep(1.125)  # MAX(0.5, 50% of 1.0) = 0.5 - Wait for PostBack and dropdown to update
                print(f"    [OK] Selected city: '{city_name}'")
                return city_name
            else:
                print(f"    [WARNING] Index {index} is out of range. Total valid cities: {len(valid_cities)}")
                # Close dropdown
                try:
                    from selenium.webdriver.common.keys import Keys
                    dropdown_trigger.send_keys(Keys.ESCAPE)
                except:
                    pass
                return None
        except Exception as e:
            print(f"    [ERROR] Could not get city by index {index}: {e}")
            return None
    
    def _get_all_bank_names(self, city_index: int, city_already_selected: bool = False) -> list:
        """Get all bank names from dropdown ext-gen1069. Returns list of bank names.
        
        Args:
            city_index: Index of the city (used to ensure correct city is selected)
            city_already_selected: If True, skip city selection (city was already selected)
        """
        try:
            # Make sure we're on the correct city first (unless already selected)
            if not city_already_selected:
                current_city = self._get_city_by_index(city_index)
                if not current_city:
                    return []
                time.sleep(1.125)  # MAX(0.5, 50% of 1.5) = 0.75 - Wait for banks to load after city selection
            else:
                time.sleep(1.125)  # MAX(0.5, 50% of 1.0) = 0.5 - Wait a bit even if city already selected
            
            # Click dropdown trigger to open
            dropdown_trigger = self.driver.find_element(By.ID, "ext-gen1069")
            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", dropdown_trigger)
            time.sleep(0.75)  # MAX(0.5, 50% of 0.5) = 0.5 - Wait before clicking
            self.driver.execute_script("arguments[0].click();", dropdown_trigger)
            
            # Wait for dropdown to appear
            time.sleep(0.75)  # MAX(0.5, 50% of 0.5) = 0.5
            wait = WebDriverWait(self.driver, 10)
            wait.until(EC.presence_of_element_located((By.XPATH, "//div[contains(@class, 'x-boundlist')] | //table")))
            
            # Wait for tbody to be present and populated
            try:
                wait.until(EC.presence_of_element_located((By.XPATH, "//tbody[@id='treeview-1022-body']")))
                time.sleep(1.125)  # MAX(0.5, 50% of 1.0) = 0.5 - Wait for spans to be rendered
            except:
                pass
            
            # Get all spans with class="x-tree-node-text" within tbody id="treeview-1022-body"
            # This is the bank dropdown, not the checkbox area
            # THE FIX: Only use visible spans for indexing - filter immediately
            # Ensure we ONLY check spans inside treeview-1022-body tbody
            all_spans = self.driver.find_elements(By.XPATH, "//tbody[@id='treeview-1022-body']//span[@class='x-tree-node-text' or contains(@class, 'x-tree-node-text')]")
            
            # Filter to only visible, non-empty spans immediately
            # The XPath already ensures we're only getting spans from treeview-1022-body
            span_elements = [
                s for s in all_spans
                if s.is_displayed() and s.text.strip()
            ]
            
            bank_names = []
            skip_labels = ["Laporan Posisi Keuangan", "Laporan Laba Rugi", "Laporan", "Posisi", "Keuangan", "Laba", "Rugi"]
            
            if span_elements:
                for span in span_elements:
                    span_text = span.text.strip()
                    
                    # Skip if it's a known label (not a bank name)
                    is_label = any(label.lower() in span_text.lower() for label in skip_labels)
                    if is_label:
                        continue
                    
                    # Bank names typically contain numbers (bank codes)
                    has_number = any(char.isdigit() for char in span_text)
                    
                    # Only add if it looks like a bank name (has number or is reasonably long)
                    if has_number or len(span_text) > 15:
                        bank_names.append(span_text)
                        print(f"    [DEBUG] Found bank: '{span_text[:50]}...'")
            
            # Close dropdown
            try:
                from selenium.webdriver.common.keys import Keys
                dropdown_trigger.send_keys(Keys.ESCAPE)
            except:
                pass
            
            return bank_names
        except Exception as e:
            print(f"    [DEBUG] Could not get all bank names: {e}")
            return []
    
    def _select_bank_by_index(self, bank_index: int, city_index: int, city_already_selected: bool = False) -> str:
        """Select a bank by its index from dropdown ext-gen1069. 
        
        Args:
            bank_index: Index of the bank to select (0 = first span, 1 = second span, etc.)
            city_index: Index of the city (used to ensure correct city is selected)
            city_already_selected: If True, skip city selection (city was already selected)
        
        Returns:
            Bank name if successful, empty string if failed.
        """
        try:
            # Make sure we're on the correct city first (unless already selected)
            if not city_already_selected:
                current_city = self._get_city_by_index(city_index)
                if not current_city:
                    return ""
                time.sleep(1.125)  # MAX(0.5, 50% of 1.5) = 0.75 - Wait for banks to load after city selection
            else:
                # If this is index 0 (first bank), wait longer to ensure dropdown is ready
                if bank_index == 0:
                    time.sleep(1.125)  # MAX(0.5, 50% of 1.0) = 0.5 - Wait for first bank after city change
                else:
                    time.sleep(1.125)  # MAX(0.5, 50% of 1.0) = 0.5 - Wait between bank selections
            
            # Click dropdown trigger to open
            dropdown_trigger = self.driver.find_element(By.ID, "ext-gen1069")
            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", dropdown_trigger)
            time.sleep(0.75)  # MAX(0.5, 50% of 0.3) = 0.5
            self.driver.execute_script("arguments[0].click();", dropdown_trigger)
            
            # Wait for dropdown to appear
            time.sleep(0.75)  # MAX(0.5, 50% of 0.5) = 0.5
            wait = WebDriverWait(self.driver, 10)
            wait.until(EC.presence_of_element_located((By.XPATH, "//div[contains(@class, 'x-boundlist')] | //table")))
            
            # Wait for tbody to be present and populated
            # For index 0 (first bank), wait longer to ensure dropdown is fully loaded
            wait_time = 1.125 if bank_index == 0 else 0.75  # Increased by 50%
            try:
                wait.until(EC.presence_of_element_located((By.XPATH, "//tbody[@id='treeview-1022-body']")))
                time.sleep(wait_time)  # Longer wait for first bank, shorter for others
            except:
                pass
            
            # Additional wait for spans to be rendered, especially for index 0
            if bank_index == 0:
                time.sleep(1.125)  # MAX(0.5, 50% of 1.0) = 0.5
            
            # Get all spans with class="x-tree-node-text" within tbody id="treeview-1022-body"
            # This ensures we're clicking the dropdown tr, not the checkbox tr
            # THE FIX: Only use visible spans for indexing - filter immediately
            # Ensure we ONLY check spans inside treeview-1022-body tbody
            all_spans = self.driver.find_elements(By.XPATH, "//tbody[@id='treeview-1022-body']//span[@class='x-tree-node-text' or contains(@class, 'x-tree-node-text')]")
            
            # Filter to only visible, non-empty spans immediately
            # The XPath already ensures we're only getting spans from treeview-1022-body
            span_elements = [
                s for s in all_spans
                if s.is_displayed() and s.text.strip()
            ]
            
            print(f"    [DEBUG] Found {len(span_elements)} visible, non-empty span elements in dropdown (from treeview-1022-body)")
            
            # Filter spans the same way as _get_all_bank_names
            skip_labels = ["Laporan Posisi Keuangan", "Laporan Laba Rugi", "Laporan", "Posisi", "Keuangan", "Laba", "Rugi"]
            valid_bank_spans = []
            
            for span in span_elements:
                try:
                    span_text = span.text.strip()
                    
                    # Skip if it's a known label (not a bank name)
                    is_label = any(label.lower() in span_text.lower() for label in skip_labels)
                    if is_label:
                        continue
                    
                    # Bank names typically contain numbers (bank codes)
                    has_number = any(char.isdigit() for char in span_text)
                    
                    # Only add if it looks like a bank name (has number or is reasonably long)
                    if has_number or len(span_text) > 15:
                        valid_bank_spans.append((span, span_text))
                        print(f"    [DEBUG] Valid bank span {len(valid_bank_spans)-1}: '{span_text[:50]}...'")
                except:
                    # Element might be stale, skip it
                    continue
            
            print(f"    [DEBUG] Total valid bank spans: {len(valid_bank_spans)}")
            
            # Select by index
            if bank_index < len(valid_bank_spans):
                selected_span, bank_name = valid_bank_spans[bank_index]
                
                # Find the parent tr within the tbody (this is the dropdown row, not checkbox)
                try:
                    # Find the closest tr ancestor (should be within treeview-1022-body)
                    parent_tr = selected_span.find_element(By.XPATH, "./ancestor::tr[1]")
                    clickable_elem = parent_tr
                except:
                    # Fallback: try to find any clickable ancestor
                    try:
                        clickable_elem = selected_span.find_element(By.XPATH, "./ancestor::*[@role='row' or contains(@class, 'x-boundlist-item')][1]")
                    except:
                        clickable_elem = selected_span
                
                # Select it
                self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", clickable_elem)
                time.sleep(1.125)  # MAX(0.5, 50% of 0.3) = 0.5
                self.driver.execute_script("arguments[0].click();", clickable_elem)
                time.sleep(1.125)  # MAX(0.5, 50% of 0.5) = 0.5 - Wait for PostBack
                print(f"    [DEBUG] Selected bank at index {bank_index}: '{bank_name[:50]}...'")
                return bank_name
            else:
                # Close dropdown if index out of range
                try:
                    from selenium.webdriver.common.keys import Keys
                    dropdown_trigger.send_keys(Keys.ESCAPE)
                except:
                    pass
                print(f"    [WARNING] Bank index {bank_index} is out of range. Total valid banks: {len(valid_bank_spans)}")
                return ""
        except Exception as e:
            print(f"    [DEBUG] Could not select bank by index {bank_index}: {e}")
            return ""
    
    def _get_bank_by_index(self, city_index: int, bank_index: int, city_already_selected: bool = False) -> str:
        """Get bank name by index from dropdown ext-gen1069. Returns bank name or None if not found.
        
        Args:
            city_index: Index of the city (used to ensure correct city is selected)
            bank_index: Index of the bank to select
            city_already_selected: If True, skip city selection (city was already selected)
        """
        # Get all bank names first
        bank_names = self._get_all_bank_names(city_index, city_already_selected)
        
        if bank_index < len(bank_names):
            bank_name = bank_names[bank_index]
            # Select the bank by name
            if self._select_bank_by_name(bank_name):
                return bank_name
        
        return None
    
    def _click_tampilkan_and_extract_data(self, year: str, city: str, bank: str, extract_mode: str = 'sheets_1_3', skip_laba_kotor: bool = False, skip_rasio: bool = False) -> dict:
        """
        Click Tampilkan button, wait for report, and extract data
        
        Args:
            year: Selected year
            city: City name
            bank: Bank name
            extract_mode: 'sheets_1_3' to extract ASET/KREDIT/DPK, 'sheets_4_5' to extract LABA KOTOR/RASIO
            skip_laba_kotor: If True, skip Laba Kotor extraction (for phase 003)
            skip_rasio: If True, skip Rasio extraction (for phase 002)
        """
        try:
            # Make sure we're on default content and close any open dropdowns
            self.driver.switch_to.default_content()
            time.sleep(0.75)  # MAX(0.5, 50% of 0.3) = 0.5
            
            # Close any open dropdowns by pressing ESC
            try:
                from selenium.webdriver.common.keys import Keys
                body = self.driver.find_element(By.TAG_NAME, "body")
                body.send_keys(Keys.ESCAPE)
                time.sleep(1.125)  # MAX(0.5, 50% of 0.3) = 0.5
            except:
                pass
            
            # Click Tampilkan button
            tampilkan_button = None
            try:
                tampilkan_button = self.driver.find_element(By.ID, "ShowReportButton-btnInnerEl")
            except:
                tampilkan_button = self.driver.find_element(By.XPATH, "//span[contains(text(), 'Tampilkan')]")
            
            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", tampilkan_button)
            time.sleep(0.75)
            self.driver.execute_script("arguments[0].click();", tampilkan_button)
            print(f"  [OK] Clicked 'Tampilkan' button")
            
            # Check for period error and handle it
            error_found, new_month, new_year = self._check_and_handle_period_error(self.current_month, self.current_year)
            if error_found:
                # Update current month/year
                self.current_month = new_month
                self.current_year = new_year
                year = new_year  # Update year for this extraction
                print(f"  [INFO] Period updated due to error. Using {new_month} {new_year}")
                # Skip wait attempts - pass flag to _extract_report_data
                skip_wait = True
            else:
                skip_wait = False
            
            # Wait for report to load (unless error was found)
            if not skip_wait:
                print(f"  [INFO] Waiting for report to load...")
                self._wait_for_report_loaded(max_wait=30)
            
            # Extract data with Bad Request retry logic (up to 2 retries)
            max_retries = 2
            for retry_attempt in range(max_retries + 1):
                extracted_data = self._extract_report_data(year, city, bank, extract_mode=extract_mode, skip_wait_attempts=skip_wait, skip_laba_kotor=skip_laba_kotor, skip_rasio=skip_rasio)
                
                # Check if Bad Request was detected (extracted_data is None)
                if extracted_data is None:
                    if retry_attempt < max_retries:
                        print(f"  [WARNING] Bad Request terdeteksi, mencoba klik Tampilkan lagi (percobaan {retry_attempt + 1}/{max_retries})...")
                        # Re-click Tampilkan button
                        try:
                            tampilkan_button = None
                            try:
                                tampilkan_button = self.driver.find_element(By.ID, "ShowReportButton-btnInnerEl")
                            except:
                                tampilkan_button = self.driver.find_element(By.XPATH, "//span[contains(text(), 'Tampilkan')]")
                            
                            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", tampilkan_button)
                            time.sleep(1.125)
                            self.driver.execute_script("arguments[0].click();", tampilkan_button)
                            print(f"  [OK] Clicked 'Tampilkan' button again (retry {retry_attempt + 1})")
                            
                            # Wait a bit before retrying
                            time.sleep(3.0)
                            continue  # Retry extraction
                        except Exception as e:
                            print(f"  [ERROR] Error re-clicking Tampilkan: {e}")
                            return None
                    else:
                        print(f"  [WARNING] Bad Request masih terjadi setelah {max_retries} percobaan, melewati bank ini...")
                        return None
                else:
                    # Success - ensure city is correctly set in extracted_data for Laba Kotor/Rasio
                    if extract_mode == 'sheets_4_5' and extracted_data:
                        # Ensure city is set correctly (don't let it be overwritten by extraction)
                        extracted_data['city'] = city
                        extracted_data['bank'] = bank
                    return extracted_data
            
            return None
        except Exception as e:
            print(f"  [ERROR] Error in _click_tampilkan_and_extract_data: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def _initialize_excel(self, year: str):
        """Initialize data storage (Excel will be created at the end)"""
        self.all_data = []  # Clear any previous data
        print(f"  [OK] Data storage initialized")
    
    def _append_to_excel(self, data: dict, year: str, city: str, bank: str, is_first_bank_in_city: bool, data_list: str = 'sheets_1_3'):
        """
        Store extracted data for later Excel generation
        
        Args:
            data: Extracted data dictionary
            year: Selected year
            city: City name
            bank: Bank name
            is_first_bank_in_city: Whether this is the first bank in the city
            data_list: Which data list to use ('sheets_1_3' or 'sheets_4_5')
        """
        try:
            previous_year = str(int(year) - 1)
            
            # Store data with all necessary fields
            record = {
                'city': city,
                'bank': bank,
            }
            
            # Add fields that exist in data (for both first 3 sheets and Laba Kotor)
            if f'Kredit {year}' in data:
                record[f'Kredit {year}'] = data.get(f'Kredit {year}', 0)
                record[f'Kredit {previous_year}'] = data.get(f'Kredit {previous_year}', 0)
            if f'Total Aset {year}' in data:
                record[f'Total Aset {year}'] = data.get(f'Total Aset {year}', 0)
                record[f'Total Aset {previous_year}'] = data.get(f'Total Aset {previous_year}', 0)
            if f'DPK {year}' in data:
                record[f'DPK {year}'] = data.get(f'DPK {year}', 0)
                record[f'DPK {previous_year}'] = data.get(f'DPK {previous_year}', 0)
            if f'Laba Kotor {year}' in data:
                record[f'Laba Kotor {year}'] = data.get(f'Laba Kotor {year}', 0)
                record[f'Laba Kotor {previous_year}'] = data.get(f'Laba Kotor {previous_year}', 0)
            
            # Add ratio fields (for Sheet 5)
            ratio_names = ['KPMM', 'PPKA', 'NPL Neto', 'NPL Gross', 'ROA', 'BOPO', 'NIM', 'LDR', 'CR']
            for ratio_name in ratio_names:
                if ratio_name in data:
                    record[ratio_name] = data.get(ratio_name, 0)
            
            # Store in the appropriate data list
            if data_list == 'sheets_1_3':
                self.sheets_1_3_data.append(record)
            elif data_list == 'sheets_4_5':
                self.sheets_4_5_data.append(record)
            else:
                # Fallback to all_data for backward compatibility
                self.all_data.append(record)
            print(f"  [OK] Stored data for {city} - {bank}")
        except Exception as e:
            print(f"  [ERROR] Error storing data: {e}")
            import traceback
            traceback.print_exc()
    
    def _get_month_number(self, month_name: str) -> str:
        """
        Convert month name to number (e.g., 'Desember' -> '12')
        
        Month mapping:
        - 03 is Maret
        - 06 is Juni
        - 09 is September
        - 12 is Desember
        """
        month_map = {
            'januari': '01', 'februari': '02', 'maret': '03', 'april': '04',
            'mei': '05', 'juni': '06', 'juli': '07', 'agustus': '08',
            'september': '09', 'oktober': '10', 'november': '11', 'desember': '12'
        }
        return month_map.get(month_name.lower(), '12')
    
    def _get_excel_filename(self, month: str, year: str) -> str:
        """
        Generate Excel filename in format: Publikasi_MM_YYYY.xlsx
        e.g., Publikasi_09_2025.xlsx
        
        Args:
            month: Month name (e.g., "September")
            year: Year (e.g., "2025")
            
        Returns:
            Filename string
        """
        month_num = self._get_month_number(month)
        # month_num is already a string with zero-padding (e.g., "09")
        # Just use it directly
        return f"Publikasi_{month_num}_{year}.xlsx"
    
    def _extract_bank_name(self, bank_full: str) -> str:
        """Extract bank name by removing prefix numbers and '-' (e.g., '600784-Perumda BPR Tuah Karimun' -> 'Perumda BPR Tuah Karimun')"""
        if '-' in bank_full:
            parts = bank_full.split('-', 1)
            if len(parts) == 2:
                # Check if first part is all digits
                if parts[0].strip().isdigit():
                    return parts[1].strip()
        return bank_full.strip()
    
    def _update_excel_row(self, workbook, sheet_name: str, bank_name: str, city: str, new_data: dict, year: str):
        """
        Update an existing row in Excel sheet based on bank name and city.
        Applies update rules: zero→non-zero (update), non-zero→zero (keep existing), non-zero→non-zero (update)
        
        Args:
            workbook: openpyxl Workbook object
            sheet_name: Name of the sheet to update
            bank_name: Bank name to match
            city: City name to match
            new_data: Dictionary with new data values
            year: Current year
            
        Returns:
            Tuple of (row_found: bool, row_num: int) - row_num is 0 if not found
        """
        if sheet_name not in workbook.sheetnames:
            return False, 0
        
        ws = workbook[sheet_name]
        previous_year = str(int(year) - 1)
        
        # Find the row by matching bank name (column B) and city (column C)
        # Headers are in row 2, data starts at row 3
        for row_num in range(3, ws.max_row + 1):
            try:
                existing_bank = ws.cell(row=row_num, column=2).value  # Column B: Nama Bank
                existing_city = ws.cell(row=row_num, column=3).value  # Column C: Lokasi
                
                if existing_bank and existing_city:
                    # Normalize for comparison
                    if (str(existing_bank).strip().lower() == bank_name.strip().lower() and
                        str(existing_city).strip().lower() == city.strip().lower()):
                        
                        # Found matching row - update based on rules
                        # Column D: Current year, Column E: Previous year
                        # Determine which data field to update based on sheet name
                        data_key = None
                        if 'ASET' in sheet_name:
                            data_key = 'Total Aset'
                        elif 'Kredit' in sheet_name:
                            data_key = 'Kredit'
                        elif 'DPK' in sheet_name:
                            data_key = 'DPK'
                        elif 'Laba Kotor' in sheet_name:
                            data_key = 'Laba Kotor'
                        
                        if data_key:
                            current_key = f'{data_key} {year}'
                            previous_key = f'{data_key} {previous_year}'
                            
                            # Get existing values
                            existing_current = ws.cell(row=row_num, column=4).value or 0
                            existing_previous = ws.cell(row=row_num, column=5).value or 0
                            
                            # Get new values
                            new_current = new_data.get(current_key, 0) or 0
                            new_previous = new_data.get(previous_key, 0) or 0
                            
                            # Apply update rules
                            # Rule 1: If existing is 0 and new is non-zero → Update
                            # Rule 2: If existing is non-zero and new is 0 → Keep existing
                            # Rule 3: If both are non-zero → Update
                            if existing_current == 0 and new_current != 0:
                                ws.cell(row=row_num, column=4).value = new_current
                            elif existing_current != 0 and new_current == 0:
                                pass  # Keep existing
                            elif existing_current != 0 and new_current != 0:
                                ws.cell(row=row_num, column=4).value = new_current
                            elif existing_current == 0 and new_current == 0:
                                pass  # Both zero, no change needed
                            
                            if existing_previous == 0 and new_previous != 0:
                                ws.cell(row=row_num, column=5).value = new_previous
                            elif existing_previous != 0 and new_previous == 0:
                                pass  # Keep existing
                            elif existing_previous != 0 and new_previous != 0:
                                ws.cell(row=row_num, column=5).value = new_previous
                            
                            # Recalculate Peningkatan (column 6)
                            current_val = ws.cell(row=row_num, column=4).value or 0
                            previous_val = ws.cell(row=row_num, column=5).value or 0
                            if previous_val and previous_val != 0:
                                peningkatan = ((current_val - previous_val) / abs(previous_val)) * 100
                            else:
                                peningkatan = 0 if current_val == 0 else 100
                            ws.cell(row=row_num, column=6).value = peningkatan / 100
                        
                        return True, row_num
            except Exception as e:
                continue
        
        return False, 0
    
    def _finalize_excel(self, month: str, year: str):
        """Create or update Excel workbook with three sheets (ASET, Kredit, DPK)"""
        if not Workbook:
            print("  [ERROR] openpyxl not installed. Cannot create Excel file.")
            return
        
        # Use sheets_1_3_data instead of all_data
        data_to_use = self.sheets_1_3_data if hasattr(self, 'sheets_1_3_data') and self.sheets_1_3_data else self.all_data
        
        if not data_to_use:
            print("  [WARNING] No data to export")
            return
        
        try:
            from openpyxl.styles import Font, Border, Side, Alignment
            from openpyxl.utils import get_column_letter
            from openpyxl import load_workbook
            
            # Get filename using new format
            filename = self._get_excel_filename(month, year)
            # Save to output/publikasi subdirectory
            publikasi_dir = self.output_dir / "publikasi"
            publikasi_dir.mkdir(parents=True, exist_ok=True)
            filepath = publikasi_dir / filename
            
            # Check if file exists
            if filepath.exists():
                print(f"  [INFO] Excel file exists: {filename}, loading and updating...")
                self.excel_wb = load_workbook(filepath)
            else:
                print(f"  [INFO] Creating new Excel file: {filename}")
                self.excel_wb = Workbook()
                # Remove default sheet
                if 'Sheet' in self.excel_wb.sheetnames:
                    self.excel_wb.remove(self.excel_wb['Sheet'])
            
            month_num = self._get_month_number(month)
            previous_year = str(int(year) - 1)
            sheet_name_prefix = f"{month_num}-{year[-2:]}"
            
            # Define border style
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Create three sheets
            sheets_data = [
                ('ASET', 'Total Aset', f'PERINGKAT ASET BPR PERIODE {month_num} {year}'),
                ('Kredit', 'Kredit', f'PERINGKAT KREDIT BPR PERIODE {month_num} {year}'),
                ('DPK', 'DPK', f'PERINGKAT DPK BPR PERIODE {month_num} {year}')
            ]
            
            for sheet_type, data_key, title in sheets_data:
                sheet_name = f"{sheet_name_prefix} {sheet_type}"
                
                # Check if sheet exists, if not create it
                if sheet_name in self.excel_wb.sheetnames:
                    ws = self.excel_wb[sheet_name]
                    print(f"  [INFO] Sheet '{sheet_name}' exists, updating...")
                else:
                    ws = self.excel_wb.create_sheet(title=sheet_name)
                    print(f"  [INFO] Created new sheet '{sheet_name}'")
                    
                    # Title row (row 1)
                    ws.merge_cells(f'A1:F1')
                    title_cell = ws['A1']
                    title_cell.value = title
                    title_cell.font = Font(bold=True, size=14)
                    title_cell.alignment = Alignment(horizontal='center', vertical='center')
                    title_cell.border = thin_border
                    
                    # Header row (row 2)
                    headers = ['No', 'Nama Bank', 'Lokasi', year, previous_year, 'Peningkatan']
                    for col_idx, header in enumerate(headers, start=1):
                        cell = ws.cell(row=2, column=col_idx)
                        cell.value = header
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = thin_border
                
                # Sort data by current year column (descending) for this sheet type
                sorted_data = sorted(data_to_use, key=lambda x: x.get(f'{data_key} {year}', 0), reverse=True)
                
                # Process each record - update existing or add new
                for record in sorted_data:
                    bank_name = self._extract_bank_name(record['bank'])
                    city = record['city']
                    
                    # Try to update existing row
                    row_found, existing_row = self._update_excel_row(self.excel_wb, sheet_name, bank_name, city, record, year)
                    
                    if not row_found:
                        # Row doesn't exist, add new row
                        row_num = ws.max_row + 1
                        if row_num < 3:
                            row_num = 3  # Ensure we start at row 3 (after headers)
                        
                        # Get values for this sheet type
                        current_value = record.get(f'{data_key} {year}', 0)
                        previous_value = record.get(f'{data_key} {previous_year}', 0)
                        
                        # Calculate Peningkatan
                        if previous_value and previous_value != 0:
                            peningkatan = ((current_value - previous_value) / abs(previous_value)) * 100
                        else:
                            peningkatan = 0 if current_value == 0 else 100
                        
                        # Write data
                        ws.cell(row=row_num, column=1).value = row_num - 2  # No (starting from 1)
                        ws.cell(row=row_num, column=2).value = bank_name  # Nama Bank
                        ws.cell(row=row_num, column=3).value = city  # Lokasi
                        ws.cell(row=row_num, column=4).value = current_value  # Current year
                        ws.cell(row=row_num, column=5).value = previous_value  # Previous year
                        ws.cell(row=row_num, column=6).value = peningkatan / 100  # Peningkatan
                        
                        # Apply formatting
                        for col_idx in range(1, 7):
                            cell = ws.cell(row=row_num, column=col_idx)
                            cell.border = thin_border
                            if col_idx == 1:  # No column
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                            elif col_idx == 6:  # Peningkatan
                                cell.number_format = '0.00%'
                                cell.alignment = Alignment(horizontal='right', vertical='center')
                            elif col_idx in [4, 5]:  # Year columns
                                cell.number_format = '#,##0'
                                cell.alignment = Alignment(horizontal='right', vertical='center')
                            else:  # Text columns
                                cell.alignment = Alignment(horizontal='left', vertical='center')
                
                # Re-sort and renumber all rows after updates
                # Collect all data rows
                data_rows = []
                for row_num in range(3, ws.max_row + 1):
                    bank_val = ws.cell(row=row_num, column=2).value
                    city_val = ws.cell(row=row_num, column=3).value
                    current_val = ws.cell(row=row_num, column=4).value or 0
                    if bank_val and city_val:
                        data_rows.append((row_num, current_val, bank_val, city_val))
                
                # Sort by current value (descending)
                data_rows.sort(key=lambda x: x[1], reverse=True)
                
                # Create temporary list with sorted data
                sorted_rows_data = []
                for orig_row, current_val, bank_val, city_val in data_rows:
                    prev_val = ws.cell(row=orig_row, column=5).value or 0
                    peningkatan_val = ws.cell(row=orig_row, column=6).value or 0
                    sorted_rows_data.append((bank_val, city_val, current_val, prev_val, peningkatan_val))
                
                # Clear existing data rows and rewrite in sorted order
                for row_num in range(3, ws.max_row + 1):
                    for col_idx in range(1, 7):
                        ws.cell(row=row_num, column=col_idx).value = None
                
                # Write sorted data
                for idx, (bank_val, city_val, current_val, prev_val, peningkatan_val) in enumerate(sorted_rows_data, start=1):
                    row_num = idx + 2
                    ws.cell(row=row_num, column=1).value = idx
                    ws.cell(row=row_num, column=2).value = bank_val
                    ws.cell(row=row_num, column=3).value = city_val
                    ws.cell(row=row_num, column=4).value = current_val
                    ws.cell(row=row_num, column=5).value = prev_val
                    ws.cell(row=row_num, column=6).value = peningkatan_val
                    
                    # Apply formatting
                    for col_idx in range(1, 7):
                        cell = ws.cell(row=row_num, column=col_idx)
                        cell.border = thin_border
                        if col_idx == 1:
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        elif col_idx == 6:
                            cell.number_format = '0.00%'
                            cell.alignment = Alignment(horizontal='right', vertical='center')
                        elif col_idx in [4, 5]:
                            cell.number_format = '#,##0'
                            cell.alignment = Alignment(horizontal='right', vertical='center')
                        else:
                            cell.alignment = Alignment(horizontal='left', vertical='center')
                
                # Set column widths
                ws.column_dimensions['A'].width = 6   # No
                ws.column_dimensions['B'].width = 50  # Nama Bank
                ws.column_dimensions['C'].width = 30  # Lokasi
                ws.column_dimensions['D'].width = 18  # Current year
                ws.column_dimensions['E'].width = 18  # Previous year
                ws.column_dimensions['F'].width = 15  # Peningkatan
                
                # Set row heights
                ws.row_dimensions[1].height = 25  # Title row
                ws.row_dimensions[2].height = 20  # Header row
            
            # Save file
            self.excel_wb.save(filepath)
            print(f"  [OK] Excel file saved to {filepath}")
            print(f"  [OK] Total records processed: {len(data_to_use)}")
            print(f"  [OK] Updated/Created 3 sheets: {sheet_name_prefix} ASET, {sheet_name_prefix} Kredit, {sheet_name_prefix} DPK")
        except Exception as e:
            print(f"  [ERROR] Error creating Excel file: {e}")
            import traceback
            traceback.print_exc()
    
    def _finalize_excel_laba_kotor(self, month: str, year: str):
        """Add or update Sheet 4 (Laba Kotor) in Excel workbook"""
        if not Workbook:
            print("  [ERROR] openpyxl not installed. Cannot add Laba Kotor sheet.")
            return
        
        # Use sheets_4_5_data instead of all_data
        data_to_use = self.sheets_4_5_data if hasattr(self, 'sheets_4_5_data') and self.sheets_4_5_data else self.all_data
        
        if not data_to_use:
            print("  [WARNING] No Laba Kotor data to export")
            return
        
        try:
            from openpyxl.styles import Font, Border, Side, Alignment
            from openpyxl import load_workbook
            
            # Get filename using new format
            filename = self._get_excel_filename(month, year)
            # Save to output/publikasi subdirectory
            publikasi_dir = self.output_dir / "publikasi"
            publikasi_dir.mkdir(parents=True, exist_ok=True)
            filepath = publikasi_dir / filename
            
            # Load existing workbook (should exist from Sheets 1-3)
            if not filepath.exists():
                print(f"  [ERROR] Excel file not found: {filepath}")
                print(f"  [ERROR] Sheets 1-3 should have been created first!")
                return
            
            self.excel_wb = load_workbook(filepath)
            
            month_num = self._get_month_number(month)
            previous_year = str(int(year) - 1)
            sheet_name_prefix = f"{month_num}-{year[-2:]}"
            
            # Define border style
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Check if sheet exists, if not create it
            sheet_name = f"{sheet_name_prefix} Laba Kotor"
            if sheet_name in self.excel_wb.sheetnames:
                ws = self.excel_wb[sheet_name]
                print(f"  [INFO] Sheet '{sheet_name}' exists, updating...")
            else:
                ws = self.excel_wb.create_sheet(title=sheet_name)
                print(f"  [INFO] Created new sheet '{sheet_name}'")
                
                # Remove default sheet only after we've created the new sheet (to avoid corrupt file)
                if 'Sheet' in self.excel_wb.sheetnames:
                    self.excel_wb.remove(self.excel_wb['Sheet'])
                title = f'PERINGKAT LABA KOTOR BPR PERIODE {month_num} {year}'
                
                # Title row (row 1)
                ws.merge_cells(f'A1:F1')
                title_cell = ws['A1']
                title_cell.value = title
                title_cell.font = Font(bold=True, size=14)
                title_cell.alignment = Alignment(horizontal='center', vertical='center')
                title_cell.border = thin_border
                
                # Header row (row 2)
                headers = ['No', 'Nama Bank', 'Lokasi', year, previous_year, 'Peningkatan']
                for col_idx, header in enumerate(headers, start=1):
                    cell = ws.cell(row=2, column=col_idx)
                    cell.value = header
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border
            
            # Filter records that have Laba Kotor data
            laba_kotor_records = [r for r in data_to_use if f'Laba Kotor {year}' in r or 'Laba Kotor' in str(r)]
            if not laba_kotor_records:
                # If no Laba Kotor data, use all records (they might have 0 values)
                laba_kotor_records = data_to_use
                print(f"  [INFO] No Laba Kotor fields found, using all {len(laba_kotor_records)} records (may have 0 values)")
            
            # Sort Laba Kotor records by current year column (descending)
            laba_kotor_records.sort(key=lambda x: x.get(f'Laba Kotor {year}', 0), reverse=True)
            
            # Process each record - update existing or add new
            for record in laba_kotor_records:
                bank_name = self._extract_bank_name(record['bank'])
                city = record['city']
                
                # Try to update existing row
                row_found, existing_row = self._update_excel_row(self.excel_wb, sheet_name, bank_name, city, record, year)
                
                if not row_found:
                    # Row doesn't exist, add new row
                    row_num = ws.max_row + 1
                    if row_num < 3:
                        row_num = 3
                    
                    # Get Laba Kotor values
                    current_value = record.get(f'Laba Kotor {year}', 0)
                    previous_value = record.get(f'Laba Kotor {previous_year}', 0)
                    
                    # Calculate Peningkatan
                    if previous_value and previous_value != 0:
                        peningkatan = ((current_value - previous_value) / abs(previous_value)) * 100
                    else:
                        peningkatan = 0 if current_value == 0 else 100
                    
                    # Write data
                    ws.cell(row=row_num, column=1).value = row_num - 2
                    ws.cell(row=row_num, column=2).value = bank_name
                    ws.cell(row=row_num, column=3).value = city
                    ws.cell(row=row_num, column=4).value = current_value
                    ws.cell(row=row_num, column=5).value = previous_value
                    ws.cell(row=row_num, column=6).value = peningkatan / 100
                    
                    # Apply formatting
                    for col_idx in range(1, 7):
                        cell = ws.cell(row=row_num, column=col_idx)
                        cell.border = thin_border
                        if col_idx == 1:
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        elif col_idx == 6:
                            cell.number_format = '0.00%'
                            cell.alignment = Alignment(horizontal='right', vertical='center')
                        elif col_idx in [4, 5]:
                            cell.number_format = '#,##0'
                            cell.alignment = Alignment(horizontal='right', vertical='center')
                        else:
                            cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Re-sort and renumber all rows after updates (similar to _finalize_excel)
            data_rows = []
            for row_num in range(3, ws.max_row + 1):
                bank_val = ws.cell(row=row_num, column=2).value
                city_val = ws.cell(row=row_num, column=3).value
                current_val = ws.cell(row=row_num, column=4).value or 0
                if bank_val and city_val:
                    data_rows.append((row_num, current_val, bank_val, city_val))
            
            data_rows.sort(key=lambda x: x[1], reverse=True)
            sorted_rows_data = []
            for orig_row, current_val, bank_val, city_val in data_rows:
                prev_val = ws.cell(row=orig_row, column=5).value or 0
                peningkatan_val = ws.cell(row=orig_row, column=6).value or 0
                sorted_rows_data.append((bank_val, city_val, current_val, prev_val, peningkatan_val))
            
            # Clear and rewrite in sorted order
            for row_num in range(3, ws.max_row + 1):
                for col_idx in range(1, 7):
                    ws.cell(row=row_num, column=col_idx).value = None
            
            for idx, (bank_val, city_val, current_val, prev_val, peningkatan_val) in enumerate(sorted_rows_data, start=1):
                row_num = idx + 2
                ws.cell(row=row_num, column=1).value = idx
                ws.cell(row=row_num, column=2).value = bank_val
                ws.cell(row=row_num, column=3).value = city_val
                ws.cell(row=row_num, column=4).value = current_val
                ws.cell(row=row_num, column=5).value = prev_val
                ws.cell(row=row_num, column=6).value = peningkatan_val
                
                for col_idx in range(1, 7):
                    cell = ws.cell(row=row_num, column=col_idx)
                    cell.border = thin_border
                    if col_idx == 1:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    elif col_idx == 6:
                        cell.number_format = '0.00%'
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    elif col_idx in [4, 5]:
                        cell.number_format = '#,##0'
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Set column widths
            ws.column_dimensions['A'].width = 6   # No
            ws.column_dimensions['B'].width = 50  # Nama Bank
            ws.column_dimensions['C'].width = 30  # Lokasi
            ws.column_dimensions['D'].width = 18  # Current year
            ws.column_dimensions['E'].width = 18  # Previous year
            ws.column_dimensions['F'].width = 15  # Peningkatan
            
            # Set row heights
            ws.row_dimensions[1].height = 25  # Title row
            ws.row_dimensions[2].height = 20  # Header row
            
            # Save file
            self.excel_wb.save(filepath)
            print(f"  [OK] Updated/Created Sheet 4: {sheet_name_prefix} Laba Kotor")
            print(f"  [OK] Total Laba Kotor records processed: {len(laba_kotor_records)}")
            print(f"  [OK] Excel file saved to: {filepath}")
        except Exception as e:
            print(f"  [ERROR] Error adding Laba Kotor sheet: {e}")
            import traceback
            traceback.print_exc()
    
    def _finalize_excel_rasio(self, month: str, year: str):
        """Add Sheet 5 (Rasio) to existing Excel workbook with 9 separate tables"""
        if not Workbook:
            print("  [ERROR] openpyxl not installed. Cannot add Rasio sheet.")
            return
        
        # Use sheets_4_5_data instead of all_data
        data_to_use = self.sheets_4_5_data if hasattr(self, 'sheets_4_5_data') and self.sheets_4_5_data else self.all_data
        
        if not data_to_use:
            print("  [WARNING] No Rasio data to export")
            # Still try to create sheet if workbook exists
            try:
                filename = self._get_excel_filename(month, year)
                filepath = self.output_dir / filename
                if filepath.exists():
                    from openpyxl import load_workbook
                    self.excel_wb = load_workbook(filepath)
                    self.excel_wb.save(filepath)
                    print(f"  [INFO] Excel file saved (no Rasio data): {filepath}")
            except:
                pass
            return
        
        # Deduplicate records by bank + city
        # If duplicates exist, keep the one with non-zero ratio values (prefer real data over 0,00)
        original_count = len(data_to_use)
        print(f"  [INFO] Deduplicating records (before: {original_count} records)...")
        seen_banks = {}
        deduplicated_data = []
        ratio_names = ['KPMM', 'PPKA', 'NPL Neto', 'NPL Gross', 'ROA', 'BOPO', 'NIM', 'LDR', 'CR']
        
        for record in data_to_use:
            bank_name = record.get('bank', '').strip()
            city = record.get('city', '').strip()
            bank_key = (bank_name, city)
            
            if bank_key not in seen_banks:
                # First time seeing this bank, add it
                seen_banks[bank_key] = record
                deduplicated_data.append(record)
            else:
                # Duplicate found - check which one has better data (non-zero values)
                existing_record = seen_banks[bank_key]
                
                # Count non-zero ratios in existing record
                existing_non_zero_count = sum(
                    1 for ratio in ratio_names 
                    if existing_record.get(ratio, 0) != 0 and existing_record.get(ratio) is not None
                )
                
                # Count non-zero ratios in new record
                new_non_zero_count = sum(
                    1 for ratio in ratio_names 
                    if record.get(ratio, 0) != 0 and record.get(ratio) is not None
                )
                
                # Replace existing if new record has more non-zero ratios
                if new_non_zero_count > existing_non_zero_count:
                    # Remove old record and add new one
                    deduplicated_data.remove(existing_record)
                    deduplicated_data.append(record)
                    seen_banks[bank_key] = record
                    print(f"  [INFO] Replaced duplicate record for {bank_name} ({city}) - new record has {new_non_zero_count} non-zero ratios vs {existing_non_zero_count}")
                elif new_non_zero_count == existing_non_zero_count and new_non_zero_count > 0:
                    # Same number of non-zero ratios - check if new record has any non-zero values where existing has zero
                    should_replace = False
                    for ratio in ratio_names:
                        new_val = record.get(ratio, 0)
                        existing_val = existing_record.get(ratio, 0)
                        if new_val != 0 and existing_val == 0:
                            should_replace = True
                            break
                    
                    if should_replace:
                        deduplicated_data.remove(existing_record)
                        deduplicated_data.append(record)
                        seen_banks[bank_key] = record
                        print(f"  [INFO] Replaced duplicate record for {bank_name} ({city}) - new record has better ratio values")
        
        data_to_use = deduplicated_data
        duplicates_removed = original_count - len(data_to_use)
        print(f"  [OK] Deduplication complete (after: {len(data_to_use)} records, removed {duplicates_removed} duplicates)")
        
        try:
            from openpyxl.styles import Font, Border, Side, Alignment
            from openpyxl import load_workbook
            
            # Get filename using new format
            filename = self._get_excel_filename(month, year)
            # Save to output/publikasi subdirectory
            publikasi_dir = self.output_dir / "publikasi"
            publikasi_dir.mkdir(parents=True, exist_ok=True)
            filepath = publikasi_dir / filename
            
            if not filepath.exists():
                print(f"  [ERROR] Excel file not found: {filepath}")
                return
            
            self.excel_wb = load_workbook(filepath)
            
            month_num = self._get_month_number(month)
            sheet_name_prefix = f"{month_num}-{year[-2:]}"
            
            # Define border style
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Check if Sheet 5: Rasio exists, if not create it
            sheet_name = f"{sheet_name_prefix} Rasio"
            if sheet_name in self.excel_wb.sheetnames:
                ws = self.excel_wb[sheet_name]
                print(f"  [INFO] Sheet '{sheet_name}' exists, updating...")
                # Clear existing data rows (keep title and structure)
                # Find where data starts (after title row 1, empty row 2, then tables start at row 3)
                # For simplicity, we'll recreate the sheet structure
                self.excel_wb.remove(ws)
                ws = self.excel_wb.create_sheet(title=sheet_name)
                print(f"  [INFO] Recreated sheet '{sheet_name}' for fresh update")
            else:
                ws = self.excel_wb.create_sheet(title=sheet_name)
                print(f"  [INFO] Created new sheet '{sheet_name}'")
            
            title = f'PERINGKAT RASIO BPR PERIODE {month_num} {year}'
            
            # Title row (row 1) - merged across all columns
            ws.merge_cells(f'A1:D1')
            title_cell = ws['A1']
            title_cell.value = title
            title_cell.font = Font(bold=True, size=14)
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            title_cell.border = thin_border
            
            # Define 9 ratios with their column names
            ratios = [
                ('KPMM', 'KPMM'),
                ('PPKA', 'PPKA'),
                ('NPL Neto', 'NPL Neto'),
                ('NPL Gross', 'NPL Gross'),
                ('ROA', 'ROA'),
                ('BOPO', 'BOPO'),
                ('NIM', 'NIM'),
                ('LDR', 'LDR'),
                ('CR', 'CR')
            ]
            
            # Starting row for each table
            current_row = 3  # Start after title (row 1) and empty row (row 2)
            
            for ratio_key, ratio_name in ratios:
                # Filter data that has this ratio
                ratio_data = [r for r in data_to_use if ratio_key in r and r.get(ratio_key) is not None]
                
                if not ratio_data:
                    # Skip if no data for this ratio
                    current_row += 50  # Reserve space for next table
                    continue
                
                # Sort by ratio value
                # NPL Neto, NPL Gross, and BOPO should be sorted ascending (lower is better)
                # Other ratios should be sorted descending (higher is better)
                ascending_ratios = ['NPL Neto', 'NPL Gross', 'BOPO']
                reverse_sort = ratio_key not in ascending_ratios
                ratio_data.sort(key=lambda x: x.get(ratio_key, 0), reverse=reverse_sort)
                
                # Table header (row current_row)
                headers = ['No', 'Nama Bank', 'Lokasi', ratio_name]
                for col_idx, header in enumerate(headers, start=1):
                    cell = ws.cell(row=current_row, column=col_idx)
                    cell.value = header
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border
                
                # Data rows
                data_start_row = current_row + 1
                for idx, record in enumerate(ratio_data, start=1):
                    row_num = data_start_row + idx - 1
                    bank_name = self._extract_bank_name(record['bank'])
                    city = record['city']
                    ratio_value = record.get(ratio_key, 0)
                    
                    # Write data
                    ws.cell(row=row_num, column=1).value = idx  # No
                    ws.cell(row=row_num, column=2).value = bank_name  # Nama Bank
                    ws.cell(row=row_num, column=3).value = city  # Lokasi
                    ws.cell(row=row_num, column=4).value = ratio_value  # Ratio value
                    
                    # Apply formatting
                    for col_idx in range(1, 5):
                        cell = ws.cell(row=row_num, column=col_idx)
                        cell.border = thin_border
                        if col_idx == 1:  # No column
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        elif col_idx == 4:  # Ratio value - number format with 2 decimals
                            cell.number_format = '0.00'
                            cell.alignment = Alignment(horizontal='right', vertical='center')
                        else:  # Text columns
                            cell.alignment = Alignment(horizontal='left', vertical='center')
                
                # Move to next table (leave 2 rows gap between tables)
                # Table ends at row (data_start_row + len(ratio_data) - 1)
                # Next table starts at row (data_start_row + len(ratio_data) + 2)
                current_row = data_start_row + len(ratio_data) + 2
            
            # Set column widths
            ws.column_dimensions['A'].width = 6   # No
            ws.column_dimensions['B'].width = 50  # Nama Bank
            ws.column_dimensions['C'].width = 30  # Lokasi
            ws.column_dimensions['D'].width = 15  # Ratio value
            
            # Set row heights
            ws.row_dimensions[1].height = 25  # Title row
            
            # Save file
            self.excel_wb.save(filepath)
            print(f"  [OK] Added Sheet 5: {sheet_name_prefix} Rasio")
            print(f"  [OK] Total Rasio records: {len(data_to_use)}")
            print(f"  [OK] Created 9 tables for ratios: KPMM, PPKA, NPL Neto, NPL Gross, ROA, BOPO, NIM, LDR, CR")
            print(f"  [OK] Excel file saved to: {filepath}")
        except Exception as e:
            print(f"  [ERROR] Error adding Rasio sheet: {e}")
            import traceback
            traceback.print_exc()
    
    def _find_combo_name_by_keyword(self, keyword: str) -> str:
        """
        Find combobox name by keyword
        
        Args:
            keyword: Keyword to search for (e.g., "month", "year", "province")
            
        Returns:
            Component name if found, empty string otherwise
        """
        if self.extjs is None:
            return ""
        
        combos = self.extjs.list_all_combos()
        keyword_lower = keyword.lower()
        
        for combo in combos:
            name = (combo.get('name') or '').lower()
            id_val = (combo.get('id') or '').lower()
            input_id = (combo.get('inputId') or '').lower()
            
            if (keyword_lower in name or 
                keyword_lower in id_val or 
                keyword_lower in input_id):
                return combo.get('name', '')
        
        return ''
    
    def _wait_for_report_loaded(self, max_wait: int = 60) -> bool:
        """
        Wait for report to load by checking for required identifiers in the page
        
        Args:
            max_wait: Maximum time to wait in seconds (default 60)
            
        Returns:
            Always returns True after max_wait seconds (will create Excel with whatever data is found)
        """
        start_time = time.time()
        check_interval = 0.75  # Check every 0.75 second
        
        # Required identifiers to check for (for logging purposes)
        required_identifiers = [
            "Kredit",
            "DPK", 
            "Total Aset",
            "Laba Kotor",
            "LABA (RUGI) TAHUN BERJALAN SEBELUM PAJAK PENGHASILAN",
            "Rasio"
        ]
        
        found_identifiers = set()
        
        while time.time() - start_time < max_wait:
            try:
                # Check main page source for report content
                self.driver.switch_to.default_content()
                page_source = self.driver.page_source
                page_source_lower = page_source.lower()
                
                # Check for all required identifiers (for logging)
                for identifier in required_identifiers:
                    if identifier not in found_identifiers:
                        if identifier in page_source or identifier.lower() in page_source_lower:
                            found_identifiers.add(identifier)
                            print(f"    [DEBUG] Found identifier: '{identifier}' ({len(found_identifiers)}/{len(required_identifiers)})")
                
                # Wait before next check
                elapsed = int(time.time() - start_time)
                remaining = max_wait - elapsed
                if remaining > 0 and remaining % 10 == 0:  # Print every 10 seconds
                    print(f"    [DEBUG] Waiting... ({remaining}s remaining, found {len(found_identifiers)}/{len(required_identifiers)} identifiers)")
                
                time.sleep(check_interval)
                
            except Exception as e:
                # If error occurs, wait and try again
                time.sleep(check_interval)
                continue
        
        # Always return True after waiting - will create Excel with whatever data is found
        print(f"    [DEBUG] Wait completed. Found identifiers: {found_identifiers}")
        return True
    
    def _extract_report_data(self, selected_year: str, city: str = None, bank: str = None, extract_mode: str = 'sheets_1_3', skip_wait_attempts: bool = False, skip_laba_kotor: bool = False, skip_rasio: bool = False) -> dict:
        """
        Extract financial data from the generated report
        
        Args:
            selected_year: The selected year (e.g., "2024")
            city: The selected city (optional, will try to extract from page if not provided)
            bank: The selected bank (optional, will try to extract from page if not provided)
            extract_mode: 'sheets_1_3' to extract ASET/KREDIT/DPK, 'sheets_4_5' to extract LABA KOTOR/RASIO only
            skip_wait_attempts: If True, skip wait attempts (used when period error was handled)
            skip_laba_kotor: If True, skip Laba Kotor extraction (for phase 003)
            skip_rasio: If True, skip Rasio extraction (for phase 002)
            
        Returns:
            Dictionary with extracted data
        """
        try:
            # Try to find report in iframe first, then main page
            print("    [DEBUG] Checking for report in iframes...")
            iframes = self.driver.find_elements(By.TAG_NAME, "iframe")
            page_source = None
            
            # Check iframes for report content
            report_iframe = None
            for iframe in iframes:
                try:
                    self.driver.switch_to.frame(iframe)
                    iframe_source = self.driver.page_source
                    if ("Kredit" in iframe_source or "Aset" in iframe_source or "DPK" in iframe_source or
                        "LABA" in iframe_source or "Rasio" in iframe_source or "KPMM" in iframe_source):
                        print("    [DEBUG] Found report content in iframe")
                        page_source = iframe_source
                        report_iframe = iframe  # Keep reference to the iframe
                        # Stay in iframe context for XPath searches
                        break
                    self.driver.switch_to.default_content()
                except:
                    self.driver.switch_to.default_content()
                    continue

            # If not in iframe, use main page
            if page_source is None:
                print("    [DEBUG] Using main page source")
                self.driver.switch_to.default_content()
                page_source = self.driver.page_source
            # else: Stay in iframe context - don't switch back yet
            
            # Identifiers to check for (at least one should be present for validation)
            # Include both Sheets 1-3 identifiers and Sheets 4-5 identifiers
            identifiers_to_check = [
                "Kepada BPR",
                "Kepada Bank Umum",
                "pihak terkait",
                "pihak tidak terkait",
                "Total Aset",
                "Tabungan",
                "Deposito",
                "LABA (RUGI) TAHUN BERJALAN SEBELUM PAJAK PENGHASILAN",
                "Kewajiban Penyediaan Modal Minimum",
                "Rasio Cadangan terhadap PPKA",
                "Non Performing Loan"
            ]
            
            def check_identifiers_in_soup(soup: BeautifulSoup, identifiers: list, extract_mode: str = 'sheets_1_3') -> tuple[bool, str]:
                """
                Check if identifiers exist in BeautifulSoup and have valid data records
                Simplified: just check if identifier exists and has next divs (data)
                
                For sheets_4_5 mode: Must find BOTH Laba Kotor and Rasio identifiers
                
                Args:
                    soup: BeautifulSoup object to check
                    identifiers: List of identifier strings to check for
                    extract_mode: 'sheets_1_3' or 'sheets_4_5'
                    
                Returns:
                    Tuple of (found: bool, identifier_name: str)
                """
                if extract_mode == 'sheets_4_5':
                    # For sheets_4_5, we need BOTH Laba Kotor and Rasio identifiers
                    laba_kotor_identifier = "LABA (RUGI) TAHUN BERJALAN SEBELUM PAJAK PENGHASILAN"
                    rasio_identifiers = [
                        "Kewajiban Penyediaan Modal Minimum",
                        "Rasio Cadangan terhadap PPKA",
                        "Non Performing Loan"
                    ]
                    
                    laba_kotor_found = False
                    rasio_found = False
                    found_rasio_identifier = ""
                    
                    # Check for Laba Kotor identifier
                    for div in soup.find_all('div'):
                        text = div.get_text(strip=True)
                        if not text or len(text) > 5000:
                            continue
                        
                        if laba_kotor_identifier.lower() in text.lower():
                            text_lower = text.lower()
                            identifier_lower = laba_kotor_identifier.lower()
                            
                            if len(text) < 200 or text_lower.startswith(identifier_lower) or text_lower.endswith(identifier_lower):
                                all_divs = soup.find_all('div')
                                try:
                                    div_index = all_divs.index(div)
                                    if div_index < len(all_divs) - 1:
                                        laba_kotor_found = True
                                        break
                                except ValueError:
                                    pass
                    
                    # Check for Rasio identifier
                    for rasio_id in rasio_identifiers:
                        for div in soup.find_all('div'):
                            text = div.get_text(strip=True)
                            if not text or len(text) > 5000:
                                continue
                            
                            if rasio_id.lower() in text.lower():
                                text_lower = text.lower()
                                identifier_lower = rasio_id.lower()
                                
                                if len(text) < 200 or text_lower.startswith(identifier_lower) or text_lower.endswith(identifier_lower):
                                    all_divs = soup.find_all('div')
                                    try:
                                        div_index = all_divs.index(div)
                                        if div_index < len(all_divs) - 1:
                                            rasio_found = True
                                            found_rasio_identifier = rasio_id
                                            break
                                    except ValueError:
                                        pass
                        if rasio_found:
                            break
                    
                    # Both must be found for sheets_4_5
                    if laba_kotor_found and rasio_found:
                        return True, f"Laba Kotor & {found_rasio_identifier}"
                    elif laba_kotor_found:
                        return False, "Laba Kotor (menunggu Rasio)"
                    elif rasio_found:
                        return False, f"{found_rasio_identifier} (menunggu Laba Kotor)"
                    else:
                        return False, ""
                else:
                    # For sheets_1_3, check any identifier
                    for identifier in identifiers:
                        # Find the <div> whose text contains our identifier
                        # Be more specific: skip very long divs (entire page content)
                        for div in soup.find_all('div'):
                            text = div.get_text(strip=True)
                            if not text:
                                continue
                            
                            # Skip divs that are too long (likely contain entire page content)
                            if len(text) > 5000:
                                continue
                            
                            if identifier.lower() in text.lower():
                                # Prefer divs where identifier is a significant part of the text
                                text_lower = text.lower()
                                identifier_lower = identifier.lower()
                                
                                # If text is short or identifier is at the start/end, it's likely the right div
                                if len(text) < 200 or text_lower.startswith(identifier_lower) or text_lower.endswith(identifier_lower):
                                    # Check if there are divs after this one (indicating data is present)
                                    all_divs = soup.find_all('div')
                                    try:
                                        div_index = all_divs.index(div)
                                        # If there are more divs after this identifier div, data is likely present
                                        if div_index < len(all_divs) - 1:
                                            return True, identifier
                                    except ValueError:
                                        # div not in all_divs (shouldn't happen, but handle it)
                                        pass
                                    break  # Found identifier, no need to continue searching
                    return False, ""
            
            def refresh_page_source(report_iframe_ref) -> tuple[str, object]:
                """
                Refresh page source from iframe or main page
                
                Args:
                    report_iframe_ref: Reference to the iframe element (or None)l
                    
                Returns:
                    Tuple of (page_source: str, updated_iframe_ref: object)
                """
                if report_iframe_ref:
                    # We're in iframe context, refresh iframe source
                    try:
                        self.driver.switch_to.frame(report_iframe_ref)
                        new_page_source = self.driver.page_source
                        return new_page_source, report_iframe_ref
                    except:
                        # Iframe might have changed, find it again
                        self.driver.switch_to.default_content()
                        iframes = self.driver.find_elements(By.TAG_NAME, "iframe")
                        for iframe in iframes:
                            try:
                                self.driver.switch_to.frame(iframe)
                                iframe_source = self.driver.page_source
                                if ("Kredit" in iframe_source or "Aset" in iframe_source or "DPK" in iframe_source or
                                    "LABA" in iframe_source or "Rasio" in iframe_source or "KPMM" in iframe_source):
                                    return iframe_source, iframe
                                self.driver.switch_to.default_content()
                            except:
                                self.driver.switch_to.default_content()
                                continue
                        # Fallback to main page if iframe not found
                        self.driver.switch_to.default_content()
                        return self.driver.page_source, None
                else:
                    # Using main page, refresh main page source
                    self.driver.switch_to.default_content()
                    return self.driver.page_source, None
            
            # Validation: Wait for page to fully load by checking if records are found by identifiers
            # Skip wait attempts if period error was found and handled
            if skip_wait_attempts:
                print("    [INFO] Skipping wait attempts (period error was handled, using updated period)...")
                max_wait_attempts = 2  # Just one quick check
                wait_interval = 2.5
            elif extract_mode == 'sheets_4_5':
                print("    [INFO] Memvalidasi halaman telah dimuat sepenuhnya (memeriksa identifier setiap 12 detik untuk Sheets 4-5)...")
                max_wait_attempts = 30  # Maximum 30 attempts = 240 seconds (20 * 12s = 240s)
                wait_interval = 10  # Wait 12 seconds between checks (for 2 files)
            else:
                print("    [INFO] Memvalidasi halaman telah dimuat sepenuhnya (memeriksa identifier setiap 10 detik)...")
                max_wait_attempts = 30  # Maximum 30 attempts = 200 seconds (20 * 10s = 200s)
                wait_interval = 10  # Wait 10 seconds between checks
            page_fully_loaded = False   
            soup = None
            
            for attempt in range(max_wait_attempts):
                # Always refresh page_source and re-parse BeautifulSoup to get latest content
                print(f"    [INFO] Percobaan {attempt + 1}/{max_wait_attempts}: Memperbarui page_source dan mem-parse ulang BeautifulSoup...")
                page_source, report_iframe = refresh_page_source(report_iframe)
                
                # Re-parse BeautifulSoup with fresh page_source
                soup = BeautifulSoup(page_source, 'html.parser')
                print(f"    [DEBUG] BeautifulSoup telah di-parse ulang (ukuran: {len(page_source)} karakter)")
                
                # Check for "Bad Request" error
                bad_request_found = False
                page_source_lower = page_source.lower()
                if 'bad request' in page_source_lower:
                    bad_request_found = True
                    print(f"    [WARNING] 'Bad Request' ditemukan dalam halaman (percobaan {attempt + 1})")
                
                # Check if identifiers exist in the parsed BeautifulSoup
                record_found, found_identifier = check_identifiers_in_soup(soup, identifiers_to_check, extract_mode)
                
                if bad_request_found:
                    # Bad Request found, will be handled by retry logic
                    print(f"    [WARNING] Bad Request terdeteksi, akan dicoba ulang...")
                    break  # Exit loop to trigger retry
                elif record_found:
                    page_fully_loaded = True
                    if extract_mode == 'sheets_4_5':
                        print(f"    [OK] Halaman telah dimuat sepenuhnya - Kedua identifier ditemukan: '{found_identifier}' (percobaan {attempt + 1})")
                    else:
                        print(f"    [OK] Halaman telah dimuat sepenuhnya - Identifier '{found_identifier}' ditemukan dengan data (percobaan {attempt + 1})")
                    break
                else:
                    if extract_mode == 'sheets_4_5' and found_identifier:
                        print(f"    [INFO] Halaman belum sepenuhnya dimuat - {found_identifier} (percobaan {attempt + 1}/{max_wait_attempts})")
                    else:
                        print(f"    [INFO] Halaman belum sepenuhnya dimuat - Identifier tidak ditemukan (percobaan {attempt + 1}/{max_wait_attempts})")
                    if attempt < max_wait_attempts - 1:  # Don't wait on last attempt
                        print(f"    [INFO] Menunggu {wait_interval} detik sebelum memeriksa lagi...")
                        time.sleep(wait_interval)
                    else:
                        print(f"    [WARNING] Mencapai batas maksimum percobaan ({max_wait_attempts})")
            
            # Check for Bad Request after wait loop
            if not page_fully_loaded and not bad_request_found:
                # Final check for Bad Request
                page_source, report_iframe = refresh_page_source(report_iframe)
                if 'bad request' in page_source.lower():
                    bad_request_found = True
                    print(f"    [WARNING] 'Bad Request' ditemukan setelah menunggu")
            
            # If Bad Request found, return None to signal retry needed
            if bad_request_found:
                return None
            
            if not page_fully_loaded:
                print("    [WARNING] Halaman mungkin belum sepenuhnya dimuat - Identifier tidak ditemukan setelah waktu tunggu maksimum")
                print("    [WARNING] Melanjutkan ekstraksi dengan data yang tersedia...")
                # Ensure soup is parsed with latest page_source
                if soup is None:
                    page_source, report_iframe = refresh_page_source(report_iframe)
                soup = BeautifulSoup(page_source, 'html.parser')
            
            # Helper function to split concatenated numbers (current year + previous year)
            def split_concatenated_numbers(text: str) -> tuple[str, str]:
                """
                Split concatenated numbers like "23,122,1223,112,122" into two numbers.
                Format: After comma, max 3 digits. If next is comma, same year. If next is digit, next year starts.
                
                Example: "23,122,1223,112,122" -> ("23,122,122", "3,112,122")
                
                Returns:
                    Tuple of (current_year_text, previous_year_text)
                """
                if not text or ',' not in text:
                    return text, ""
                
                import re
                
                # Find the split point: look for pattern where after comma, we have 3 digits, then a digit (not comma)
                # Pattern: comma, then exactly 3 digits, then a digit (not comma)
                # This indicates the start of the next year
                # Example: "23,122,1223" -> split after "122" (3 digits), before "3" (digit)
                pattern = r',(\d{3})(\d)(?=,|\d|$)'
                match = re.search(pattern, text)
                
                if match:
                    # Found split point: the digit after the 3-digit group starts the next year
                    # Split position is right before that digit
                    split_pos = match.end(1)  # Position after the 3 digits (before the next digit)
                    current_year_text = text[:split_pos].rstrip(',')
                    previous_year_text = text[split_pos:]
                    
                    # Remove leading comma from previous year if any
                    previous_year_text = previous_year_text.lstrip(',')
                    
                    print(f"      [DEBUG] Split '{text}' -> Current: '{current_year_text}', Previous: '{previous_year_text}'")
                    return current_year_text, previous_year_text
                
                # No split found, return original text as current year
                return text, ""
            
            # Helper function to extract numeric value from text
            def extract_number(text: str) -> float:
                """Extract an integer-like number from Indonesian-style formatted text."""
                if not text:
                    return 0.0
                
                original_text = text
                # Normalize spaces
                text = text.replace('\xa0', ' ').replace('&nbsp;', ' ').strip()
                
                # Keep digits only (we just need whole numbers)
                digits_only = re.sub(r'\D', '', text)
                if not digits_only:
                    print(f"      [DEBUG]     Failed to extract number from: '{original_text}'")
                    return 0.0
                
                value = float(digits_only)
                if original_text.strip() != digits_only:
                    print(f"      [DEBUG]     Raw: '{original_text}' -> Digits: '{digits_only}' -> Number: {value}")
                return value
            
            # Helper function to find identifier and get next 2 div values using BeautifulSoup
            def find_and_extract(identifier: str):
                """
                Find the div containing `identifier` and return the next 2 div values
                (current year, previous year) - simplified approach using inline divs
                """
                values = []
                
                # 1. Find the <div> whose text contains our identifier
                # Be more specific: look for divs where identifier is the main/only text, not in huge text blocks
                label_div = None
                all_divs = soup.find_all('div')
                
                for div in all_divs:
                    text = div.get_text(strip=True)
                    if not text:
                        continue
                    
                    # Skip divs that are too long (likely contain entire page content)
                    if len(text) > 5000:  # Skip very long divs (entire page content)
                        continue
                    
                    # Check if identifier is in text
                    if identifier.lower() in text.lower():
                        # Prefer divs where identifier is a significant part of the text
                        # or where text is relatively short (more specific match)
                        text_lower = text.lower()
                        identifier_lower = identifier.lower()
                        
                        # If text is short or identifier is at the start/end, it's likely the right div
                        if len(text) < 200 or text_lower.startswith(identifier_lower) or text_lower.endswith(identifier_lower):
                            label_div = div
                            print(f"    [DEBUG] Found identifier '{identifier}' in <div>: '{text[:100]}...' (length: {len(text)})")
                            break
                        # Otherwise, continue searching for a better match
                
                if not label_div:
                    print(f"    [DEBUG] Identifier '{identifier}' NOT FOUND in page")
                    return values
                
                # 2. Find the index of this div and get the next divs that contain numeric values
                try:
                    div_index = all_divs.index(label_div)
                    # Look through next divs to find numeric values (up to 10 divs ahead)
                    numeric_count = 0
                    for j in range(1, min(11, len(all_divs) - div_index)):  # Check up to 10 next divs
                        if numeric_count >= 2:  # We need 2 values
                            break
                            
                        if div_index + j < len(all_divs):
                            next_div = all_divs[div_index + j]
                            div_text = next_div.get_text(strip=True)
                            
                            if not div_text:
                                continue
                            
                            # Skip very long divs (likely contain entire page content)
                            if len(div_text) > 5000:
                                continue
                            
                            # Skip if it's clearly another identifier (contains common identifier keywords)
                            if any(keyword in div_text.lower() for keyword in ['kepada', 'pihak', 'bank', 'bpr', 'report viewer', 'configuration error']):
                                # This is likely another identifier or error message, skip it
                                continue
                            
                            # Check if this div contains numbers (might be formatted like "1.234.567" or "1,234,567")
                            # It might also contain concatenated numbers like "23,122,1223,112,122"
                            
                            # Check if it contains concatenated numbers (has comma and might have split pattern)
                            if ',' in div_text and len(div_text) > 10:
                                # Try to split concatenated numbers
                                current_year_text, prev_year_text = split_concatenated_numbers(div_text)
                                
                                if prev_year_text:
                                    # Found concatenated numbers, extract both
                                    current_number = extract_number(current_year_text)
                                    prev_number = extract_number(prev_year_text)
                                    
                                    # Validate numbers are reasonable
                                    if (current_number > 0 and current_number < 1e15 and current_number != float('inf') and
                                        prev_number >= 0 and prev_number < 1e15 and prev_number != float('inf')):
                                        
                                        if numeric_count == 0:
                                            # Add current year first
                                            print(f"    [DEBUG]   Next div[{j}] (Year {selected_year}): "
                                                  f"'{div_text}' -> Split: '{current_year_text}' = {current_number}")
                                            values.append(current_number)
                                            numeric_count += 1
                                        
                                        if numeric_count == 1:
                                            # Add previous year
                                            print(f"    [DEBUG]   Next div[{j}] (Year {previous_year}): "
                                                  f"'{div_text}' -> Split: '{prev_year_text}' = {prev_number}")
                                            values.append(prev_number)
                                            numeric_count += 1
                                        
                                        # Got both values, break
                                        if numeric_count >= 2:
                                            break
                                        continue
                            
                            # Check if this div looks like it contains a single formatted number
                            # It should be relatively short and contain digits with possible formatting
                            if len(div_text) > 100:
                                # Too long, probably not a single number (unless it's concatenated, which we handled above)
                                continue
                            
                            # Extract number from single number text
                            number = extract_number(div_text)
                            
                            # Validate the number is reasonable (not infinity and not too large)
                            # Indonesian Rupiah values are typically in millions/billions, so cap at 1e15
                            if number > 0 and number < 1e15 and number != float('inf'):
                                year_label = selected_year if numeric_count == 0 else previous_year
                                print(f"    [DEBUG]   Next div[{j}] (Year {year_label}): "
                                      f"'{div_text}' -> {number}")
                                values.append(number)
                                numeric_count += 1
                            elif number == 0 and any(char.isdigit() for char in div_text) and len(div_text) < 50:
                                # Zero value is valid if it's a short text with digits
                                year_label = selected_year if numeric_count == 0 else previous_year
                                print(f"    [DEBUG]   Next div[{j}] (Year {year_label}): "
                                      f"'{div_text}' -> {number}")
                                values.append(number)
                                numeric_count += 1
                except ValueError:
                    print(f"    [DEBUG] Identifier '{identifier}' found but couldn't find its index")
                    return values
                
                if not values:
                    print(f"    [DEBUG] Identifier '{identifier}' found but no numeric values extracted from next divs")
                elif len(values) == 1:
                    print(f"    [DEBUG] Identifier '{identifier}' found but only 1 value extracted")
                
                return values
            
            result = {}
            previous_year = str(int(selected_year) - 1)
            
            # Use provided city and bank, or extract from page source
            print("    [INFO] Extracting Kota/Kabupaten and Bank...")
            extracted_city = city if city else ""
            extracted_bank = bank if bank else ""
            
            # For sheets_4_5 mode, always use provided city/bank - don't extract from page to avoid getting month name
            if extract_mode == 'sheets_4_5':
                # Always use provided city and bank for Laba Kotor/Rasio
                extracted_city = city if city else "N/A"
                extracted_bank = bank if bank else "N/A"
            # Only try to extract from page if not provided (for sheets_1_3 mode)
            elif not extracted_city or not extracted_bank:
                try:
                    # Switch to default content to access form fields
                    self.driver.switch_to.default_content()
                    
                    # Try to find city and bank from input fields using Selenium
                    if not extracted_city:
                        try:
                            # Look for city input field
                            city_inputs = [
                                self.driver.find_elements(By.XPATH, "//input[contains(@id, 'City') or contains(@id, 'Kota') or contains(@id, 'Kabupaten')]"),
                                self.driver.find_elements(By.XPATH, "//input[contains(@name, 'City') or contains(@name, 'Kota') or contains(@name, 'Kabupaten')]")
                            ]
                            for inputs in city_inputs:
                                for inp in inputs:
                                    try:
                                        value = inp.get_attribute('value')
                                        if value and value.strip():
                                            # List of month names to reject (to avoid getting month name as city)
                                            month_names = ['januari', 'februari', 'maret', 'april', 'mei', 'juni', 
                                                         'juli', 'agustus', 'september', 'oktober', 'november', 'desember']
                                            value_lower = value.strip().lower()
                                            # Reject if it's a month name
                                            if value_lower in month_names:
                                                print(f"    [DEBUG] Rejected month name as city: '{value}'")
                                                continue
                                            extracted_city = value.strip()
                                            print(f"    [DEBUG] Found city from input field: '{extracted_city}'")
                                            break
                                    except:
                                        continue
                                if extracted_city:
                                    break
                        except Exception as e:
                            print(f"    [DEBUG] Could not find city input: {e}")
                    
                    if not extracted_bank:
                        try:
                            # Look for bank input field
                            bank_inputs = [
                                self.driver.find_elements(By.XPATH, "//input[contains(@id, 'Bank')]"),
                                self.driver.find_elements(By.XPATH, "//input[contains(@name, 'Bank')]")
                            ]
                            for inputs in bank_inputs:
                                for inp in inputs:
                                    try:
                                        value = inp.get_attribute('value')
                                        if value and value.strip():
                                            extracted_bank = value.strip()
                                            print(f"    [DEBUG] Found bank from input field: '{extracted_bank}'")
                                            break
                                    except:
                                        continue
                                if extracted_bank:
                                    break
                        except Exception as e:
                            print(f"    [DEBUG] Could not find bank input: {e}")
                    
                    # Fallback: Try to find from BeautifulSoup parsed page
                    if not extracted_city or not extracted_bank:
                        city_inputs = soup.find_all('input', {'id': lambda x: x and ('city' in x.lower() or 'kota' in x.lower() or 'kabupaten' in x.lower())})
                        bank_inputs = soup.find_all('input', {'id': lambda x: x and 'bank' in x.lower()})
                        
                        if not extracted_city:
                            # List of month names to reject (to avoid getting month name as city)
                            month_names = ['januari', 'februari', 'maret', 'april', 'mei', 'juni', 
                                         'juli', 'agustus', 'september', 'oktober', 'november', 'desember']
                            for inp in city_inputs:
                                value = inp.get('value', '')
                                if value and value.strip():
                                    value_lower = value.strip().lower()
                                    # Reject if it's a month name
                                    if value_lower in month_names:
                                        print(f"    [DEBUG] Rejected month name as city: '{value}'")
                                        continue
                                    extracted_city = value.strip()
                                    print(f"    [DEBUG] Found city from soup: '{extracted_city}'")
                                    break
                        
                        if not extracted_bank:
                            for inp in bank_inputs:
                                value = inp.get('value', '')
                                if value and value.strip():
                                    extracted_bank = value.strip()
                                    print(f"    [DEBUG] Found bank from soup: '{extracted_bank}'")
                                    break
                
                except Exception as e:
                    print(f"    [WARNING] Could not extract city/bank: {e}")
            
            result['city'] = extracted_city if extracted_city else "N/A"
            result['bank'] = extracted_bank if extracted_bank else "N/A"
            
            # Only extract ASET/KREDIT/DPK if in sheets_1_3 mode
            if extract_mode == 'sheets_1_3':
                # Extract Kredit data only - Sum of 4 identifiers
                print("    [INFO] Extracting Kredit data...")
                # Only extract: Kepada BPR, Kepada Bank Umum, pihak terkait (first occurrence), pihak tidak terkait (first occurrence)
                kredit_identifiers = [
                    "Kepada BPR",
                    "Kepada Bank Umum",
                    "pihak terkait",
                    "pihak tidak terkait"
                ]
                kredit_selected_year = 0
                kredit_previous_year = 0
                found_identifiers = set()  # Track which identifiers were found
                
                for identifier in kredit_identifiers:
                    values = find_and_extract(identifier)
                    if len(values) >= 2:
                        # Check if we already found this identifier (avoid duplicates - only take first occurrence)
                        identifier_key = identifier.strip().lower()
                        if identifier_key not in found_identifiers:
                            # First div = current year, second div = previous year
                            kredit_selected_year += values[0]  # First div (current year)
                            kredit_previous_year += values[1]  # Second div (previous year)
                            found_identifiers.add(identifier_key)
                            print(f"    [DEBUG] Added Kredit from '{identifier}': {values[0]} (2024) + {values[1]} (2023)")
                    elif len(values) == 1:
                        identifier_key = identifier.strip().lower()
                        if identifier_key not in found_identifiers:
                            kredit_selected_year += values[0]  # Only current year available
                            found_identifiers.add(identifier_key)
                            print(f"    [DEBUG] Added Kredit from '{identifier}': {values[0]} (2024 only)")
                
                result[f'Kredit {selected_year}'] = kredit_selected_year
                result[f'Kredit {previous_year}'] = kredit_previous_year
                
                print(f"    [OK] Extracted Kredit data: {selected_year}={kredit_selected_year}, {previous_year}={kredit_previous_year}")
                
                # Extract Total Aset data - only one div per year (no sum needed)
                print("    [INFO] Extracting Total Aset data...")
                total_aset_identifier = "Total Aset"
                total_aset_values = find_and_extract(total_aset_identifier)
                
                if len(total_aset_values) >= 2:
                    result[f'Total Aset {selected_year}'] = total_aset_values[0]
                    result[f'Total Aset {previous_year}'] = total_aset_values[1]
                    print(f"    [OK] Extracted Total Aset data: {selected_year}={total_aset_values[0]}, {previous_year}={total_aset_values[1]}")
                elif len(total_aset_values) == 1:
                    result[f'Total Aset {selected_year}'] = total_aset_values[0]
                    result[f'Total Aset {previous_year}'] = 0
                    print(f"    [OK] Extracted Total Aset data: {selected_year}={total_aset_values[0]}, {previous_year}=0 (only current year found)")
                else:
                    result[f'Total Aset {selected_year}'] = 0
                    result[f'Total Aset {previous_year}'] = 0
                    print(f"    [WARNING] Total Aset data not found")
                
                # Extract DPK data - Sum of Tabungan, Deposito, and Simpanan dari Bank Lain for each year
                print("    [INFO] Extracting DPK data (Tabungan + Deposito + Simpanan dari Bank Lain)...")
                dpk_identifiers = [
                    "Tabungan",
                    "Deposito",
                    "Simpanan dari Bank Lain"
                ]
                dpk_selected_year = 0
                dpk_previous_year = 0
                found_dpk_identifiers = set()
                
                for identifier in dpk_identifiers:
                    values = find_and_extract(identifier)
                    if len(values) >= 2:
                        identifier_key = identifier.strip().lower()
                        if identifier_key not in found_dpk_identifiers:
                            dpk_selected_year += values[0]  # First div (current year)
                            dpk_previous_year += values[1]  # Second div (previous year)
                            found_dpk_identifiers.add(identifier_key)
                            print(f"    [DEBUG] Added DPK from '{identifier}': {values[0]} (2024) + {values[1]} (2023)")
                    elif len(values) == 1:
                        identifier_key = identifier.strip().lower()
                        if identifier_key not in found_dpk_identifiers:
                            dpk_selected_year += values[0]  # Only current year available
                            found_dpk_identifiers.add(identifier_key)
                            print(f"    [DEBUG] Added DPK from '{identifier}': {values[0]} (2024 only)")
                
                result[f'DPK {selected_year}'] = dpk_selected_year
                result[f'DPK {previous_year}'] = dpk_previous_year
                
                print(f"    [OK] Extracted DPK data: {selected_year}={dpk_selected_year}, {previous_year}={dpk_previous_year}")
                
                # In sheets_1_3 mode, explicitly set Laba Kotor and Rasio to 0 (not extracted)
                result[f'Laba Kotor {selected_year}'] = 0
                result[f'Laba Kotor {previous_year}'] = 0
                ratio_names = ['KPMM', 'PPKA', 'NPL Neto', 'NPL Gross', 'ROA', 'BOPO', 'NIM', 'LDR', 'CR']
                for ratio_name in ratio_names:
                    result[ratio_name] = 0
                print("    [INFO] Skipping Laba Kotor/Rasio extraction (sheets_1_3 mode - only extracting ASET/KREDIT/DPK)")
            else:
                # In sheets_4_5 mode, skip ASET/KREDIT/DPK extraction
                print("    [INFO] Skipping ASET/KREDIT/DPK extraction (sheets_4_5 mode)")
                result[f'Kredit {selected_year}'] = 0
                result[f'Kredit {previous_year}'] = 0
                result[f'Total Aset {selected_year}'] = 0
                result[f'Total Aset {previous_year}'] = 0
                result[f'DPK {selected_year}'] = 0
                result[f'DPK {previous_year}'] = 0
            
            # Extract Laba Kotor and Ratios (for Sheets 4-5) - only if in sheets_4_5 mode
            if extract_mode == 'sheets_4_5':
                # Check if Laba Kotor/Rasio identifiers exist in the page
                laba_kotor_identifier = "LABA (RUGI) TAHUN BERJALAN SEBELUM PAJAK PENGHASILAN"
                ratio_identifiers_check = [
                    "Kewajiban Penyediaan Modal Minimum (KPMM)",
                    "Rasio Cadangan terhadap PPKA",
                    "Non Performing Loan"
                ]
                
                # Check if any ratio identifiers exist
                has_ratio_data = False
                for div in soup.find_all('div'):
                    text = div.get_text(strip=True)
                    if laba_kotor_identifier.upper() in text.upper() or any(rid.upper() in text.upper() for rid in ratio_identifiers_check):
                        has_ratio_data = True
                        break
                
                if has_ratio_data:
                    print("    [INFO] Detected Laba Kotor/Rasio data in report, extracting...")
                    
                    def extract_laba_kotor_value(identifier_text: str) -> tuple[float, float]:
                        """
                        Extract Laba Kotor value from table structure.
                        Structure: identifier in <td><div>, values in subsequent <td><div> elements.
                        Values can be negative in parentheses like (677,555,231).
                        Returns tuple of (current_year_value, previous_year_value)
                        """
                        import re
                        
                        for div in soup.find_all('div'):
                            text = div.get_text(strip=True)
                            if identifier_text.upper() in text.upper() and len(text) < 5000:
                                # Find the parent <td> or <tr> (table row)
                                parent_td = div.find_parent('td')
                                if not parent_td:
                                    continue
                                
                                # Find the parent <tr> (table row) to get all <td> elements in the row
                                parent_tr = parent_td.find_parent('tr')
                                if not parent_tr:
                                    # If no <tr>, try to find subsequent <td> elements from the same parent
                                    parent_td_siblings = parent_td.find_next_siblings('td')
                                    if not parent_td_siblings:
                                        continue
                                    tds = [parent_td] + list(parent_td_siblings)
                                else:
                                    # Get all <td> elements in the row
                                    tds = parent_tr.find_all('td')
                                
                                # Find the index of the identifier <td>
                                try:
                                    identifier_td_index = tds.index(parent_td)
                                except:
                                    continue
                                
                                # Extract values from subsequent <td> elements (skip the identifier <td>)
                                # Look for numeric values in <td> elements (may have multiple <td> elements to check)
                                values = []
                                import re
                                for td in tds[identifier_td_index + 1:identifier_td_index + 10]:  # Check more <td> elements to find numeric values
                                    # Look for <div> inside the <td>
                                    td_div = td.find('div')
                                    if td_div:
                                        td_text = td_div.get_text(strip=True)
                                        if td_text:
                                            # Check if this looks like a number (contains digits, possibly with parentheses and commas)
                                            # Pattern: optional parentheses, digits with commas, optional decimal part
                                            # Examples: "(677,555,231)", "1,223", "123.45"
                                            numeric_pattern = r'^\(?[\d,]+\.?\d*\)?$'
                                            if re.match(numeric_pattern, td_text.replace(' ', '')):
                                                values.append(td_text)
                                                if len(values) >= 2:  # Stop after finding 2 numeric values
                                                    break
                                
                                # Parse the values
                                if len(values) >= 2:
                                    # Two values: first is previous year, second is current year
                                    try:
                                        # Parse first value (previous year)
                                        val1_text = values[0].strip()
                                        is_negative1 = val1_text.startswith('(') and val1_text.endswith(')')
                                        if is_negative1:
                                            val1_text = val1_text[1:-1].strip()  # Remove parentheses
                                        # Remove commas (thousands separators) and convert to float
                                        val1_text = val1_text.replace(',', '')
                                        val1 = float(val1_text)
                                        if is_negative1:
                                            val1 = -val1
                                        
                                        # Parse second value (current year)
                                        val2_text = values[1].strip()
                                        is_negative2 = val2_text.startswith('(') and val2_text.endswith(')')
                                        if is_negative2:
                                            val2_text = val2_text[1:-1].strip()  # Remove parentheses
                                        # Remove commas (thousands separators) and convert to float
                                        val2_text = val2_text.replace(',', '')
                                        val2 = float(val2_text)
                                        if is_negative2:
                                            val2 = -val2
                                        
                                        # Return as (current_year, previous_year)
                                        return (val2, val1)
                                    except Exception as e:
                                        print(f"    [DEBUG] Error parsing two values: {e}, values: {values}")
                                        pass
                                elif len(values) == 1:
                                    # Single value: use for both years
                                    try:
                                        val_text = values[0].strip()
                                        is_negative = val_text.startswith('(') and val_text.endswith(')')
                                        if is_negative:
                                            val_text = val_text[1:-1].strip()  # Remove parentheses
                                        # Remove commas (thousands separators) and convert to float
                                        val_text = val_text.replace(',', '')
                                        val = float(val_text)
                                        if is_negative:
                                            val = -val
                                        # Use same value for both years
                                        return (val, val)
                                    except Exception as e:
                                        print(f"    [DEBUG] Error parsing single value: {e}, value: {values[0]}")
                                        pass
                        
                        return (0.0, 0.0)
                    
                    def extract_ratio_value(identifier_text: str) -> float:
                        """
                        Extract Rasio value from table structure.
                        After finding identifier in <td>, check each sibling <td>:
                        - If <td> has no <div> child, skip to next <td>
                        - If <td> has <div> child, extract text and check if it's a number with decimal point
                        """
                        import re
                        print(f"    [DEBUG] Searching for ratio identifier: '{identifier_text}'")
                        found_identifier = False
                        for div in soup.find_all('div'):
                            text = div.get_text(strip=True)
                            if identifier_text.upper() in text.upper() and len(text) < 5000:
                                found_identifier = True
                                # Find the parent <td> element
                                parent_td = div.find_parent('td')
                                if not parent_td:
                                    continue
                                
                                # Find the parent <tr> (table row) to get all <td> elements
                                parent_tr = parent_td.find_parent('tr')
                                if not parent_tr:
                                    continue
                                
                                # Get all <td> elements in the row
                                tds = parent_tr.find_all('td')
                                
                                # Find the index of the identifier <td>
                                try:
                                    identifier_td_index = tds.index(parent_td)
                                except:
                                    continue
                                
                                print(f"    [DEBUG] Found identifier '{identifier_text}' at td index {identifier_td_index}, checking next tds...")
                                
                                # Check each sibling <td> after the identifier
                                for td_idx, td in enumerate(tds[identifier_td_index + 1:identifier_td_index + 30], start=identifier_td_index + 1):  # Check up to 30 <td> elements
                                    # Check if this <td> has a <div> child
                                    td_div = td.find('div', recursive=False)  # Only direct child, not nested
                                    if not td_div:
                                        # Try to find any div (including nested)
                                        td_div = td.find('div')
                                    
                                    if not td_div:
                                        # No div found in this td, skip to next
                                        print(f"    [DEBUG]   td[{td_idx}]: no div child, skipping")
                                        continue
                                    
                                    # Get text from the <div>
                                    div_text = td_div.get_text(strip=True)
                                    
                                    # Remove &nbsp; entities and check if empty
                                    div_text_clean = div_text.replace('\xa0', ' ').replace('&nbsp;', ' ').strip()
                                    
                                    print(f"    [DEBUG]   td[{td_idx}]: found div with text='{div_text_clean[:50]}'")
                                    
                                    # Skip if empty or only whitespace
                                    if not div_text_clean or div_text_clean == '':
                                        print(f"    [DEBUG]   td[{td_idx}]: div text is empty, skipping")
                                        continue
                                    
                                    # Try to extract number - check if it's a valid numeric text
                                    # Remove spaces and try to parse
                                    cleaned_text = div_text_clean.replace(' ', '').replace(',', '.')
                                    
                                    # Check for negative (in parentheses)
                                    is_negative = False
                                    if cleaned_text.startswith('(') and cleaned_text.endswith(')'):
                                        is_negative = True
                                        cleaned_text = cleaned_text[1:-1].strip()
                                    elif cleaned_text.startswith('-'):
                                        is_negative = True
                                        cleaned_text = cleaned_text[1:].strip()
                                    
                                    # Try to parse as float - if it's a valid number like 0.33 or 3.25
                                    try:
                                        number = float(cleaned_text)
                                        if is_negative:
                                            number = -number
                                        # Reasonable range check
                                        if abs(number) < 1e15:
                                            print(f"    [DEBUG] Found {identifier_text} ratio value at td index {td_idx}: {number}")
                                            return number
                                    except ValueError:
                                        # Not a valid number, skip to next td
                                        print(f"    [DEBUG]   td[{td_idx}]: '{cleaned_text[:30]}' is not a valid number, skipping")
                                        pass
                                
                                print(f"    [DEBUG] No ratio value found for '{identifier_text}' after checking {min(30, len(tds) - identifier_td_index - 1)} tds")
                        if not found_identifier:
                            print(f"    [DEBUG] Identifier '{identifier_text}' not found in any div element")
                        return 0.0
                    
                    # Extract Laba Kotor (for Sheet 4) - skip if skip_laba_kotor is True
                    if not skip_laba_kotor:
                        laba_kotor_current, laba_kotor_previous = extract_laba_kotor_value(laba_kotor_identifier)
                        # Swap years: 2025 should be in 2024 column, 2024 should be in 2025 column
                        result[f'Laba Kotor {selected_year}'] = laba_kotor_previous  # Swap: put previous year value in current year column
                        result[f'Laba Kotor {previous_year}'] = laba_kotor_current   # Swap: put current year value in previous year column
                        print(f"    [OK] Extracted Laba Kotor: {selected_year}={laba_kotor_previous}, {previous_year}={laba_kotor_current} (swapped)")
                    else:
                        print(f"    [INFO] Skipping Laba Kotor extraction (skip_laba_kotor=True)")
                        result[f'Laba Kotor {selected_year}'] = 0
                        result[f'Laba Kotor {previous_year}'] = 0
                    
                    # Extract all 9 ratios (for Sheet 5) - from iframe (1 iframe per checkbox) - skip if skip_rasio is True
                    if not skip_rasio:
                        print("    [INFO] Switching to iframe for ratio extraction...")
                        self.driver.switch_to.default_content()
                        iframes = self.driver.find_elements(By.TAG_NAME, "iframe")
                        ratio_iframe = None
                        ratio_soup = None
                        
                        if iframes:
                            # Use the first (and only) iframe for ratios (1 iframe per checkbox)
                            try:
                                self.driver.switch_to.frame(iframes[0])
                                ratio_page_source = self.driver.page_source
                                ratio_soup = BeautifulSoup(ratio_page_source, 'html.parser')
                                ratio_iframe = iframes[0]
                                print(f"    [OK] Switched to iframe for ratio extraction")
                            except Exception as e:
                                print(f"    [WARNING] Could not switch to iframe: {e}")
                                # Fallback: use current soup
                                ratio_soup = soup
                        else:
                            # No iframes found, use current soup (main page)
                            print(f"    [WARNING] No iframes found, using current page source for ratios.")
                            ratio_soup = soup
                        
                        # Update extract_ratio_value to use ratio_soup instead of soup
                        def extract_ratio_value_from_soup(identifier_text: str, soup_to_use: BeautifulSoup) -> float:
                            """
                            Extract Rasio value from table structure using provided soup.
                            After finding identifier in <td>, check each sibling <td>:
                            - If <td> has no <div> child, skip to next <td>
                            - If <td> has <div> child, extract text and check if it's a number with decimal point
                            """
                            import re
                            print(f"    [DEBUG] Searching for ratio identifier: '{identifier_text}'")
                            found_identifier = False
                            for div in soup_to_use.find_all('div'):
                                text = div.get_text(strip=True)
                                if identifier_text.upper() in text.upper() and len(text) < 5000:
                                    found_identifier = True
                                    # Find the parent <td> element
                                    parent_td = div.find_parent('td')
                                    if not parent_td:
                                        continue
                                    
                                    # Find the parent <tr> (table row) to get all <td> elements
                                    parent_tr = parent_td.find_parent('tr')
                                    if not parent_tr:
                                        continue
                                    
                                    # Get all <td> elements in the row
                                    tds = parent_tr.find_all('td')
                                    
                                    # Find the index of the identifier <td>
                                    try:
                                        identifier_td_index = tds.index(parent_td)
                                    except:
                                        continue
                                    
                                    print(f"    [DEBUG] Found identifier '{identifier_text}' at td index {identifier_td_index}, checking next tds...")
                                    
                                    # Check each sibling <td> after the identifier
                                    for td_idx, td in enumerate(tds[identifier_td_index + 1:identifier_td_index + 30], start=identifier_td_index + 1):  # Check up to 30 <td> elements
                                        # Check if this <td> has a <div> child
                                        td_div = td.find('div', recursive=False)  # Only direct child, not nested
                                        if not td_div:
                                            # Try to find any div (including nested)
                                            td_div = td.find('div')
                                        
                                        if not td_div:
                                            # No div found in this td, skip to next
                                            print(f"    [DEBUG]   td[{td_idx}]: no div child, skipping")
                                            continue
                                        
                                        # Get text from the <div>
                                        div_text = td_div.get_text(strip=True)
                                        
                                        # Remove &nbsp; entities and check if empty
                                        div_text_clean = div_text.replace('\xa0', ' ').replace('&nbsp;', ' ').strip()
                                        
                                        print(f"    [DEBUG]   td[{td_idx}]: found div with text='{div_text_clean[:50]}'")
                                        
                                        # Skip if empty or only whitespace
                                        if not div_text_clean or div_text_clean == '':
                                            print(f"    [DEBUG]   td[{td_idx}]: div text is empty, skipping")
                                            continue
                                        
                                        # Try to extract number - check if it's a valid numeric text
                                        # Remove spaces and try to parse
                                        cleaned_text = div_text_clean.replace(' ', '').replace(',', '.')
                                        
                                        # Check for negative (in parentheses)
                                        is_negative = False
                                        if cleaned_text.startswith('(') and cleaned_text.endswith(')'):
                                            is_negative = True
                                            cleaned_text = cleaned_text[1:-1].strip()
                                        elif cleaned_text.startswith('-'):
                                            is_negative = True
                                            cleaned_text = cleaned_text[1:].strip()
                                        
                                        # Try to parse as float - if it's a valid number like 0.33 or 3.25
                                        try:
                                            number = float(cleaned_text)
                                            if is_negative:
                                                number = -number
                                            # Reasonable range check
                                            if abs(number) < 1e15:
                                                print(f"    [DEBUG] Found {identifier_text} ratio value at td index {td_idx}: {number}")
                                                return number
                                        except ValueError:
                                            # Not a valid number, skip to next td
                                            print(f"    [DEBUG]   td[{td_idx}]: '{cleaned_text[:30]}' is not a valid number, skipping")
                                            pass
                                    
                                    print(f"    [DEBUG] No ratio value found for '{identifier_text}' after checking {min(30, len(tds) - identifier_td_index - 1)} tds")
                            if not found_identifier:
                                print(f"    [DEBUG] Identifier '{identifier_text}' not found in any div element")
                            return 0.0
                        
                        ratio_identifiers = [
                            ("Kewajiban Penyediaan Modal Minimum (KPMM)", "KPMM"),
                            ("Rasio Cadangan terhadap PPKA", "PPKA"),
                            ("Non Performing Loan (NPL) Neto", "NPL Neto"),
                            ("Non Performing Loan (NPL) Gross", "NPL Gross"),
                            ("Return on Assets (ROA)", "ROA"),
                            ("Biaya Operasional terhadap Pendapatan Operasional (BOPO)", "BOPO"),
                            ("Net Interest Margin (NIM)", "NIM"),
                            ("Loan to Deposit Ratio (LDR)", "LDR"),
                            ("Cash Ratio", "CR")
                        ]
                        
                        for identifier, ratio_name in ratio_identifiers:
                            value = extract_ratio_value_from_soup(identifier, ratio_soup)
                            result[ratio_name] = value
                            print(f"    [DEBUG] Extracted {ratio_name}: {value}")
                        
                        # Switch back to first iframe (or default content) after ratio extraction
                        if ratio_iframe:
                            try:
                                self.driver.switch_to.default_content()
                                if report_iframe:
                                    self.driver.switch_to.frame(report_iframe)
                                print("    [DEBUG] Switched back to first iframe after ratio extraction")
                            except:
                                self.driver.switch_to.default_content()
                    else:
                        print(f"    [INFO] Skipping Rasio extraction (skip_rasio=True)")
                        ratio_names = ['KPMM', 'PPKA', 'NPL Neto', 'NPL Gross', 'ROA', 'BOPO', 'NIM', 'LDR', 'CR']
                        for ratio_name in ratio_names:
                            result[ratio_name] = 0
                else:
                    # No ratio data found, set defaults
                    if not skip_laba_kotor:
                        result[f'Laba Kotor {selected_year}'] = 0
                        result[f'Laba Kotor {previous_year}'] = 0
                    if not skip_rasio:
                        ratio_names = ['KPMM', 'PPKA', 'NPL Neto', 'NPL Gross', 'ROA', 'BOPO', 'NIM', 'LDR', 'CR']
                        for ratio_name in ratio_names:
                            result[ratio_name] = 0
            
            print(f"    [OK] Total extracted data points: {len(result)}")
            
            # Switch back to default content after extraction
            try:
                self.driver.switch_to.default_content()
                print("    [DEBUG] Switched back to default content")
            except:
                pass
            
            return result
            
        except Exception as e:
            print(f"    [ERROR] Error extracting report data: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def _extract_laba_kotor_data(self, selected_year: str, city: str = None, bank: str = None) -> dict:
        """Extract Laba Kotor and all 9 Rasio data from the report. Returns dict with all ratio values."""
        from bs4 import BeautifulSoup
        import re
        
        try:
            # Always use provided city and bank - don't extract from page to avoid getting month name
            result = {
                'city': city if city else 'N/A',  # Use provided city, don't extract from page
                'bank': bank if bank else 'N/A'    # Use provided bank, don't extract from page
            }
            
            previous_year = str(int(selected_year) - 1)
            
            # Switch to iframe if report is in iframe
            try:
                iframes = self.driver.find_elements(By.TAG_NAME, "iframe")
                if iframes:
                    self.driver.switch_to.frame(iframes[0])
                    print("    [DEBUG] Switched to iframe for data extraction")
            except:
                pass
            
            # Get page source and parse with BeautifulSoup
            page_source = self.driver.page_source
            soup = BeautifulSoup(page_source, 'html.parser')
            
            # Wait for page to fully load by checking for identifier
            print("    [INFO] Memvalidasi halaman telah dimuat sepenuhnya (memeriksa identifier setiap 10 detik)...")
            max_attempts = 20
            check_interval = 15  # Check every 15 seconds
            
            identifier_found = False
            bad_request_found = False
            for attempt in range(max_attempts):
                # Check for "Bad Request" error
                page_source_lower = page_source.lower()
                if 'bad request' in page_source_lower:
                    bad_request_found = True
                    print(f"    [WARNING] 'Bad Request' ditemukan dalam halaman (percobaan {attempt + 1})")
                    break  # Exit loop if Bad Request found
                
                # Check if identifier exists (check for any ratio identifier)
                identifiers_to_check = [
                    "LABA (RUGI) TAHUN BERJALAN SEBELUM PAJAK PENGHASILAN",
                    "Kewajiban Penyediaan Modal Minimum",
                    "Rasio Cadangan terhadap PPKA",
                    "Non Performing Loan"
                ]
                for identifier in identifiers_to_check:
                    for div in soup.find_all('div'):
                        text = div.get_text(strip=True)
                        if identifier.upper() in text.upper():
                            identifier_found = True
                            print(f"    [OK] Halaman telah dimuat sepenuhnya - Identifier ditemukan (percobaan {attempt + 1})")
                            break
                    if identifier_found:
                        break
                
                if identifier_found:
                    break
                
                if attempt < max_attempts - 1:
                    print(f"    [INFO] Percobaan {attempt + 1}/{max_attempts}: Memperbarui page_source dan mem-parse ulang BeautifulSoup...")
                    time.sleep(check_interval)
                    page_source = self.driver.page_source
                    soup = BeautifulSoup(page_source, 'html.parser')
                    print(f"    [DEBUG] BeautifulSoup telah di-parse ulang (ukuran: {len(page_source)} karakter)")
            
            # If Bad Request found, return None to signal retry needed
            if bad_request_found:
                return None
            
            def extract_laba_kotor_value(identifier_text: str) -> tuple[float, float]:
                """
                Extract Laba Kotor value from table structure.
                Structure: identifier in <td><div>, values in subsequent <td><div> elements.
                Values can be negative in parentheses like (677,555,231).
                Returns tuple of (current_year_value, previous_year_value)
                """
                import re
                
                for div in soup.find_all('div'):
                    text = div.get_text(strip=True)
                    if identifier_text.upper() in text.upper() and len(text) < 5000:
                        # Find the parent <td> or <tr> (table row)
                        parent_td = div.find_parent('td')
                        if not parent_td:
                            continue
                        
                        # Find the parent <tr> (table row) to get all <td> elements in the row
                        parent_tr = parent_td.find_parent('tr')
                        if not parent_tr:
                            # If no <tr>, try to find subsequent <td> elements from the same parent
                            parent_td_siblings = parent_td.find_next_siblings('td')
                            if not parent_td_siblings:
                                continue
                            tds = [parent_td] + list(parent_td_siblings)
                        else:
                            # Get all <td> elements in the row
                            tds = parent_tr.find_all('td')
                        
                        # Find the index of the identifier <td>
                        try:
                            identifier_td_index = tds.index(parent_td)
                        except:
                            continue
                        
                        # Extract values from subsequent <td> elements (skip the identifier <td>)
                        # Look for numeric values in <td> elements (may have multiple <td> elements to check)
                        values = []
                        import re
                        for td in tds[identifier_td_index + 1:identifier_td_index + 10]:  # Check more <td> elements to find numeric values
                            # Look for <div> inside the <td>
                            td_div = td.find('div')
                            if td_div:
                                td_text = td_div.get_text(strip=True)
                                if td_text:
                                    # Check if this looks like a number (contains digits, possibly with parentheses and commas)
                                    # Pattern: optional parentheses, digits with commas, optional decimal part
                                    # Examples: "(677,555,231)", "1,223", "123.45"
                                    numeric_pattern = r'^\(?[\d,]+\.?\d*\)?$'
                                    if re.match(numeric_pattern, td_text.replace(' ', '')):
                                        values.append(td_text)
                                        if len(values) >= 2:  # Stop after finding 2 numeric values
                                            break
                        
                        # Parse the values
                        if len(values) >= 2:
                            # Two values: first is previous year, second is current year
                            try:
                                # Parse first value (previous year)
                                val1_text = values[0].strip()
                                is_negative1 = val1_text.startswith('(') and val1_text.endswith(')')
                                if is_negative1:
                                    val1_text = val1_text[1:-1].strip()  # Remove parentheses
                                # Remove commas (thousands separators) and convert to float
                                val1_text = val1_text.replace(',', '')
                                val1 = float(val1_text)
                                if is_negative1:
                                    val1 = -val1
                                
                                # Parse second value (current year)
                                val2_text = values[1].strip()
                                is_negative2 = val2_text.startswith('(') and val2_text.endswith(')')
                                if is_negative2:
                                    val2_text = val2_text[1:-1].strip()  # Remove parentheses
                                # Remove commas (thousands separators) and convert to float
                                val2_text = val2_text.replace(',', '')
                                val2 = float(val2_text)
                                if is_negative2:
                                    val2 = -val2
                                
                                # Return as (current_year, previous_year)
                                return (val2, val1)
                            except Exception as e:
                                print(f"    [DEBUG] Error parsing two values: {e}, values: {values}")
                                pass
                        elif len(values) == 1:
                            # Single value: use for both years
                            try:
                                val_text = values[0].strip()
                                is_negative = val_text.startswith('(') and val_text.endswith(')')
                                if is_negative:
                                    val_text = val_text[1:-1].strip()  # Remove parentheses
                                # Remove commas (thousands separators) and convert to float
                                val_text = val_text.replace(',', '')
                                val = float(val_text)
                                if is_negative:
                                    val = -val
                                # Use same value for both years
                                return (val, val)
                            except Exception as e:
                                print(f"    [DEBUG] Error parsing single value: {e}, value: {values[0]}")
                                pass
                
                return (0.0, 0.0)
            
            def extract_ratio_value(identifier_text: str) -> float:
                """
                Extract Rasio value - just take the integer that has '.' in it
                If there is bracket, meaning it's negative, parse it as is
                """
                for div in soup.find_all('div'):
                    text = div.get_text(strip=True)
                    if identifier_text.upper() in text.upper() and len(text) < 5000:
                        # Find the next div with numeric value
                        all_divs = soup.find_all('div')
                        try:
                            div_index = all_divs.index(div)
                            # Look at next divs for the value
                            for i in range(div_index + 1, min(div_index + 10, len(all_divs))):
                                next_div = all_divs[i]
                                next_text = next_div.get_text(strip=True)
                                
                                # Skip if too long or contains identifier keywords
                                if len(next_text) > 100:
                                    continue
                                
                                # Extract number - must have '.' in it (decimal point)
                                cleaned_text = next_text.strip()
                                
                                # Check if it has a decimal point (dot)
                                if '.' not in cleaned_text:
                                    continue
                                
                                # Check for negative (in parentheses)
                                is_negative = False
                                if cleaned_text.startswith('(') and cleaned_text.endswith(')'):
                                    is_negative = True
                                    cleaned_text = cleaned_text[1:-1].strip()
                                elif cleaned_text.startswith('-'):
                                    is_negative = True
                                    cleaned_text = cleaned_text[1:].strip()
                                
                                # Replace comma with dot for decimal separator (if comma is used)
                                # But keep the dot that's already there
                                cleaned_text = cleaned_text.replace(',', '.')
                                
                                # Try to extract number
                                try:
                                    number = float(cleaned_text)
                                    if is_negative:
                                        number = -number
                                    # Reasonable range check
                                    if abs(number) < 1e15:
                                        return number
                                except:
                                    pass
                        except:
                            pass
                return 0.0
            
            # Extract Laba Kotor (for Sheet 4)
            print("    [INFO] Extracting Laba Kotor data...")
            laba_kotor_current, laba_kotor_previous = extract_laba_kotor_value("LABA (RUGI) TAHUN BERJALAN SEBELUM PAJAK PENGHASILAN")
            result[f'Laba Kotor {selected_year}'] = laba_kotor_current
            result[f'Laba Kotor {previous_year}'] = laba_kotor_previous
            print(f"    [OK] Extracted Laba Kotor: {selected_year}={laba_kotor_current}, {previous_year}={laba_kotor_previous}")
            
            # Define 9 ratio identifiers (for Sheet 5 only)
            ratio_identifiers = [
                ("Kewajiban Penyediaan Modal Minimum", "KPMM"),
                ("Rasio Cadangan terhadap PPKA", "PPKA"),
                ("Non Performing Loan (NPL) Neto", "NPL Neto"),
                ("Non Performing Loan (NPL) Gross", "NPL Gross"),
                ("Return on Assets (ROA)", "ROA"),
                ("Biaya Operasional terhadap Pendapatan Operasional (BOPO)", "BOPO"),
                ("Net Interest Margin (NIM)", "NIM"),
                ("Loan to Deposit Ratio (LDR)", "LDR"),
                ("Cash Ratio", "CR")
            ]
            
            # Extract all 9 ratios (for Sheet 5)
            print("    [INFO] Extracting all 9 Rasio data (Sheet 5)...")
            for identifier, ratio_name in ratio_identifiers:
                value = extract_ratio_value(identifier)
                result[ratio_name] = value
                print(f"    [DEBUG] Extracted {ratio_name}: {value}")
            
            print(f"    [OK] Extracted Laba Kotor and all 9 Rasio data")
            print(f"    [OK] Total extracted data points: {len(result)}")
            
            # Switch back to default content
            try:
                self.driver.switch_to.default_content()
                print("    [DEBUG] Switched back to default content")
            except:
                pass
            
            return result
            
        except Exception as e:
            print(f"    [ERROR] Error extracting Laba Kotor and Rasio data: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def cleanup(self, kill_processes: bool = False):
        """
        Close browser and cleanup all Selenium resources
        
        Args:
            kill_processes: If True, also kill any lingering Chrome/ChromeDriver processes
        """
        print("[INFO] Membersihkan sumber daya Selenium...")
        
        # Cleanup ExtJS helper first
        if self.extjs:
            try:
                self.extjs = None
            except:
                pass
        
        # Cleanup WebDriverWait
        if self.wait:
            try:
                self.wait = None
            except:
                pass
        
        # Cleanup WebDriver - this is the most important
        if self.driver:
            try:
                # Close all windows
                try:
                    self.driver.close()
                except:
                    pass
                
                # Quit driver (closes all windows and ends session)
                self.driver.quit()
                print("[OK] Browser ditutup")
            except Exception as e:
                print(f"[WARNING] Error saat menutup browser: {e}")
            finally:
                # Clear reference
                self.driver = None
        
        # Force garbage collection to free memory
        import gc
        gc.collect()
        print("[OK] Sumber daya Selenium telah dibersihkan")
        
        # Optionally kill lingering processes
        if kill_processes:
            try:
                from .utils import kill_chrome_processes
                kill_chrome_processes()
            except ImportError:
                try:
                    from utils import kill_chrome_processes
                    kill_chrome_processes()
                except ImportError:
                    print("[WARNING] Tidak dapat mengimpor fungsi kill_chrome_processes")
    
    # ============================================================================
    # Direct URL Retry Methods (for zero value banks)
    # ============================================================================
    
    def _format_bank_code_for_url(self, bank_name: str) -> list:
        """
        Format bank name for URL with fallback options.
        Returns a list of possible bank code formats to try.
        
        Similar to Sindikasi scraper - formats bank name for direct URL usage.
        
        Args:
            bank_name: Bank name (e.g., "PT Bpr Rangkiang Aur Denai" or "Bprs Al-Makmur")
            
        Returns:
            List of formatted bank codes: [original_format, expanded_format]
            If no expansion needed, returns [original_format]
        """
        # Step 1: Replace "-" with spaces
        name = bank_name.replace("-", " ")
        
        # Step 2: Split by spaces and clean
        words = [w.strip() for w in name.split() if w.strip()]
        
        if not words:
            return [""]
        
        # Step 3: Ensure name starts with "PT" or "Perumda"
        first_word_upper = words[0].upper()
        if first_word_upper not in ["PT", "PERUMDA"]:
            # Prepend "PT" if it doesn't start with PT or Perumda
            words.insert(0, "PT")
        
        # Step 4: Handle PERUMDA case (normalize to "Perumda")
        if words[0].upper() == "PERUMDA":
            words[0] = "Perumda"
        
        # Step 5: Create original format (keep Bpr/Bprs as-is)
        # PT should always be uppercase, other words capitalized
        original_words = []
        for w in words:
            if w.upper() == "PT":
                original_words.append("PT")
            else:
                original_words.append(w.capitalize())
        original_format = "+".join(original_words)
        
        # Step 6: Check if we need expanded format
        has_bpr = False
        has_bprs = False
        expanded_words = []
        
        for word in words:
            word_upper = word.upper()
            if word_upper in ["BPR", "BPRS"]:
                if word_upper == "BPRS":
                    has_bprs = True
                    expanded_words.extend(["Bank", "Perekonomian", "Rakyat", "Syariah"])
                else:
                    has_bpr = True
                    expanded_words.extend(["Bank", "Perekonomian", "Rakyat"])
            else:
                expanded_words.append(word)
        
        # Step 7: Create expanded format if needed
        formats = [original_format]
        
        if has_bpr or has_bprs:
            # Capitalize expanded words, but keep PT uppercase
            expanded_formatted = []
            for w in expanded_words:
                if w.upper() == "PT":
                    expanded_formatted.append("PT")
                else:
                    expanded_formatted.append(w.capitalize())
            expanded_format = "+".join(expanded_formatted)
            formats.append(expanded_format)
        
        return formats
    
    def _month_name_to_number(self, month_name: str) -> int:
        """
        Convert month name to number (for direct URL)
        
        Args:
            month_name: Month name (Maret, Juni, September, Desember)
            
        Returns:
            Month number (3, 6, 9, or 12)
        """
        month_map = {
            "Maret": 3,
            "Juni": 6,
            "September": 9,
            "Desember": 12
        }
        return month_map.get(month_name, 3)  # Default to 3 if not found
    
    def _build_report_url(self, bank_code: str, month: int, year: str, form_number: int) -> str:
        """
        Build the report viewer URL for direct access
        
        Args:
            bank_code: Formatted bank code (e.g., "PT+Bpr+Rangkiang+Aur+Denai")
            month: Month number (3, 6, 9, or 12)
            year: Year (e.g., "2025")
            form_number: Form number (1, 2, or 3)
            
        Returns:
            Complete URL string
        """
        base_url = "https://cfs.ojk.go.id/cfs/ReportViewerForm.aspx"
        
        # Form code for Konvensional
        form_code = f"BPK-901-{form_number:06d}"
        
        # Build URL parameters
        params = {
            "BankCodeNumber": "620000",  # Static placeholder
            "BankCode": bank_code,
            "Month": str(month),
            "Year": year,
            "FinancialReportPeriodTypeCode": "R",
            "FinancialReportTypeCode": form_code
        }
        
        # Build query string
        query_string = "&".join([f"{k}={v}" for k, v in params.items()])
        url = f"{base_url}?{query_string}"
        
        return url
    
    def _check_for_server_error(self) -> bool:
        """
        Check if the current page has a server error
        
        Returns:
            True if server error found, False otherwise
        """
        try:
            page_source = self.driver.page_source
            error_text = "Server Error in '/cfs' Application"
            
            if error_text in page_source:
                print(f"  [WARNING] Server error detected in page source")
                return True
            
            return False
        except Exception as e:
            print(f"  [WARNING] Error checking for server error: {e}")
            return False
    
    def _get_page_source_with_iframe(self) -> tuple[str, object]:
        """
        Get page source, checking iframes first (similar to Sindikasi)
        
        Returns:
            Tuple of (page_source: str, report_iframe: object or None)
        """
        # Try to find report in iframe first, then main page
        iframes = self.driver.find_elements(By.TAG_NAME, "iframe")
        page_source = None
        report_iframe = None
        
        # Check iframes for report content
        for iframe in iframes:
            try:
                self.driver.switch_to.frame(iframe)
                iframe_source = self.driver.page_source
                if ("Piutang" in iframe_source or "Aset" in iframe_source or "DPK" in iframe_source or
                    "LABA" in iframe_source or "Rasio" in iframe_source or "KPMM" in iframe_source or
                    "Kredit" in iframe_source):
                    print(f"  [DEBUG] Found report content in iframe")
                    page_source = iframe_source
                    report_iframe = iframe
                    break
                self.driver.switch_to.default_content()
            except:
                self.driver.switch_to.default_content()
                continue
        
        # If not in iframe, use main page
        if page_source is None:
            print(f"  [DEBUG] Using main page source")
            self.driver.switch_to.default_content()
            page_source = self.driver.page_source
        
        return page_source, report_iframe
    
    def _clean_numeric_text(self, text: str) -> float:
        """
        Clean numeric text by reversing . and , then parse as float
        
        Args:
            text: Text containing numeric value (e.g., "230,115,190" becomes "230.115.190")
            
        Returns:
            Parsed float value, or 0.0 if parsing fails
        """
        try:
            import re
            
            # Remove whitespace
            cleaned = text.strip()
            
            # Reverse . and , (swap all occurrences)
            # "230,115,190" → "230.115.190"
            # "230.115.190" → "230,115,190"
            cleaned = cleaned.replace('.', 'TEMP_DOT')
            cleaned = cleaned.replace(',', '.')
            cleaned = cleaned.replace('TEMP_DOT', ',')
            
            # After swapping, parse the number
            # Determine which separator is decimal and which is thousands
            # Pattern: if there are multiple separators, the last one is usually decimal
            # If only one separator, it could be either
            
            dot_count = cleaned.count('.')
            comma_count = cleaned.count(',')
            
            if dot_count > 0 and comma_count > 0:
                # Both present - last separator is decimal
                last_dot_pos = cleaned.rfind('.')
                last_comma_pos = cleaned.rfind(',')
                
                if last_dot_pos > last_comma_pos:
                    # Last separator is dot, so dots are thousands, comma is decimal
                    # Remove all dots (thousands), replace comma with dot for parsing
                    cleaned = cleaned.replace('.', '')
                    cleaned = cleaned.replace(',', '.')
                else:
                    # Last separator is comma, so commas are thousands, dot is decimal
                    # Remove all commas (thousands), keep dot as decimal
                    cleaned = cleaned.replace(',', '')
            elif dot_count > 1:
                # Multiple dots - all are thousands separators, remove all
                cleaned = cleaned.replace('.', '')
            elif comma_count > 1:
                # Multiple commas - all are thousands separators, remove all
                cleaned = cleaned.replace(',', '')
            elif dot_count == 1:
                # Single dot - could be decimal or thousands
                # Check position: if near end (last 3 chars), likely decimal, otherwise thousands
                dot_pos = cleaned.find('.')
                if dot_pos >= len(cleaned) - 3:
                    # Near end, likely decimal separator
                    pass  # Keep it
                else:
                    # Likely thousands separator, remove it
                    cleaned = cleaned.replace('.', '')
            elif comma_count == 1:
                # Single comma - could be decimal or thousands
                # Check position: if near end (last 3 chars), likely decimal, otherwise thousands
                comma_pos = cleaned.find(',')
                if comma_pos >= len(cleaned) - 3:
                    # Near end, likely decimal separator, convert to dot
                    cleaned = cleaned.replace(',', '.')
                else:
                    # Likely thousands separator, remove it
                    cleaned = cleaned.replace(',', '')
            
            # Remove any remaining non-numeric characters except decimal point and minus
            cleaned = re.sub(r'[^\d.\-]', '', cleaned)
            
            if not cleaned or cleaned == '-':
                return 0.0
            
            return float(cleaned)
        except Exception as e:
            print(f"  [DEBUG] Error parsing numeric text '{text}': {e}")
            return 0.0
    
    def _extract_identifier_value(self, soup: BeautifulSoup, identifier: str) -> dict:
        """
        Extract values for a single identifier from the page
        Uses the same approach as Sindikasi scraper:
        - Find div containing identifier text
        - Look at next divs to find numeric values
        - First numeric = current year, second = previous year
        
        Args:
            soup: BeautifulSoup parsed page
            identifier: Identifier text to find
            
        Returns:
            dict with {'2025': value, '2024': value} or {'2025': 0.0, '2024': 0.0} if not found
        """
        result = {'2025': 0.0, '2024': 0.0}
        
        try:
            # Helper function to extract number from text
            def extract_number(text: str) -> float:
                """Extract a number from Indonesian-style formatted text (handles decimals correctly)."""
                if not text:
                    return 0.0
                
                # Normalize spaces
                text = text.replace('\xa0', ' ').replace('&nbsp;', ' ').strip()
                
                # Use _clean_numeric_text to properly handle Indonesian format (handles both whole numbers and decimals)
                return self._clean_numeric_text(text)
            
            # Helper function to split concatenated numbers (current year + previous year)
            def split_concatenated_numbers(text: str) -> tuple[str, str]:
                """
                Split concatenated numbers like "23,122,1223,112,122" into two numbers.
                Format: After comma, max 3 digits. If next is comma, same year. If next is digit, next year starts.
                """
                if not text or ',' not in text:
                    return text, ""
                
                # Find the split point: look for pattern where after comma, we have 3 digits, then a digit (not comma)
                pattern = r',(\d{3})(\d)(?=,|\d|$)'
                match = re.search(pattern, text)
                
                if match:
                    split_pos = match.end(1)  # Position after the 3 digits (before the next digit)
                    current_year_text = text[:split_pos].rstrip(',')
                    previous_year_text = text[split_pos:].lstrip(',')
                    return current_year_text, previous_year_text
                
                return text, ""
            
            # Find the <div> whose text contains our identifier
            label_div = None
            all_divs = soup.find_all('div')
            
            for div in all_divs:
                text = div.get_text(strip=True)
                if not text:
                    continue
                
                # Skip divs that are too long (likely contain entire page content)
                if len(text) > 5000:
                    continue
                
                # Check if identifier is in text
                if identifier.lower() in text.lower():
                    # Prefer divs where identifier is a significant part of the text
                    text_lower = text.lower()
                    identifier_lower = identifier.lower()
                    
                    # If text is short or identifier is at the start/end, it's likely the right div
                    if len(text) < 200 or text_lower.startswith(identifier_lower) or text_lower.endswith(identifier_lower):
                        label_div = div
                        print(f"  [DEBUG] Found identifier '{identifier}' in <div>: '{text[:100]}...'")
                        break
            
            if not label_div:
                print(f"  [DEBUG] Identifier '{identifier}' NOT FOUND in page")
                return result
            
            # Find the index of this div and get the next divs that contain numeric values
            try:
                div_index = all_divs.index(label_div)
                # Look through next divs to find numeric values (up to 10 divs ahead)
                numeric_count = 0
                for j in range(1, min(11, len(all_divs) - div_index)):  # Check up to 10 next divs
                    if numeric_count >= 2:  # We need 2 values
                        break
                    
                    if div_index + j < len(all_divs):
                        next_div = all_divs[div_index + j]
                        div_text = next_div.get_text(strip=True)
                        
                        if not div_text:
                            continue
                        
                        # Skip very long divs (likely contain entire page content)
                        if len(div_text) > 5000:
                            continue
                        
                        # Skip if it's clearly another identifier (contains common identifier keywords)
                        if any(keyword in div_text.lower() for keyword in ['kepada', 'pihak', 'bank', 'bpr', 'report viewer', 'configuration error', 'piutang', 'aset', 'dpk', 'laba', 'rasio']):
                            continue
                        
                        # Check if it contains concatenated numbers (has comma and might have split pattern)
                        if ',' in div_text and len(div_text) > 10:
                            # Try to split concatenated numbers
                            current_year_text, prev_year_text = split_concatenated_numbers(div_text)
                            
                            if prev_year_text:
                                # Found concatenated numbers, extract both
                                current_number = extract_number(current_year_text)
                                prev_number = extract_number(prev_year_text)
                                
                                # Validate numbers are reasonable
                                if (current_number >= 0 and current_number < 1e15 and current_number != float('inf') and
                                    prev_number >= 0 and prev_number < 1e15 and prev_number != float('inf')):
                                    
                                    if numeric_count == 0:
                                        result['2025'] = current_number
                                        numeric_count += 1
                                    
                                    if numeric_count == 1:
                                        result['2024'] = prev_number
                                        numeric_count += 1
                                    
                                    if numeric_count >= 2:
                                        break
                                    continue
                        
                        # Check if this div looks like it contains a single formatted number
                        if len(div_text) > 100:
                            continue
                        
                        # Extract number from single number text
                        number = extract_number(div_text)
                        
                        # Validate the number is reasonable
                        if number >= 0 and number < 1e15 and number != float('inf'):
                            if numeric_count == 0:
                                result['2025'] = number
                                numeric_count += 1
                            elif numeric_count == 1:
                                result['2024'] = number
                                numeric_count += 1
                        elif number == 0 and any(char.isdigit() for char in div_text) and len(div_text) < 50:
                            # Zero value is valid if it's a short text with digits
                            if numeric_count == 0:
                                result['2025'] = number
                                numeric_count += 1
                            elif numeric_count == 1:
                                result['2024'] = number
                                numeric_count += 1
            except ValueError:
                print(f"  [DEBUG] Identifier '{identifier}' found but couldn't find its index")
                return result
            
            if result['2025'] == 0.0 and result['2024'] == 0.0:
                print(f"  [DEBUG] Identifier '{identifier}' found but no numeric values extracted from next divs")
            elif result['2024'] == 0.0:
                print(f"  [DEBUG] Identifier '{identifier}' found but only 1 value extracted")
            
        except Exception as e:
            print(f"  [DEBUG] Error extracting identifier '{identifier}': {e}")
            import traceback
            print(traceback.format_exc())
        
        return result
    
    def _parse_form1_direct_url(self) -> dict:
        """
        Parse BPK-901-000001 data using direct URL
        
        Extracts:
        - ASET: Total Aset
        - KREDIT: Sum of (a) Kepada BPR, (b) Kepada Bank Umum, (c) Kepada non bank – pihak terkait, (d) Kepada non bank – pihak tidak terkait
        - DPK: Sum of (a) Tabungan, (b) Deposito, (c) Simpanan dari Bank Lain
        
        Returns:
            dict with structure:
            {
                'ASET': {'2025': value, '2024': value},
                'KREDIT': {'2025': sum, '2024': sum, 'individual': {...}},
                'DPK': {'2025': sum, '2024': sum, 'individual': {...}}
            }
        """
        result = {
            'ASET': {'2025': 0.0, '2024': 0.0},
            'KREDIT': {'2025': 0.0, '2024': 0.0, 'individual': {}},
            'DPK': {'2025': 0.0, '2024': 0.0, 'individual': {}}
        }
        
        try:
            # Wait a bit for page to load
            time.sleep(3.0)
            
            # Get page source (check iframes first)
            page_source, report_iframe = self._get_page_source_with_iframe()
            soup = BeautifulSoup(page_source, 'html.parser')
            
            # Extract ASET
            aset_values = self._extract_identifier_value(soup, "Total Aset")
            result['ASET'] = aset_values
            print(f"    ASET (Total Aset): 2025={aset_values['2025']:,.2f}, 2024={aset_values['2024']:,.2f}")
            
            # Extract KREDIT components
            kredit_identifiers = [
                "Kepada BPR",
                "Kepada Bank Umum",
                "Kepada non bank – pihak terkait",
                "Kepada non bank – pihak tidak terkait"
            ]
            
            for identifier in kredit_identifiers:
                values = self._extract_identifier_value(soup, identifier)
                result['KREDIT']['individual'][identifier] = values
                result['KREDIT']['2025'] += values['2025']
                result['KREDIT']['2024'] += values['2024']
                print(f"    {identifier}: 2025={values['2025']:,.2f}, 2024={values['2024']:,.2f}")
            
            print(f"    KREDIT Total: 2025={result['KREDIT']['2025']:,.2f}, 2024={result['KREDIT']['2024']:,.2f}")
            
            # Extract DPK components
            dpk_identifiers = [
                "Tabungan",
                "Deposito",
                "Simpanan dari Bank Lain"
            ]
            
            for identifier in dpk_identifiers:
                values = self._extract_identifier_value(soup, identifier)
                result['DPK']['individual'][identifier] = values
                result['DPK']['2025'] += values['2025']
                result['DPK']['2024'] += values['2024']
                print(f"    {identifier}: 2025={values['2025']:,.2f}, 2024={values['2024']:,.2f}")
            
            print(f"    DPK Total: 2025={result['DPK']['2025']:,.2f}, 2024={result['DPK']['2024']:,.2f}")
            
            return result
            
        except Exception as e:
            print(f"  [ERROR] Error parsing form 1: {e}")
            import traceback
            print(traceback.format_exc())
            return result
    
    def _parse_form2_direct_url(self) -> dict:
        """
        Parse BPK-901-000002 data using direct URL
        
        Extracts:
        - LABA KOTOR: LABA (RUGI) TAHUN BERJALAN SEBELUM PAJAK PENGHASILAN
        - LABA BERSIH: JUMLAH LABA (RUGI) TAHUN BERJALAN
        
        Returns:
            dict with structure:
            {
                'LABA KOTOR': {'2025': value, '2024': value},
                'LABA BERSIH': {'2025': value, '2024': value}
            }
        """
        result = {
            'LABA KOTOR': {'2025': 0.0, '2024': 0.0},
            'LABA BERSIH': {'2025': 0.0, '2024': 0.0}
        }
        
        try:
            # Wait a bit for page to load
            time.sleep(3.0)
            
            # Get page source (check iframes first)
            page_source, report_iframe = self._get_page_source_with_iframe()
            soup = BeautifulSoup(page_source, 'html.parser')
            
            # Extract LABA KOTOR
            laba_kotor_values = self._extract_identifier_value(soup, "LABA (RUGI) TAHUN BERJALAN SEBELUM PAJAK PENGHASILAN")
            result['LABA KOTOR'] = laba_kotor_values
            print(f"    LABA KOTOR: 2025={laba_kotor_values['2025']:,.2f}, 2024={laba_kotor_values['2024']:,.2f}")
            
            # Extract LABA BERSIH
            laba_bersih_values = self._extract_identifier_value(soup, "JUMLAH LABA (RUGI) TAHUN BERJALAN")
            result['LABA BERSIH'] = laba_bersih_values
            print(f"    LABA BERSIH: 2025={laba_bersih_values['2025']:,.2f}, 2024={laba_bersih_values['2024']:,.2f}")
            
            return result
            
        except Exception as e:
            print(f"  [ERROR] Error parsing form 2: {e}")
            import traceback
            print(traceback.format_exc())
            return result
    
    def _parse_form3_direct_url(self) -> dict:
        """
        Parse BPK-901-000003 data using direct URL
        
        Extracts ratios (separate, not summed):
        - KPMM: Kewajiban Penyediaan Modal Minimum (KPMM)
        - NPL: Non Performing Loan (NPL)
        - ROA: Return on Assets (ROA)
        - BOPO: Biaya Operasional terhadap Pendapatan Operasional (BOPO)
        - NIM: Net Interest Margin (NIM)
        - LDR: Loan to Deposit Ratio (LDR)
        - Cash Ratio: Cash Ratio
        
        Returns:
            dict with structure:
            {
                'KPMM': {'2025': value, '2024': value},
                'NPL': {'2025': value, '2024': value},
                ...
            }
        """
        result = {}
        
        ratios = [
            ("KPMM", "Kewajiban Penyediaan Modal Minimum (KPMM)"),
            ("NPL", "Non Performing Loan (NPL)"),
            ("ROA", "Return on Assets (ROA)"),
            ("BOPO", "Biaya Operasional terhadap Pendapatan Operasional (BOPO)"),
            ("NIM", "Net Interest Margin (NIM)"),
            ("LDR", "Loan to Deposit Ratio (LDR)"),
            ("Cash Ratio", "Cash Ratio")
        ]
        
        try:
            # Wait a bit for page to load
            time.sleep(3.0)
            
            # Get page source (check iframes first)
            page_source, report_iframe = self._get_page_source_with_iframe()
            soup = BeautifulSoup(page_source, 'html.parser')
            
            for ratio_name, identifier in ratios:
                values = self._extract_identifier_value(soup, identifier)
                result[ratio_name] = values
                # Format ratios without thousands separators (they're usually percentages)
                print(f"    {ratio_name}: 2025={values['2025']:.2f}, 2024={values['2024']:.2f}")
            
            return result
            
        except Exception as e:
            print(f"  [ERROR] Error parsing form 3: {e}")
            import traceback
            print(traceback.format_exc())
            return result
    
    def _read_excel_for_zero_values(self, month: str, year: str) -> list:
        """
        Read Excel file and find banks with 0 or 0,00 values in KPMM (first ratio table in Rasio sheet)
        
        Args:
            month: Month name (e.g., "Desember")
            year: Year (e.g., "2025")
            
        Returns:
            List of dicts: [{'bank_name': str, 'city': str, 'sheets_with_zero': [sheet_names]}, ...]
        """
        banks_with_zero = []
        
        try:
            from openpyxl import load_workbook
            
            # Get filename and filepath
            filename = self._get_excel_filename(month, year)
            publikasi_dir = self.output_dir / "publikasi"
            filepath = publikasi_dir / filename
            
            if not filepath.exists():
                print(f"  [WARNING] Excel file not found: {filepath}")
                return banks_with_zero
            
            print(f"  [INFO] Reading Excel file for zero KPMM values: {filename}")
            wb = load_workbook(filepath, data_only=True)
            
            month_num = self._get_month_number(month)
            sheet_name_prefix = f"{month_num}-{year[-2:]}"
            
            # Only check Rasio sheet, specifically KPMM (first ratio table)
            rasio_sheet_name = f"{sheet_name_prefix} Rasio"
            if rasio_sheet_name not in wb.sheetnames:
                print(f"  [WARNING] Rasio sheet not found: {rasio_sheet_name}")
                wb.close()
                return banks_with_zero
            
            ws = wb[rasio_sheet_name]
            print(f"    Checking sheet: {rasio_sheet_name} (KPMM only)")
            
            # Track banks with zero KPMM values
            bank_zero_map = {}  # {bank_key: {'bank_name': str, 'city': str, 'sheets_with_zero': set}}
            
            # Find KPMM table by looking for header row with "KPMM" as column header
            # Structure: Header row has columns: No, Nama Bank, Lokasi, KPMM
            kpmm_header_row = None
            kpmm_data_start_row = None
            
            # Look for header row that contains "KPMM" as a column header
            for row_num in range(1, min(50, ws.max_row + 1)):  # Check first 50 rows for header
                # Check if this row has "KPMM" in any cell (likely column D/4)
                for col_num in range(1, 5):  # Check columns A-D
                    cell_value = ws.cell(row=row_num, column=col_num).value
                    if cell_value and "KPMM" in str(cell_value).upper():
                        # Also check if this looks like a header row (has "No", "Nama Bank", "Lokasi")
                        no_cell = ws.cell(row=row_num, column=1).value
                        nama_bank_cell = ws.cell(row=row_num, column=2).value
                        lokasi_cell = ws.cell(row=row_num, column=3).value
                        
                        if (no_cell and ("No" in str(no_cell) or str(no_cell).isdigit()) and
                            nama_bank_cell and "Nama" in str(nama_bank_cell) and "Bank" in str(nama_bank_cell)):
                            kpmm_header_row = row_num
                            kpmm_data_start_row = row_num + 1  # Data starts after header
                            print(f"    Found KPMM table header at row {row_num}")
                            break
                
                if kpmm_header_row:
                    break
            
            if not kpmm_header_row:
                print(f"  [WARNING] KPMM table header not found in Rasio sheet")
                wb.close()
                return banks_with_zero
            
            # Find where KPMM table ends (next header row or empty row)
            kpmm_data_end_row = ws.max_row + 1
            for row_num in range(kpmm_data_start_row + 1, min(kpmm_data_start_row + 200, ws.max_row + 1)):
                # Check if this row is a header for next table (has ratio name like "PPKA", "NPL", etc.)
                cell_value = ws.cell(row=row_num, column=4).value  # Column D
                if cell_value:
                    cell_str = str(cell_value).upper()
                    # Check if it's another ratio header (not KPMM)
                    if cell_str in ["PPKA", "NPL NETO", "NPL GROSS", "ROA", "BOPO", "NIM", "LDR", "CR", "CASH RATIO"]:
                        kpmm_data_end_row = row_num
                        print(f"    Found end of KPMM table at row {row_num} (next ratio: {cell_str})")
                        break
                # Also check if row is empty (might indicate end of table)
                bank_name = ws.cell(row=row_num, column=2).value
                if not bank_name:
                    # Check a few more rows to see if table really ended
                    empty_count = 0
                    for check_row in range(row_num, min(row_num + 3, ws.max_row + 1)):
                        if not ws.cell(check_row, 2).value:
                            empty_count += 1
                    if empty_count >= 2:
                        kpmm_data_end_row = row_num
                        print(f"    Found end of KPMM table at row {row_num} (empty rows)")
                        break
            
            # Check KPMM data rows
            for row_num in range(kpmm_data_start_row, kpmm_data_end_row):
                bank_name_cell = ws.cell(row=row_num, column=2)  # Column B: Nama Bank
                lokasi_cell = ws.cell(row=row_num, column=3)  # Column C: Lokasi
                value_cell = ws.cell(row=row_num, column=4)  # Column D: KPMM value
                
                bank_name = bank_name_cell.value
                lokasi = lokasi_cell.value if lokasi_cell.value else ""
                value = value_cell.value
                
                if not bank_name:
                    continue
                
                bank_name = str(bank_name).strip()
                city = str(lokasi).strip() if lokasi else ""
                
                # Check if value is 0, None, or "0,00"/"0.00" format
                is_zero = False
                if value is None:
                    is_zero = True
                elif isinstance(value, (int, float)):
                    is_zero = (value == 0 or value == 0.0)
                elif isinstance(value, str):
                    # Check for string formats like "0,00", "0.00", "0"
                    value_clean = value.strip()
                    # Remove all commas, dots, and spaces to check if it's effectively zero
                    cleaned = value_clean.replace(',', '').replace('.', '').replace(' ', '').replace('-', '')
                    # Check if it's empty, just zeros, or matches common zero patterns
                    if cleaned == '' or cleaned == '0' or cleaned == '00' or cleaned == '000':
                        is_zero = True
                    # Also check direct string matches for common zero formats
                    elif value_clean.lower() in ['0', '0,00', '0.00', '0,0', '0.0', '0,000', '0.000', '-', '']:
                        is_zero = True
                    else:
                        # Try to parse as float to check if it's zero
                        try:
                            # Replace comma with dot for parsing
                            parsed = float(value_clean.replace(',', '.'))
                            is_zero = (parsed == 0.0)
                        except:
                            pass
                
                # If KPMM is zero, add bank to retry list
                if is_zero:
                    bank_key = f"{bank_name}|{city}"
                    if bank_key not in bank_zero_map:
                        bank_zero_map[bank_key] = {
                            'bank_name': bank_name,
                            'city': city,
                            'sheets_with_zero': set()
                        }
                    bank_zero_map[bank_key]['sheets_with_zero'].add(rasio_sheet_name)
                    print(f"      Found zero KPMM: {bank_name} ({city}) - value: {value}")
            
            # Convert to list format
            for bank_data in bank_zero_map.values():
                banks_with_zero.append({
                    'bank_name': bank_data['bank_name'],
                    'city': bank_data['city'],
                    'sheets_with_zero': list(bank_data['sheets_with_zero'])
                })
            
            print(f"  [INFO] Found {len(banks_with_zero)} banks with zero KPMM values")
            for bank in banks_with_zero:
                print(f"    - {bank['bank_name']} ({bank['city']})")
            
            wb.close()
            return banks_with_zero
            
        except Exception as e:
            print(f"  [ERROR] Error reading Excel for zero values: {e}")
            import traceback
            print(traceback.format_exc())
            return banks_with_zero
    
    def _retry_bank_with_direct_url(self, bank_name: str, month: str, year: str) -> dict:
        """
        Retry single bank using direct URL method
        
        Args:
            bank_name: Bank name to retry
            month: Month name (e.g., "Desember")
            year: Year (e.g., "2025")
            
        Returns:
            dict with extracted data: {'form1': {...}, 'form2': {...}, 'form3': {...}}
        """
        result = {
            'form1': None,
            'form2': None,
            'form3': None
        }
        
        try:
            # Get target month and year
            month_num = self._month_name_to_number(month)
            form_numbers = [1, 2, 3]
            
            print(f"  [INFO] Retrying bank: {bank_name[:80]}...")
            
            # Get bank code formats (original + expanded if needed)
            bank_code_formats = self._format_bank_code_for_url(bank_name)
            print(f"  [DEBUG] Bank code formats to try: {bank_code_formats}")
            
            # Try each bank code format
            for format_idx, bank_code in enumerate(bank_code_formats, 1):
                print(f"  [INFO] Trying bank code format {format_idx}/{len(bank_code_formats)}: {bank_code}")
                
                format_success = False
                server_error_on_form1 = False
                
                # Process each form
                for form_num in form_numbers:
                    print(f"    Processing form {form_num}...")
                    
                    # If form 1 had server error, skip forms 2 and 3 and try next bank code format
                    if server_error_on_form1 and form_num > 1:
                        print(f"    Skipping form {form_num} (form 1 had server error, trying next bank code format)")
                        break
                    
                    # Build URL
                    url = self._build_report_url(bank_code, month_num, year, form_num)
                    print(f"    URL: {url}")
                    
                    # Navigate to URL
                    max_retries = 2
                    parsed_data = None
                    
                    for retry_attempt in range(max_retries + 1):
                        try:
                            if retry_attempt > 0:
                                print(f"    Retrying form {form_num} (attempt {retry_attempt + 1}/{max_retries + 1})...")
                                # Refresh the page on retry
                                self.driver.refresh()
                                time.sleep(3.0)
                            else:
                                self.driver.get(url)
                                time.sleep(2.0)  # Wait for page to load
                            
                            # Check for server error
                            if self._check_for_server_error():
                                print(f"    [WARNING] Server error for form {form_num}")
                                if form_num == 1:
                                    server_error_on_form1 = True
                                break  # Break from retry loop, try next form or next format
                            
                            # Parse form data
                            if form_num == 1:
                                parsed_data = self._parse_form1_direct_url()
                            elif form_num == 2:
                                parsed_data = self._parse_form2_direct_url()
                            elif form_num == 3:
                                parsed_data = self._parse_form3_direct_url()
                            
                            # If we got data, break from retry loop
                            if parsed_data:
                                break
                            else:
                                print(f"    [WARNING] Form {form_num} - no data parsed")
                                if retry_attempt < max_retries:
                                    continue  # Retry
                                else:
                                    break  # No more retries
                                
                        except Exception as e:
                            error_msg = str(e)
                            # Check if it's a Chrome timeout error
                            is_timeout_error = "timeout" in error_msg.lower() or "Timed out receiving message from renderer" in error_msg
                            
                            if is_timeout_error and retry_attempt < max_retries:
                                print(f"    [WARNING] Chrome timeout error on form {form_num}, refreshing page and retrying...")
                                try:
                                    self.driver.refresh()
                                    time.sleep(3.0)
                                    continue  # Retry
                                except:
                                    pass
                            
                            print(f"    [ERROR] Error processing form {form_num}: {e}")
                            import traceback
                            print(traceback.format_exc())
                            
                            if retry_attempt < max_retries:
                                continue  # Retry
                            else:
                                break  # No more retries
                    
                    # Store parsed data if we got it
                    if parsed_data:
                        if form_num == 1:
                            result['form1'] = parsed_data
                        elif form_num == 2:
                            result['form2'] = parsed_data
                        elif form_num == 3:
                            result['form3'] = parsed_data
                        
                        format_success = True
                        print(f"    [OK] Form {form_num} - data parsed successfully!")
                    
                    # If form 1 had server error, break from form loop to try next bank code format
                    if server_error_on_form1:
                        break
                
                # If this format worked, we can break
                if format_success:
                    print(f"  [OK] Bank code format {format_idx} succeeded")
                    break  # Break from format loop
            
            return result
            
        except Exception as e:
            print(f"  [ERROR] Error retrying bank {bank_name}: {e}")
            import traceback
            print(traceback.format_exc())
            return result
    
    def _update_excel_with_retry_data(self, month: str, year: str, retry_results: dict):
        """
        Update Excel file with retry data
        
        Args:
            month: Month name (e.g., "Desember")
            year: Year (e.g., "2025")
            retry_results: Dict mapping bank_name to retry data: {bank_name: {'city': str, 'form1': {...}, 'form2': {...}, 'form3': {...}}}
        """
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Border, Side, Alignment
            
            # Get filename and filepath
            filename = self._get_excel_filename(month, year)
            publikasi_dir = self.output_dir / "publikasi"
            filepath = publikasi_dir / filename
            
            if not filepath.exists():
                print(f"  [WARNING] Excel file not found: {filepath}")
                return
            
            print(f"  [INFO] Updating Excel file with retry data: {filename}")
            wb = load_workbook(filepath)
            
            month_num = self._get_month_number(month)
            sheet_name_prefix = f"{month_num}-{year[-2:]}"
            previous_year = str(int(year) - 1)
            
            # Define border style
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Update each bank in retry_results
            for bank_name, bank_data in retry_results.items():
                city = bank_data.get('city', '')
                form1 = bank_data.get('form1')
                form2 = bank_data.get('form2')
                form3 = bank_data.get('form3')
                
                print(f"  [INFO] Updating bank: {bank_name} ({city})")
                
                # Update ASET sheet
                if form1 and 'ASET' in form1:
                    sheet_name = f"{sheet_name_prefix} ASET"
                    if sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        self._update_excel_row_for_retry(ws, bank_name, city, 'ASET', form1['ASET'], year, previous_year, thin_border)
                
                # Update Kredit sheet
                if form1 and 'KREDIT' in form1:
                    sheet_name = f"{sheet_name_prefix} Kredit"
                    if sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        self._update_excel_row_for_retry(ws, bank_name, city, 'Kredit', form1['KREDIT'], year, previous_year, thin_border)
                
                # Update DPK sheet
                if form1 and 'DPK' in form1:
                    sheet_name = f"{sheet_name_prefix} DPK"
                    if sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        self._update_excel_row_for_retry(ws, bank_name, city, 'DPK', form1['DPK'], year, previous_year, thin_border)
                
                # Update Laba Kotor sheet
                if form2 and 'LABA KOTOR' in form2:
                    sheet_name = f"{sheet_name_prefix} Laba Kotor"
                    if sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        self._update_excel_row_for_retry(ws, bank_name, city, 'Laba Kotor', form2['LABA KOTOR'], year, previous_year, thin_border)
                
                # Update Rasio sheet
                if form3:
                    sheet_name = f"{sheet_name_prefix} Rasio"
                    if sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        self._update_rasio_sheet_for_retry(ws, bank_name, form3, year, previous_year, thin_border)
            
            # Save updated Excel file
            wb.save(filepath)
            print(f"  [OK] Excel file updated: {filepath}")
            wb.close()
            
        except Exception as e:
            print(f"  [ERROR] Error updating Excel with retry data: {e}")
            import traceback
            print(traceback.format_exc())
    
    def _update_excel_row_for_retry(self, ws, bank_name: str, city: str, data_type: str, values: dict, year: str, previous_year: str, border):
        """
        Update a single row in Excel sheet with retry data
        
        Args:
            ws: Worksheet object
            bank_name: Bank name
            city: City name
            data_type: Type of data (ASET, Kredit, DPK, Laba Kotor)
            values: Dict with {'2025': value, '2024': value}
            year: Current year
            previous_year: Previous year
            border: Border style
        """
        try:
            # Find the row with matching bank name and city
            for row_num in range(3, ws.max_row + 1):
                bank_name_cell = ws.cell(row=row_num, column=2)  # Column B: Nama Bank
                city_cell = ws.cell(row=row_num, column=3)  # Column C: Lokasi
                
                if (bank_name_cell.value and str(bank_name_cell.value).strip() == str(bank_name).strip() and
                    city_cell.value and str(city_cell.value).strip() == str(city).strip()):
                    
                    # Update values
                    val_2025 = values.get('2025', 0.0) if isinstance(values, dict) else 0.0
                    val_2024 = values.get('2024', 0.0) if isinstance(values, dict) else 0.0
                    
                    # Only update if we have non-zero values
                    if val_2025 != 0.0 or val_2024 != 0.0:
                        ws.cell(row=row_num, column=4).value = val_2025  # Current year
                        ws.cell(row=row_num, column=5).value = val_2024  # Previous year
                        
                        # Recalculate Peningkatan
                        if val_2024 and val_2024 != 0:
                            peningkatan = ((val_2025 - val_2024) / abs(val_2024)) * 100
                        else:
                            peningkatan = 0 if val_2025 == 0 else 100
                        
                        ws.cell(row=row_num, column=6).value = peningkatan / 100  # Peningkatan
                        
                        print(f"    Updated {data_type}: {bank_name} - 2025={val_2025:,.2f}, 2024={val_2024:,.2f}")
                        break
        except Exception as e:
            print(f"    [ERROR] Error updating row for {bank_name}: {e}")
    
    def _update_rasio_sheet_for_retry(self, ws, bank_name: str, form3_data: dict, year: str, previous_year: str, border):
        """
        Update Rasio sheet with retry data
        
        Args:
            ws: Worksheet object
            bank_name: Bank name
            form3_data: Dict with ratio data
            year: Current year
            previous_year: Previous year
            border: Border style
        """
        try:
            # Find rows with matching bank name and update ratio values
            for row_num in range(3, ws.max_row + 1):
                bank_name_cell = ws.cell(row=row_num, column=1)  # Column A: Nama Bank
                ratio_name_cell = ws.cell(row=row_num, column=2)  # Column B: Rasio
                
                if bank_name_cell.value and str(bank_name_cell.value).strip() == str(bank_name).strip():
                    ratio_name = ratio_name_cell.value
                    if ratio_name and ratio_name in form3_data:
                        values = form3_data[ratio_name]
                        if isinstance(values, dict):
                            val_2025 = values.get('2025', 0.0)
                            # Only update if we have non-zero value
                            if val_2025 != 0.0:
                                ws.cell(row=row_num, column=3).value = val_2025  # %
                                print(f"    Updated Rasio {ratio_name}: {bank_name} - {val_2025:,.2f}")
        except Exception as e:
            print(f"    [ERROR] Error updating Rasio sheet for {bank_name}: {e}")
    
    def _retry_zero_value_banks(self, month: str, year: str):
        """
        Main orchestrator: Check Excel for zero values and retry those banks
        
        This method is called AFTER the main scraping job (phase='all') completes.
        It reads the latest Excel file based on the quarterly logic:
        - Jan, Feb, Mar → December (YYYY-1)
        - Apr, May, Jun → March YYYY
        - Jul, Aug, Sep → June YYYY
        - Oct, Nov, Dec → September YYYY
        
        Args:
            month: Month name (e.g., "Desember") - already determined by quarterly logic
            year: Year (e.g., "2025") - already determined by quarterly logic
        """
        try:
            print("")
            print("=" * 70)
            print("RETRY ZERO VALUE BANKS")
            print("=" * 70)
            print(f"  [INFO] Checking Excel file for: {month} {year}")
            print(f"  [INFO] This uses the same quarterly logic as main scraping")
            print("=" * 70)
            
            # Check if browser is still initialized
            if self.driver is None:
                print("  [WARNING] Browser not initialized, initializing...")
                self.initialize()
            
            # Read Excel for zero values (uses the same month/year as main scraping)
            banks_with_zero = self._read_excel_for_zero_values(month, year)
            
            if not banks_with_zero:
                print("  [INFO] No banks with zero values found, skipping retry")
                return
            
            print(f"  [INFO] Found {len(banks_with_zero)} banks with zero values to retry")
            
            # Retry each bank
            retry_results = {}
            for i, bank_info in enumerate(banks_with_zero, 1):
                bank_name = bank_info['bank_name']
                city = bank_info['city']
                
                print("")
                print(f"  [{i}/{len(banks_with_zero)}] Retrying: {bank_name} ({city})")
                
                # Retry bank
                retry_data = self._retry_bank_with_direct_url(bank_name, month, year)
                
                # Store results
                if retry_data['form1'] or retry_data['form2'] or retry_data['form3']:
                    retry_results[bank_name] = {
                        'city': city,
                        'form1': retry_data['form1'],
                        'form2': retry_data['form2'],
                        'form3': retry_data['form3']
                    }
                
                # Small delay before next bank
                time.sleep(1.0)
            
            # Update Excel with retry results
            if retry_results:
                print("")
                print("  [INFO] Updating Excel with retry results...")
                self._update_excel_with_retry_data(month, year, retry_results)
                print(f"  [OK] Updated {len(retry_results)} banks in Excel")
            else:
                print("  [WARNING] No retry data to update in Excel")
            
            print("")
            print("=" * 70)
            print("RETRY COMPLETED")
            print("=" * 70)
            
        except Exception as e:
            print(f"  [ERROR] Error in retry zero value banks: {e}")
            import traceback
            print(traceback.format_exc())
    
    def unload_selenium(self, kill_processes: bool = True):
        """
        Explicitly unload Selenium and optionally kill lingering processes
        
        Args:
            kill_processes: If True, kill any lingering Chrome/ChromeDriver processes (default: True)
        """
        self.cleanup(kill_processes=kill_processes)
    
    def __enter__(self):
        """Context manager entry"""
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        """Context manager exit"""
        self.cleanup()


if __name__ == "__main__":
    # Example usage
    scraper = OJKExtJSScraper(headless=False)
    try:
        scraper.initialize()
        scraper.navigate_to_page()
        scraper.select_tab_bpr_konvensional()
        scraper.scrape_all_data(month="Desember", year="2024")
    finally:
        scraper.cleanup()

