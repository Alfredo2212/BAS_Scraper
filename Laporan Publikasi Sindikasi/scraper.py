"""
Sindikasi Scraper
Finds listed BPRs from both BPR Konvensional and BPR Syariah report pages
"""

import time
import logging
import re
import shutil
from pathlib import Path
from datetime import datetime
from selenium.webdriver.remote.webdriver import WebDriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from bs4 import BeautifulSoup

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[WARNING] openpyxl not installed. Excel export will not work. Install with: pip install openpyxl")
    Workbook = None

# Handle imports for both package and direct execution
try:
    from helper import ExtJSHelper
    from selenium_setup import SeleniumSetup
except ImportError:
    # If relative imports fail, try absolute imports
    import sys
    from pathlib import Path
    module_dir = Path(__file__).parent.parent / "Laporan Publikasi BPR Konvensional"
    if str(module_dir) not in sys.path:
        sys.path.insert(0, str(module_dir))
    from helper import ExtJSHelper
    from selenium_setup import SeleniumSetup

from config.settings import OJKConfig


class SindikasiScraper:
    """Scraper for finding BPRs in Sindikasi reports"""
    
    # URLs for both report types
    URL_KONVENSIONAL = "https://cfs.ojk.go.id/cfs/Report.aspx?BankTypeCode=BPK&BankTypeName=BPR%20Konvensional"
    URL_SYARIAH = "https://cfs.ojk.go.id/cfs/Report.aspx?BankTypeCode=BPS&BankTypeName=BPR%20Syariah"
    
    def __init__(self, headless: bool = False):
        """
        Initialize the scraper
        
        Args:
            headless: Whether to run browser in headless mode (default: False for visibility)
        """
        self.driver: WebDriver = None
        self.wait: WebDriverWait = None
        self.all_bank_data = []  # Store all bank data for Excel export
        self.extjs: ExtJSHelper = None
        self.headless = headless
        
        # Setup logging
        log_dir = Path(__file__).parent.parent / "logs"
        log_dir.mkdir(exist_ok=True)
        log_file = log_dir / f"sindikasi_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
        
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(log_file, encoding='utf-8'),
                logging.StreamHandler()
            ]
        )
        self.logger = logging.getLogger(__name__)
        self.logger.info(f"Log file: {log_file}")
    
    def initialize(self):
        """Initialize WebDriver and ExtJS helper"""
        if self.driver is None:
            self.driver = SeleniumSetup.create_driver(headless=self.headless)
            self.wait = SeleniumSetup.create_wait(self.driver)
            self.extjs = ExtJSHelper(self.driver, self.wait)
            
            # Don't minimize - keep window visible for monitoring
            self.logger.info("Chrome browser initialized")
    
    def cleanup(self, kill_processes: bool = True):
        """Cleanup resources"""
        if self.driver:
            try:
                # Suppress urllib3 warnings during cleanup
                import warnings
                try:
                    import urllib3
                    urllib3.disable_warnings()
                except ImportError:
                    pass
                
                # Suppress warnings during cleanup
                with warnings.catch_warnings():
                    warnings.simplefilter("ignore")
                    # Small delay to let connections close naturally
                    time.sleep(0.5)
                    self.driver.quit()
                    self.logger.info("Browser closed")
            except Exception as e:
                # Ignore connection errors during cleanup (browser is already closing)
                error_str = str(e).lower()
                if "connection" not in error_str and "refused" not in error_str and "session" not in error_str:
                    self.logger.error(f"Error closing browser: {e}")
                # Connection errors are expected during cleanup, so we silently ignore them
        
        if kill_processes:
            try:
                import sys
                from pathlib import Path
                module_dir = Path(__file__).parent.parent / "Laporan Publikasi BPR Konvensional"
                if str(module_dir) not in sys.path:
                    sys.path.insert(0, str(module_dir))
                from utils import kill_chrome_processes
                kill_chrome_processes()
            except Exception as e:
                self.logger.debug(f"Error killing Chrome processes: {e}")
    
    def read_list_file(self, file_path: Path) -> dict:
        """
        Read and parse list file
        
        Args:
            file_path: Path to the list file
            
        Returns:
            dict with keys: 'scrape' (bool), 'name' (str), 'banks' (list)
        """
        result = {
            'scrape': False,
            'name': '',
            'banks': []
        }
        
        if not file_path.exists():
            self.logger.error(f"List file not found: {file_path}")
            return result
        
        self.logger.info(f"Reading list file: {file_path}")
        
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                lines = f.readlines()
            
            for line in lines:
                line = line.strip()
                
                # Skip empty lines
                if not line:
                    continue
                
                # Parse SCRAPE flag
                if line.upper().startswith('SCRAPE'):
                    scrape_match = re.search(r'SCRAPE\s*=\s*(TRUE|FALSE)', line, re.IGNORECASE)
                    if scrape_match:
                        result['scrape'] = scrape_match.group(1).upper() == 'TRUE'
                
                # Parse NAME
                elif line.upper().startswith('NAME'):
                    name_match = re.search(r'NAME\s*=\s*(.+)', line, re.IGNORECASE)
                    if name_match:
                        result['name'] = name_match.group(1).strip()
                
                # Bank names (everything else after NAME)
                elif result['name']:  # Only add banks after NAME is found
                    result['banks'].append(line)
            
            self.logger.info(f"SCRAPE = {result['scrape']}, NAME = {result['name']}")
            self.logger.info(f"Found {len(result['banks'])} BPRs in list")
            
            return result
            
        except Exception as e:
            self.logger.error(f"Error reading list file: {e}")
            return result
    
    def navigate_to_page(self, url: str):
        """Navigate to a report page and wait for ExtJS to load"""
        if self.driver is None:
            self.initialize()
        
        # Switch to default content in case we're in an iframe
        try:
            self.driver.switch_to.default_content()
        except:
            pass
        
        self.logger.info(f"Navigating to: {url}")
        self.driver.get(url)
        
        # Wait for page to load
        time.sleep(0.75)
        
        # Check if page is in iframe or main page
        max_attempts = 10
        for attempt in range(max_attempts):
            try:
                if self.extjs.check_extjs_available():
                    self.logger.debug("ExtJS is available in main page")
                    return
            except:
                pass
            time.sleep(0.75)
        
        # If not in main page, check for iframes
        iframes = self.driver.find_elements(By.TAG_NAME, "iframe")
        if iframes:
            self.logger.debug(f"Found {len(iframes)} iframe(s), checking inside...")
            for i, iframe in enumerate(iframes):
                try:
                    self.driver.switch_to.frame(iframe)
                    time.sleep(1.125)
                    
                    if self.extjs.check_extjs_available():
                        self.logger.debug(f"ExtJS is available in iframe {i+1}")
                        return
                    
                    self.driver.switch_to.default_content()
                except:
                    self.driver.switch_to.default_content()
                    continue
        
        # Final check
        time.sleep(0.75)
        if self.extjs.check_extjs_available():
            self.logger.debug("ExtJS is now available")
            return
        
        self.logger.warning("ExtJS not immediately available, continuing anyway...")
    
    def _format_bank_code_for_url(self, bank_name: str) -> list:
        """
        Format bank name for URL with fallback options.
        Returns a list of possible bank code formats to try.
        
        Bank names always start with either "PT" or "Perumda".
        If it doesn't start with "PT", it should start with "Perumda".
        
        Args:
            bank_name: Bank name (e.g., "PT Bpr Rangkiang Aur Denai" or "Bprs Al-Makmur")
            
        Returns:
            List of formatted bank codes: [original_format, expanded_format]
            If no expansion needed, returns [original_format]
        """
        # Step 1: Replace "-" with spaces
        name = bank_name.replace("-", " ")
        
        # Step 2: Split by spaces and clean
        words = [w.strip() for w in name.split() if w.strip()]
        
        if not words:
            return [""]
        
        # Step 3: Ensure name starts with "PT" or "Perumda"
        first_word_upper = words[0].upper()
        if first_word_upper not in ["PT", "PERUMDA"]:
            # Prepend "PT" if it doesn't start with PT or Perumda
            words.insert(0, "PT")
        
        # Step 4: Handle PERUMDA case (normalize to "Perumda")
        if words[0].upper() == "PERUMDA":
            words[0] = "Perumda"
        
        # Step 5: Create original format (keep Bpr/Bprs as-is)
        # PT should always be uppercase, other words capitalized
        original_words = []
        for w in words:
            if w.upper() == "PT":
                original_words.append("PT")
            else:
                original_words.append(w.capitalize())
        original_format = "+".join(original_words)
        
        # Step 6: Check if we need expanded format
        has_bpr = False
        has_bprs = False
        expanded_words = []
        
        for word in words:
            word_upper = word.upper()
            if word_upper in ["BPR", "BPRS"]:
                if word_upper == "BPRS":
                    has_bprs = True
                    expanded_words.extend(["Bank", "Perekonomian", "Rakyat", "Syariah"])
                else:
                    has_bpr = True
                    expanded_words.extend(["Bank", "Perekonomian", "Rakyat"])
            else:
                expanded_words.append(word)
        
        # Step 7: Create expanded format if needed
        formats = [original_format]
        
        if has_bpr or has_bprs:
            # Capitalize expanded words, but keep PT uppercase
            expanded_formatted = []
            for w in expanded_words:
                if w.upper() == "PT":
                    expanded_formatted.append("PT")
                else:
                    expanded_formatted.append(w.capitalize())
            expanded_format = "+".join(expanded_formatted)
            formats.append(expanded_format)
        
        return formats
    
    def _get_target_month_year(self) -> tuple[str, str]:
        """
        Determine target month and year based on current date.
        Uses quarterly mapping logic:
        - Jan, Feb, Mar → use December (YYYY-1)
        - Apr, May, Jun → use March YYYY
        - Jul, Aug, Sep → use June YYYY
        - Oct, Nov, Dec → use September YYYY
        Where YYYY is the current server year.
        
        Returns:
            Tuple of (month_name, year) e.g., ("September", "2025")
        """
        from datetime import datetime
        
        now = datetime.now()
        current_month = now.month
        current_year = now.year
        
        # Quarterly mapping logic
        month_names = {
            3: "Maret",      # 03
            6: "Juni",       # 06
            9: "September",  # 09
            12: "Desember"   # 12
        }
        
        # Map current month to quarterly report month
        if current_month in [1, 2, 3]:  # Jan, Feb, Mar
            target_month_num = 12  # Desember
            target_year = current_year - 1
        elif current_month in [4, 5, 6]:  # Apr, May, Jun
            target_month_num = 3  # Maret
            target_year = current_year
        elif current_month in [7, 8, 9]:  # Jul, Aug, Sep
            target_month_num = 6  # Juni
            target_year = current_year
        else:  # Oct, Nov, Dec
            target_month_num = 9  # September
            target_year = current_year
        
        month_name = month_names[target_month_num]
        
        self.logger.info(f"Current date: {now.strftime('%B %Y')}")
        self.logger.info(f"Target month/year: {month_name} {target_year}")
        
        return month_name, str(target_year)
    
    def _month_name_to_number(self, month_name: str) -> int:
        """
        Convert month name to number
        
        Args:
            month_name: Month name (Maret, Juni, September, Desember)
            
        Returns:
            Month number (3, 6, 9, or 12)
        """
        month_map = {
            "Maret": 3,
            "Juni": 6,
            "September": 9,
            "Desember": 12
        }
        return month_map.get(month_name, 3)  # Default to 3 if not found
    
    
    def _build_report_url(self, bank_code: str, month: int, year: str, bank_type: str, form_number: int) -> str:
        """
        Build the report viewer URL
        
        Args:
            bank_code: Formatted bank code (e.g., "PT+Bpr+Rangkiang+Aur+Denai")
            month: Month number (3, 6, 9, or 12)
            year: Year (e.g., "2025")
            bank_type: "syariah" or "konvensional"
            form_number: Form number (1, 2, 3, or 4)
            
        Returns:
            Complete URL string
        """
        base_url = "https://cfs.ojk.go.id/cfs/ReportViewerForm.aspx"
        
        # Determine form code prefix
        if bank_type.lower() == "syariah":
            form_code = f"BPS-901-{form_number:06d}"
        else:
            form_code = f"BPK-901-{form_number:06d}"
        
        # Build URL parameters
        params = {
            "BankCodeNumber": "620000",  # Static placeholder
            "BankCode": bank_code,
            "Month": str(month),
            "Year": year,
            "FinancialReportPeriodTypeCode": "R",
            "FinancialReportTypeCode": form_code
        }
        
        # Build query string
        query_string = "&".join([f"{k}={v}" for k, v in params.items()])
        url = f"{base_url}?{query_string}"
        
        return url
    
    def _check_for_server_error(self) -> bool:
        """
        Check if the current page has a server error
        
        Returns:
            True if server error found, False otherwise
        """
        try:
            page_source = self.driver.page_source
            error_text = "Server Error in '/cfs' Application"
            
            if error_text in page_source:
                self.logger.debug("  Server error detected in page source")
                return True
            
            return False
        except Exception as e:
            self.logger.debug(f"  Error checking for server error: {e}")
            return False
    
    def _remove_bpr_prefix(self, bank_name: str) -> str:
        """
        Remove BPR/BPRS prefix from bank name for matching
        
        Args:
            bank_name: Bank name with or without prefix
            
        Returns:
            Bank name without BPR/BPRS prefix
        """
        # Remove common prefixes
        name = bank_name.strip()
        name = re.sub(r'^(BPRS|BPR|PT\s*BANK|BANK)\s+', '', name, flags=re.IGNORECASE)
        return name.strip()
    
    
    def _check_identifiers(self, identifiers: list) -> bool:
        """
        Check if all identifiers are present in the page using BeautifulSoup
        
        Args:
            identifiers: List of identifier strings to check for (exact match)
            
        Returns:
            True if all identifiers are found, False otherwise
        """
        try:
            # Get page source and parse with BeautifulSoup
            page_source = self.driver.page_source
            soup = BeautifulSoup(page_source, 'html.parser')
            
            # Get all text from the page
            page_text = soup.get_text()
            
            # Check for each identifier (exact match)
            found_identifiers = []
            missing_identifiers = []
            
            for identifier in identifiers:
                if identifier in page_text:
                    found_identifiers.append(identifier)
                else:
                    missing_identifiers.append(identifier)
            
            if missing_identifiers:
                self.logger.debug(f"  Missing identifiers: {missing_identifiers}")
                return False
            else:
                self.logger.info(f"  [OK] All identifiers found: {found_identifiers}")
                return True
                
        except Exception as e:
            self.logger.error(f"  [ERROR] Error checking identifiers: {e}")
            return False
    
    def _clean_numeric_text(self, text: str) -> float:
        """
        Clean numeric text by reversing . and , then parse as float
        
        Args:
            text: Text containing numeric value (e.g., "230,115,190" becomes "230.115.190")
            
        Returns:
            Parsed float value, or 0.0 if parsing fails
        """
        try:
            # Remove whitespace
            cleaned = text.strip()
            
            # Reverse . and , (swap all occurrences)
            # "230,115,190" → "230.115.190"
            # "230.115.190" → "230,115,190"
            cleaned = cleaned.replace('.', 'TEMP_DOT')
            cleaned = cleaned.replace(',', '.')
            cleaned = cleaned.replace('TEMP_DOT', ',')
            
            # After swapping, parse the number
            # Determine which separator is decimal and which is thousands
            # Pattern: if there are multiple separators, the last one is usually decimal
            # If only one separator, it could be either
            
            dot_count = cleaned.count('.')
            comma_count = cleaned.count(',')
            
            if dot_count > 0 and comma_count > 0:
                # Both present - last separator is decimal
                last_dot_pos = cleaned.rfind('.')
                last_comma_pos = cleaned.rfind(',')
                
                if last_dot_pos > last_comma_pos:
                    # Last separator is dot, so dots are thousands, comma is decimal
                    # Remove all dots (thousands), replace comma with dot for parsing
                    cleaned = cleaned.replace('.', '')
                    cleaned = cleaned.replace(',', '.')
                else:
                    # Last separator is comma, so commas are thousands, dot is decimal
                    # Remove all commas (thousands), keep dot as decimal
                    cleaned = cleaned.replace(',', '')
            elif dot_count > 1:
                # Multiple dots - all are thousands separators, remove all
                cleaned = cleaned.replace('.', '')
            elif comma_count > 1:
                # Multiple commas - all are thousands separators, remove all
                cleaned = cleaned.replace(',', '')
            elif dot_count == 1:
                # Single dot - could be decimal or thousands
                # Check position: if near end (last 3 chars), likely decimal, otherwise thousands
                dot_pos = cleaned.find('.')
                if dot_pos >= len(cleaned) - 3:
                    # Near end, likely decimal separator
                    pass  # Keep it
                else:
                    # Likely thousands separator, remove it
                    cleaned = cleaned.replace('.', '')
            elif comma_count == 1:
                # Single comma - could be decimal or thousands
                # Check position: if near end (last 3 chars), likely decimal, otherwise thousands
                comma_pos = cleaned.find(',')
                if comma_pos >= len(cleaned) - 3:
                    # Near end, likely decimal separator, convert to dot
                    cleaned = cleaned.replace(',', '.')
                else:
                    # Likely thousands separator, remove it
                    cleaned = cleaned.replace(',', '')
            
            # Remove any remaining non-numeric characters except decimal point and minus
            cleaned = re.sub(r'[^\d.\-]', '', cleaned)
            
            if not cleaned or cleaned == '-':
                return 0.0
            
            return float(cleaned)
        except Exception as e:
            self.logger.debug(f"  Error parsing numeric text '{text}': {e}")
            return 0.0
    
    def _extract_identifier_value_from_table(self, soup: BeautifulSoup, identifier: str) -> dict:
        """
        Extract values for a single identifier from table structure (td/tr).
        Similar to how publikasi scraper extracts kredit values.
        - Find div containing identifier text
        - Find parent <td> and <tr>
        - Look at next <td> elements for numeric values
        - First td = current year (2025), second td = previous year (2024)
        
        Args:
            soup: BeautifulSoup parsed page
            identifier: Identifier text to find
            
        Returns:
            dict with {'2025': value, '2024': value} or {'2025': 0.0, '2024': 0.0} if not found
        """
        result = {'2025': 0.0, '2024': 0.0}
        
        try:
            # Helper function to extract number from text
            def extract_number(text: str) -> float:
                """Extract an integer-like number from Indonesian-style formatted text."""
                if not text:
                    return 0.0
                
                # Normalize spaces
                text = text.replace('\xa0', ' ').replace('&nbsp;', ' ').strip()
                
                # Keep digits only (we just need whole numbers)
                digits_only = re.sub(r'\D', '', text)
                if not digits_only:
                    return 0.0
                
                value = float(digits_only)
                return value
            
            # Find the <div> whose text contains our identifier
            label_div = None
            for div in soup.find_all('div'):
                text = div.get_text(strip=True)
                if not text:
                    continue
                
                # Skip divs that are too long (likely contain entire page content)
                if len(text) > 5000:
                    continue
                
                # Check if identifier is in text
                if identifier.lower() in text.lower():
                    # Prefer divs where identifier is a significant part of the text
                    text_lower = text.lower()
                    identifier_lower = identifier.lower()
                    
                    # If text is short or identifier is at the start/end, it's likely the right div
                    if len(text) < 200 or text_lower.startswith(identifier_lower) or text_lower.endswith(identifier_lower):
                        label_div = div
                        self.logger.debug(f"    Found identifier '{identifier}' in <div>: '{text[:100]}...'")
                        break
            
            if not label_div:
                self.logger.debug(f"    Identifier '{identifier}' NOT FOUND in page")
                return result
            
            # Find the parent <td> element
            parent_td = label_div.find_parent('td')
            if not parent_td:
                self.logger.debug(f"    Identifier '{identifier}' found but not in a <td> element")
                return result
            
            # Find the parent <tr> (table row) to get all <td> elements
            parent_tr = parent_td.find_parent('tr')
            if not parent_tr:
                self.logger.debug(f"    Identifier '{identifier}' found but not in a <tr> element")
                return result
            
            # Get all <td> elements in the row
            tds = parent_tr.find_all('td')
            
            # Find the index of the identifier <td>
            try:
                identifier_td_index = tds.index(parent_td)
            except:
                self.logger.debug(f"    Identifier '{identifier}' found but couldn't find its td index")
                return result
            
            self.logger.debug(f"    Found identifier '{identifier}' at td index {identifier_td_index}, checking next tds...")
            
            # Extract values from subsequent <td> elements (skip the identifier <td>)
            numeric_count = 0
            for td_idx, td in enumerate(tds[identifier_td_index + 1:identifier_td_index + 10], start=identifier_td_index + 1):  # Check up to 10 <td> elements
                if numeric_count >= 2:  # We need 2 values
                    break
                
                # Get text from the <td> (check for div first, then direct text)
                td_div = td.find('div', recursive=False)
                if not td_div:
                    td_div = td.find('div')
                
                if td_div:
                    td_text = td_div.get_text(strip=True)
                else:
                    td_text = td.get_text(strip=True)
                
                # Remove &nbsp; entities and check if empty
                td_text_clean = td_text.replace('\xa0', ' ').replace('&nbsp;', ' ').strip()
                
                if not td_text_clean:
                    continue
                
                # Skip if it's clearly another identifier
                if any(keyword in td_text_clean.lower() for keyword in ['kepada', 'pihak', 'bank', 'bpr', 'report viewer', 'configuration error', 'piutang', 'aset', 'dpk', 'laba', 'rasio']):
                    continue
                
                # Extract number from text
                number = extract_number(td_text_clean)
                
                # Validate the number is reasonable
                if number >= 0 and number < 1e15 and number != float('inf'):
                    if numeric_count == 0:
                        result['2025'] = number
                        numeric_count += 1
                        self.logger.debug(f"    td[{td_idx}]: found 2025 value = {number}")
                    elif numeric_count == 1:
                        result['2024'] = number
                        numeric_count += 1
                        self.logger.debug(f"    td[{td_idx}]: found 2024 value = {number}")
                elif number == 0 and any(char.isdigit() for char in td_text_clean) and len(td_text_clean) < 50:
                    # Zero value is valid if it's a short text with digits
                    if numeric_count == 0:
                        result['2025'] = number
                        numeric_count += 1
                        self.logger.debug(f"    td[{td_idx}]: found 2025 value = {number} (zero)")
                    elif numeric_count == 1:
                        result['2024'] = number
                        numeric_count += 1
                        self.logger.debug(f"    td[{td_idx}]: found 2024 value = {number} (zero)")
            
            if result['2025'] == 0.0 and result['2024'] == 0.0:
                self.logger.debug(f"    Identifier '{identifier}' found but no numeric values extracted from next tds")
            elif result['2024'] == 0.0:
                self.logger.debug(f"    Identifier '{identifier}' found but only 1 value extracted")
            
        except Exception as e:
            self.logger.debug(f"    Error extracting identifier '{identifier}' from table: {e}")
            import traceback
            self.logger.debug(traceback.format_exc())
        
        return result
    
    def _extract_ratio_value(self, soup: BeautifulSoup, identifier: str) -> dict:
        """
        Extract rasio values preserving decimal points (like publikasi targeted scraping).
        Uses div-based extraction with _clean_numeric_text to properly handle decimals.
        
        Args:
            soup: BeautifulSoup parsed page
            identifier: Identifier text to find
            
        Returns:
            dict with {'2025': value, '2024': value} or {'2025': 0.0, '2024': 0.0} if not found
        """
        result = {'2025': 0.0, '2024': 0.0}
        
        try:
            # Helper function to extract number preserving decimal points (like publikasi)
            def extract_decimal_number(text: str) -> float:
                """Extract a number preserving decimal points using _clean_numeric_text."""
                if not text:
                    return 0.0
                
                # Normalize spaces
                text = text.replace('\xa0', ' ').replace('&nbsp;', ' ').strip()
                
                # Check for negative (in parentheses)
                is_negative = False
                if text.startswith('(') and text.endswith(')'):
                    is_negative = True
                    text = text[1:-1].strip()
                elif text.startswith('-'):
                    is_negative = True
                    text = text[1:].strip()
                
                # Use _clean_numeric_text to properly handle Indonesian format (handles both whole numbers and decimals)
                number = self._clean_numeric_text(text)
                if is_negative:
                    number = -number
                return number
            
            # Find the <div> whose text contains our identifier
            label_div = None
            all_divs = soup.find_all('div')
            
            for div in all_divs:
                text = div.get_text(strip=True)
                if not text:
                    continue
                
                # Skip divs that are too long (likely contain entire page content)
                if len(text) > 5000:
                    continue
                
                # Check if identifier is in text
                if identifier.lower() in text.lower():
                    # Prefer divs where identifier is a significant part of the text
                    text_lower = text.lower()
                    identifier_lower = identifier.lower()
                    
                    # If text is short or identifier is at the start/end, it's likely the right div
                    if len(text) < 200 or text_lower.startswith(identifier_lower) or text_lower.endswith(identifier_lower):
                        label_div = div
                        self.logger.debug(f"    Found identifier '{identifier}' in <div>: '{text[:100]}...'")
                        break
            
            if not label_div:
                self.logger.debug(f"    Identifier '{identifier}' NOT FOUND in page")
                return result
            
            # Find the index of this div and get the next divs that contain numeric values
            try:
                div_index = all_divs.index(label_div)
                # Look through next divs to find numeric values (up to 10 divs ahead)
                numeric_count = 0
                for j in range(1, min(11, len(all_divs) - div_index)):  # Check up to 10 next divs
                    if numeric_count >= 2:  # We need 2 values
                        break
                    
                    if div_index + j < len(all_divs):
                        next_div = all_divs[div_index + j]
                        div_text = next_div.get_text(strip=True)
                        
                        if not div_text:
                            continue
                        
                        # Skip very long divs (likely contain entire page content)
                        if len(div_text) > 5000:
                            continue
                        
                        # Skip if it's clearly another identifier (contains common identifier keywords)
                        if any(keyword in div_text.lower() for keyword in ['kepada', 'pihak', 'bank', 'bpr', 'report viewer', 'configuration error', 'piutang', 'aset', 'dpk', 'laba', 'rasio']):
                            continue
                        
                        # Skip if too long (likely not a single ratio value)
                        if len(div_text) > 100:
                            continue
                        
                        # Extract number preserving decimal point
                        number = extract_decimal_number(div_text)
                        
                        # Validate the number is reasonable
                        if number != 0.0 or (number == 0.0 and any(char.isdigit() for char in div_text) and len(div_text) < 50):
                            if numeric_count == 0:
                                result['2025'] = number
                                numeric_count += 1
                                self.logger.debug(f"    Found 2025 value: {number} from '{div_text}'")
                            elif numeric_count == 1:
                                result['2024'] = number
                                numeric_count += 1
                                self.logger.debug(f"    Found 2024 value: {number} from '{div_text}'")
            except ValueError:
                self.logger.debug(f"    Identifier '{identifier}' found but couldn't find its index")
                return result
            
            if result['2025'] == 0.0 and result['2024'] == 0.0:
                self.logger.debug(f"    Identifier '{identifier}' found but no numeric values extracted from next divs")
            elif result['2024'] == 0.0:
                self.logger.debug(f"    Identifier '{identifier}' found but only 1 value extracted")
            
        except Exception as e:
            self.logger.debug(f"    Error extracting ratio '{identifier}': {e}")
            import traceback
            self.logger.debug(traceback.format_exc())
        
        return result
    
    def _extract_identifier_value(self, soup: BeautifulSoup, identifier: str) -> dict:
        """
        Extract values for a single identifier from the page
        Uses the same approach as BPR Konvensional scraper:
        - Find div containing identifier text
        - Look at next divs to find numeric values
        - First numeric = current year, second = previous year
        
        Args:
            soup: BeautifulSoup parsed page
            identifier: Identifier text to find
            
        Returns:
            dict with {'2025': value, '2024': value} or {'2025': 0.0, '2024': 0.0} if not found
        """
        result = {'2025': 0.0, '2024': 0.0}
        
        try:
            # Helper function to extract number from text
            def extract_number(text: str) -> float:
                """Extract an integer-like number from Indonesian-style formatted text."""
                if not text:
                    return 0.0
                
                original_text = text
                # Normalize spaces
                text = text.replace('\xa0', ' ').replace('&nbsp;', ' ').strip()
                
                # Keep digits only (we just need whole numbers)
                digits_only = re.sub(r'\D', '', text)
                if not digits_only:
                    return 0.0
                
                value = float(digits_only)
                return value
            
            # Helper function to split concatenated numbers (current year + previous year)
            def split_concatenated_numbers(text: str) -> tuple[str, str]:
                """
                Split concatenated numbers like "23,122,1223,112,122" into two numbers.
                Format: After comma, max 3 digits. If next is comma, same year. If next is digit, next year starts.
                """
                if not text or ',' not in text:
                    return text, ""
                
                # Find the split point: look for pattern where after comma, we have 3 digits, then a digit (not comma)
                pattern = r',(\d{3})(\d)(?=,|\d|$)'
                match = re.search(pattern, text)
                
                if match:
                    split_pos = match.end(1)  # Position after the 3 digits (before the next digit)
                    current_year_text = text[:split_pos].rstrip(',')
                    previous_year_text = text[split_pos:].lstrip(',')
                    return current_year_text, previous_year_text
                
                return text, ""
            
            # Find the <div> whose text contains our identifier
            label_div = None
            all_divs = soup.find_all('div')
            
            for div in all_divs:
                text = div.get_text(strip=True)
                if not text:
                    continue
                
                # Skip divs that are too long (likely contain entire page content)
                if len(text) > 5000:
                    continue
                
                # Check if identifier is in text
                if identifier.lower() in text.lower():
                    # Prefer divs where identifier is a significant part of the text
                    text_lower = text.lower()
                    identifier_lower = identifier.lower()
                    
                    # If text is short or identifier is at the start/end, it's likely the right div
                    if len(text) < 200 or text_lower.startswith(identifier_lower) or text_lower.endswith(identifier_lower):
                        label_div = div
                        self.logger.debug(f"    Found identifier '{identifier}' in <div>: '{text[:100]}...'")
                        break
            
            if not label_div:
                self.logger.debug(f"    Identifier '{identifier}' NOT FOUND in page")
                return result
            
            # Find the index of this div and get the next divs that contain numeric values
            try:
                div_index = all_divs.index(label_div)
                # Look through next divs to find numeric values (up to 10 divs ahead)
                numeric_count = 0
                for j in range(1, min(11, len(all_divs) - div_index)):  # Check up to 10 next divs
                    if numeric_count >= 2:  # We need 2 values
                        break
                    
                    if div_index + j < len(all_divs):
                        next_div = all_divs[div_index + j]
                        div_text = next_div.get_text(strip=True)
                        
                        if not div_text:
                            continue
                        
                        # Skip very long divs (likely contain entire page content)
                        if len(div_text) > 5000:
                            continue
                        
                        # Skip if it's clearly another identifier (contains common identifier keywords)
                        if any(keyword in div_text.lower() for keyword in ['kepada', 'pihak', 'bank', 'bpr', 'report viewer', 'configuration error', 'piutang', 'aset', 'dpk', 'laba', 'rasio']):
                            continue
                        
                        # Check if it contains concatenated numbers (has comma and might have split pattern)
                        if ',' in div_text and len(div_text) > 10:
                            # Try to split concatenated numbers
                            current_year_text, prev_year_text = split_concatenated_numbers(div_text)
                            
                            if prev_year_text:
                                # Found concatenated numbers, extract both
                                current_number = extract_number(current_year_text)
                                prev_number = extract_number(prev_year_text)
                                
                                # Validate numbers are reasonable
                                if (current_number >= 0 and current_number < 1e15 and current_number != float('inf') and
                                    prev_number >= 0 and prev_number < 1e15 and prev_number != float('inf')):
                                    
                                    if numeric_count == 0:
                                        result['2025'] = current_number
                                        numeric_count += 1
                                    
                                    if numeric_count == 1:
                                        result['2024'] = prev_number
                                        numeric_count += 1
                                    
                                    if numeric_count >= 2:
                                        break
                                    continue
                        
                        # Check if this div looks like it contains a single formatted number
                        if len(div_text) > 100:
                            continue
                        
                        # Extract number from single number text
                        number = extract_number(div_text)
                        
                        # Validate the number is reasonable
                        if number >= 0 and number < 1e15 and number != float('inf'):
                            if numeric_count == 0:
                                result['2025'] = number
                                numeric_count += 1
                            elif numeric_count == 1:
                                result['2024'] = number
                                numeric_count += 1
                        elif number == 0 and any(char.isdigit() for char in div_text) and len(div_text) < 50:
                            # Zero value is valid if it's a short text with digits
                            if numeric_count == 0:
                                result['2025'] = number
                                numeric_count += 1
                            elif numeric_count == 1:
                                result['2024'] = number
                                numeric_count += 1
            except ValueError:
                self.logger.debug(f"    Identifier '{identifier}' found but couldn't find its index")
                return result
            
            if result['2025'] == 0.0 and result['2024'] == 0.0:
                self.logger.debug(f"    Identifier '{identifier}' found but no numeric values extracted from next divs")
            elif result['2024'] == 0.0:
                self.logger.debug(f"    Identifier '{identifier}' found but only 1 value extracted")
            
        except Exception as e:
            self.logger.debug(f"    Error extracting identifier '{identifier}': {e}")
            import traceback
            self.logger.debug(traceback.format_exc())
        
        return result
    
    def _get_page_source_with_iframe(self) -> tuple[str, object]:
        """
        Get page source, checking iframes first (similar to BPR Konvensional)
        
        Returns:
            Tuple of (page_source: str, report_iframe: object or None)
        """
        # Try to find report in iframe first, then main page
        iframes = self.driver.find_elements(By.TAG_NAME, "iframe")
        page_source = None
        report_iframe = None
        
        # Check iframes for report content
        for iframe in iframes:
            try:
                self.driver.switch_to.frame(iframe)
                iframe_source = self.driver.page_source
                if ("Piutang" in iframe_source or "Aset" in iframe_source or "DPK" in iframe_source or
                    "LABA" in iframe_source or "Rasio" in iframe_source or "KPMM" in iframe_source or
                    "Kredit" in iframe_source):
                    self.logger.debug("    Found report content in iframe")
                    page_source = iframe_source
                    report_iframe = iframe
                    break
                self.driver.switch_to.default_content()
            except:
                self.driver.switch_to.default_content()
                continue
        
        # If not in iframe, use main page
        if page_source is None:
            self.logger.debug("    Using main page source")
            self.driver.switch_to.default_content()
            page_source = self.driver.page_source
        
        return page_source, report_iframe
    
    def _parse_syariah_form1(self) -> dict:
        """
        Parse Syariah form 1 (BPS-901-000001) data
        
        Extracts:
        - ASET: Total Aset
        - PIUTANG: Sum of (a) Piutang Murabahah, (b) Piutang Istishna, (c) Piutang Multijasa, (d) Piutang Qardh, (e) Piutang Sewa
        - DPK: Sum of (1) Liabilitas Segera, (2) Tabungan Wadiah, (3) Simpanan Mudarabah
        
        Returns:
            dict with structure:
            {
                'ASET': {'2025': value, '2024': value},
                'PIUTANG': {'2025': sum, '2024': sum, 'individual': {...}},
                'DPK': {'2025': sum, '2024': sum, 'individual': {...}}
            }
        """
        result = {
            'ASET': {'2025': 0.0, '2024': 0.0},
            'PIUTANG': {'2025': 0.0, '2024': 0.0, 'individual': {}},
            'DPK': {'2025': 0.0, '2024': 0.0, 'individual': {}}
        }
        
        try:
            # Wait a bit for page to load
            time.sleep(3.0)
            
            # Get page source (check iframes first)
            page_source, report_iframe = self._get_page_source_with_iframe()
            soup = BeautifulSoup(page_source, 'html.parser')
            
            # Extract ASET
            aset_values = self._extract_identifier_value(soup, "Total Aset")
            result['ASET'] = aset_values
            self.logger.info(f"    ASET (Total Aset): 2025={aset_values['2025']:,.2f}, 2024={aset_values['2024']:,.2f}")
            
            # Extract PIUTANG components
            piutang_identifiers = [
                "Piutang Murabahah",
                "Piutang Istishna",
                "Piutang Multijasa",
                "Piutang Qardh",
                "Piutang Sewa"
            ]
            
            for identifier in piutang_identifiers:
                values = self._extract_identifier_value(soup, identifier)
                result['PIUTANG']['individual'][identifier] = values
                result['PIUTANG']['2025'] += values['2025']
                result['PIUTANG']['2024'] += values['2024']
                self.logger.info(f"    {identifier}: 2025={values['2025']:,.2f}, 2024={values['2024']:,.2f}")
            
            self.logger.info(f"    PIUTANG Total: 2025={result['PIUTANG']['2025']:,.2f}, 2024={result['PIUTANG']['2024']:,.2f}")
            
            # Extract DPK components
            # For BPRS, use table structure extraction (like kredit) to get correct 2024 values
            dpk_identifiers = [
                "Liabilitas Segera",
                "Tabungan Wadiah",
                "Simpanan Mudarabah"
            ]
            
            for identifier in dpk_identifiers:
                # Use table structure extraction for BPRS DPK (like kredit extraction)
                values = self._extract_identifier_value_from_table(soup, identifier)
                result['DPK']['individual'][identifier] = values
                result['DPK']['2025'] += values['2025']
                result['DPK']['2024'] += values['2024']
                self.logger.info(f"    {identifier}: 2025={values['2025']:,.2f}, 2024={values['2024']:,.2f}")
            
            self.logger.info(f"    DPK Total: 2025={result['DPK']['2025']:,.2f}, 2024={result['DPK']['2024']:,.2f}")
            
            return result
            
        except Exception as e:
            self.logger.error(f"  [ERROR] Error parsing Syariah form 1: {e}")
            import traceback
            self.logger.debug(traceback.format_exc())
            return result
    
    def _parse_syariah_form2(self) -> dict:
        """
        Parse Syariah form 2 (BPS-901-000002) data
        
        Extracts:
        - LABA KOTOR: Laba Rugi Tahun Berjalan
        - LABA BERSIH: Laba Rugi Bersih
        
        Returns:
            dict with structure:
            {
                'LABA KOTOR': {'2025': value, '2024': value},
                'LABA BERSIH': {'2025': value, '2024': value}
            }
        """
        result = {
            'LABA KOTOR': {'2025': 0.0, '2024': 0.0},
            'LABA BERSIH': {'2025': 0.0, '2024': 0.0}
        }
        
        try:
            # Wait a bit for page to load
            time.sleep(3.0)
            
            # Get page source (check iframes first)
            page_source, report_iframe = self._get_page_source_with_iframe()
            soup = BeautifulSoup(page_source, 'html.parser')
            
            # Extract LABA KOTOR
            laba_kotor_values = self._extract_identifier_value(soup, "Laba Rugi Tahun Berjalan")
            result['LABA KOTOR'] = laba_kotor_values
            self.logger.info(f"    LABA KOTOR: 2025={laba_kotor_values['2025']:,.2f}, 2024={laba_kotor_values['2024']:,.2f}")
            
            # Extract LABA BERSIH
            laba_bersih_values = self._extract_identifier_value(soup, "Laba Rugi Bersih")
            result['LABA BERSIH'] = laba_bersih_values
            self.logger.info(f"    LABA BERSIH: 2025={laba_bersih_values['2025']:,.2f}, 2024={laba_bersih_values['2024']:,.2f}")
            
            return result
            
        except Exception as e:
            self.logger.error(f"  [ERROR] Error parsing Syariah form 2: {e}")
            import traceback
            self.logger.debug(traceback.format_exc())
            return result
    
    def _parse_syariah_form3(self) -> dict:
        """
        Parse Syariah form 3 (BPS-901-000003) data
        
        Extracts ratios (separate, not summed):
        - KPMM: Kewajiban Penyediaan Modal Minimum (KPMM)
        - NPF Neto: Non Performing Financing (NPF) Neto
        - ROA: Return on Asset (ROA)
        - BOPO: Beban Operasional terhadap Pendapatan Operasional (BOPO)
        - NI: Net Imbalan (NI)
        - FDR: Financing to Deposit Ratio (FDR)
        - Cash Ratio: Cash Ratio
        
        Returns:
            dict with structure:
            {
                'KPMM': {'2025': value, '2024': value},
                'NPF Neto': {'2025': value, '2024': value},
                ...
            }
        """
        result = {}
        
        ratios = [
            ("KPMM", "Kewajiban Penyediaan Modal Minimum (KPMM)"),
            ("NPF Neto", "Non Performing Financing (NPF) Neto"),
            ("ROA", "Return on Asset (ROA)"),
            ("BOPO", "Beban Operasional terhadap Pendapatan Operasional (BOPO)"),
            ("NI", "Net Imbalan (NI)"),
            ("FDR", "Financing to Deposit Ratio (FDR)"),
            ("Cash Ratio", "Cash Ratio")
        ]
        
        try:
            # Wait a bit for page to load
            time.sleep(3.0)
            
            # Get page source (check iframes first)
            page_source, report_iframe = self._get_page_source_with_iframe()
            soup = BeautifulSoup(page_source, 'html.parser')
            
            for ratio_name, identifier in ratios:
                # Use ratio extraction that preserves decimal points (like publikasi targeted scraping)
                values = self._extract_ratio_value(soup, identifier)
                result[ratio_name] = values
                self.logger.info(f"    {ratio_name}: 2025={values['2025']:.2f}, 2024={values['2024']:.2f}")
            
            return result
            
        except Exception as e:
            self.logger.error(f"  [ERROR] Error parsing Syariah form 3: {e}")
            import traceback
            self.logger.debug(traceback.format_exc())
            return result
    
    def _parse_konvensional_form1(self) -> dict:
        """
        Parse Konvensional form 1 (BPK-901-000001) data
        
        Extracts:
        - ASET: Total Aset
        - KREDIT: Sum of (a) Kepada BPR, (b) Kepada Bank Umum, (c) Kepada non bank – pihak terkait, (d) Kepada non bank – pihak tidak terkait
        - DPK: Sum of (a) Tabungan, (b) Deposito, (c) Simpanan dari Bank Lain
        
        Returns:
            dict with structure:
            {
                'ASET': {'2025': value, '2024': value},
                'KREDIT': {'2025': sum, '2024': sum, 'individual': {...}},
                'DPK': {'2025': sum, '2024': sum, 'individual': {...}}
            }
        """
        result = {
            'ASET': {'2025': 0.0, '2024': 0.0},
            'KREDIT': {'2025': 0.0, '2024': 0.0, 'individual': {}},
            'DPK': {'2025': 0.0, '2024': 0.0, 'individual': {}}
        }
        
        try:
            page_source = self.driver.page_source
            soup = BeautifulSoup(page_source, 'html.parser')
            
            # Extract ASET
            aset_values = self._extract_identifier_value(soup, "Total Aset")
            result['ASET'] = aset_values
            self.logger.info(f"    ASET (Total Aset): 2025={aset_values['2025']:,.2f}, 2024={aset_values['2024']:,.2f}")
            
            # Extract KREDIT components
            kredit_identifiers = [
                "Kepada BPR",
                "Kepada Bank Umum",
                "Kepada non bank – pihak terkait",
                "Kepada non bank – pihak tidak terkait"
            ]
            
            for identifier in kredit_identifiers:
                values = self._extract_identifier_value(soup, identifier)
                result['KREDIT']['individual'][identifier] = values
                result['KREDIT']['2025'] += values['2025']
                result['KREDIT']['2024'] += values['2024']
                self.logger.info(f"    {identifier}: 2025={values['2025']:,.2f}, 2024={values['2024']:,.2f}")
            
            self.logger.info(f"    KREDIT Total: 2025={result['KREDIT']['2025']:,.2f}, 2024={result['KREDIT']['2024']:,.2f}")
            
            # Extract DPK components
            dpk_identifiers = [
                "Tabungan",
                "Deposito",
                "Simpanan dari Bank Lain"
            ]
            
            for identifier in dpk_identifiers:
                values = self._extract_identifier_value(soup, identifier)
                result['DPK']['individual'][identifier] = values
                result['DPK']['2025'] += values['2025']
                result['DPK']['2024'] += values['2024']
                self.logger.info(f"    {identifier}: 2025={values['2025']:,.2f}, 2024={values['2024']:,.2f}")
            
            self.logger.info(f"    DPK Total: 2025={result['DPK']['2025']:,.2f}, 2024={result['DPK']['2024']:,.2f}")
            
            return result
            
        except Exception as e:
            self.logger.error(f"  [ERROR] Error parsing Konvensional form 1: {e}")
            import traceback
            self.logger.debug(traceback.format_exc())
            return result
    
    def _parse_konvensional_form2(self) -> dict:
        """
        Parse Konvensional form 2 (BPK-901-000002) data
        
        Extracts:
        - LABA KOTOR: LABA (RUGI) TAHUN BERJALAN SEBELUM PAJAK PENGHASILAN
        - LABA BERSIH: JUMLAH LABA (RUGI) TAHUN BERJALAN
        
        Returns:
            dict with structure:
            {
                'LABA KOTOR': {'2025': value, '2024': value},
                'LABA BERSIH': {'2025': value, '2024': value}
            }
        """
        result = {
            'LABA KOTOR': {'2025': 0.0, '2024': 0.0},
            'LABA BERSIH': {'2025': 0.0, '2024': 0.0}
        }
        
        try:
            # Wait a bit for page to load
            time.sleep(3.0)
            
            # Get page source (check iframes first)
            page_source, report_iframe = self._get_page_source_with_iframe()
            soup = BeautifulSoup(page_source, 'html.parser')
            
            # Extract LABA KOTOR
            laba_kotor_values = self._extract_identifier_value(soup, "LABA (RUGI) TAHUN BERJALAN SEBELUM PAJAK PENGHASILAN")
            result['LABA KOTOR'] = laba_kotor_values
            self.logger.info(f"    LABA KOTOR: 2025={laba_kotor_values['2025']:,.2f}, 2024={laba_kotor_values['2024']:,.2f}")
            
            # Extract LABA BERSIH
            laba_bersih_values = self._extract_identifier_value(soup, "JUMLAH LABA (RUGI) TAHUN BERJALAN")
            result['LABA BERSIH'] = laba_bersih_values
            self.logger.info(f"    LABA BERSIH: 2025={laba_bersih_values['2025']:,.2f}, 2024={laba_bersih_values['2024']:,.2f}")
            
            return result
            
        except Exception as e:
            self.logger.error(f"  [ERROR] Error parsing Konvensional form 2: {e}")
            import traceback
            self.logger.debug(traceback.format_exc())
            return result
    
    def _parse_konvensional_form3(self) -> dict:
        """
        Parse Konvensional form 3 (BPK-901-000003) data
        
        Extracts ratios (separate, not summed):
        - KPMM: Kewajiban Penyediaan Modal Minimum (KPMM)
        - NPL: Non Performing Loan (NPL)
        - ROA: Return on Assets (ROA)
        - BOPO: Biaya Operasional terhadap Pendapatan Operasional (BOPO)
        - NIM: Net Interest Margin (NIM)
        - LDR: Loan to Deposit Ratio (LDR)
        - Cash Ratio: Cash Ratio
        
        Returns:
            dict with structure:
            {
                'KPMM': {'2025': value, '2024': value},
                'NPL': {'2025': value, '2024': value},
                ...
            }
        """
        result = {}
        
        ratios = [
            ("KPMM", "Kewajiban Penyediaan Modal Minimum (KPMM)"),
            ("NPL", "Non Performing Loan (NPL)"),
            ("ROA", "Return on Assets (ROA)"),
            ("BOPO", "Biaya Operasional terhadap Pendapatan Operasional (BOPO)"),
            ("NIM", "Net Interest Margin (NIM)"),
            ("LDR", "Loan to Deposit Ratio (LDR)"),
            ("Cash Ratio", "Cash Ratio")
        ]
        
        try:
            # Wait a bit for page to load
            time.sleep(3.0)
            
            # Get page source (check iframes first)
            page_source, report_iframe = self._get_page_source_with_iframe()
            soup = BeautifulSoup(page_source, 'html.parser')
            
            for ratio_name, identifier in ratios:
                # Use ratio extraction that preserves decimal points (like publikasi targeted scraping)
                values = self._extract_ratio_value(soup, identifier)
                result[ratio_name] = values
                self.logger.info(f"    {ratio_name}: 2025={values['2025']:.2f}, 2024={values['2024']:.2f}")
            
            return result
            
        except Exception as e:
            self.logger.error(f"  [ERROR] Error parsing Konvensional form 3: {e}")
            import traceback
            self.logger.debug(traceback.format_exc())
            return result
    
    def process_bank(self, bank_name: str, url_type: str):
        """
        Process a single bank using direct URL construction
        
        Args:
            bank_name: Bank name to process
            url_type: Type of URL (e.g., "BPR Konvensional" or "BPR Syariah")
        """
        # Get target month and year
        target_month_name, target_year = self._get_target_month_year()
        target_month_num = self._month_name_to_number(target_month_name)
        
        # Determine bank type
        bank_type = "syariah" if "syariah" in url_type.lower() else "konvensional"
        
        # Process all 3 forms
        form_numbers = [1, 2, 3] if bank_type == "konvensional" else [1, 2, 3]  # All forms for both types
        
        self.logger.info(f"  Processing bank: {bank_name[:80]}...")
        
        # Get bank code formats (original + expanded if needed)
        bank_code_formats = self._format_bank_code_for_url(bank_name)
        self.logger.debug(f"  Bank code formats to try: {bank_code_formats}")
        
        # Store bank data
        bank_data = {
            'bank_name': bank_name,
            'bank_type': bank_type,
            'form1': None,
            'form2': None,
            'form3': None
        }
        
        # Try each bank code format
        for format_idx, bank_code in enumerate(bank_code_formats, 1):
            self.logger.info(f"  Trying bank code format {format_idx}/{len(bank_code_formats)}: {bank_code}")
            
            format_success = False
            
            # Process each form
            server_error_on_form1 = False
            for form_num in form_numbers:
                self.logger.info(f"    Processing form {form_num}...")
                
                # If form 1 had server error, skip forms 2 and 3 and try next bank code format
                if server_error_on_form1 and form_num > 1:
                    self.logger.info(f"    Skipping form {form_num} (form 1 had server error, trying next bank code format)")
                    break
                
                # Build URL
                url = self._build_report_url(bank_code, target_month_num, target_year, bank_type, form_num)
                self.logger.debug(f"    URL: {url}")
                
                # Navigate to URL
                max_retries = 2
                parsed_data = None
                
                for retry_attempt in range(max_retries + 1):
                    try:
                        if retry_attempt > 0:
                            self.logger.info(f"    Retrying form {form_num} (attempt {retry_attempt + 1}/{max_retries + 1})...")
                            # Refresh the page on retry
                            self.driver.refresh()
                            time.sleep(3.0)
                        else:
                            self.driver.get(url)
                            time.sleep(2.0)  # Wait for page to load
                        
                        # Check for server error
                        if self._check_for_server_error():
                            self.logger.warning(f"    [WARNING] Server error for form {form_num}")
                            if form_num == 1:
                                server_error_on_form1 = True
                            break  # Break from retry loop, try next form or next format
                        
                        # Parse form data based on type and form number
                        if bank_type == "syariah":
                            if form_num == 1:
                                parsed_data = self._parse_syariah_form1()
                            elif form_num == 2:
                                parsed_data = self._parse_syariah_form2()
                            elif form_num == 3:
                                parsed_data = self._parse_syariah_form3()
                        else:  # konvensional
                            if form_num == 1:
                                parsed_data = self._parse_konvensional_form1()
                            elif form_num == 2:
                                parsed_data = self._parse_konvensional_form2()
                            elif form_num == 3:
                                parsed_data = self._parse_konvensional_form3()
                        
                        # If we got data, break from retry loop
                        if parsed_data:
                            break
                        else:
                            self.logger.warning(f"    [WARNING] Form {form_num} - no data parsed")
                            if retry_attempt < max_retries:
                                continue  # Retry
                            else:
                                break  # No more retries
                            
                    except Exception as e:
                        error_msg = str(e)
                        # Check if it's a Chrome timeout error
                        is_timeout_error = "timeout" in error_msg.lower() or "Timed out receiving message from renderer" in error_msg
                        
                        if is_timeout_error and retry_attempt < max_retries:
                            self.logger.warning(f"    [WARNING] Chrome timeout error on form {form_num}, refreshing page and retrying...")
                            try:
                                self.driver.refresh()
                                time.sleep(3.0)
                                continue  # Retry
                            except:
                                pass
                        
                        self.logger.error(f"    [ERROR] Error processing form {form_num}: {e}")
                        import traceback
                        self.logger.debug(traceback.format_exc())
                        
                        if retry_attempt < max_retries:
                            continue  # Retry
                        else:
                            break  # No more retries
                
                # Store parsed data if we got it
                if parsed_data:
                    if form_num == 1:
                        bank_data['form1'] = parsed_data
                    elif form_num == 2:
                        bank_data['form2'] = parsed_data
                    elif form_num == 3:
                        bank_data['form3'] = parsed_data
                    
                    format_success = True
                    self.logger.info(f"    [OK] Form {form_num} - data parsed successfully!")
                
                # If form 1 had server error, break from form loop to try next bank code format
                if server_error_on_form1:
                    break
            
            # If this format worked, we can break
            if format_success:
                self.logger.info(f"  [OK] Bank code format {format_idx} succeeded")
                break  # Break from format loop
        
        # Update existing bank data if it already exists, otherwise add new
        # Find if this bank already exists in all_bank_data
        existing_index = None
        for idx, existing_data in enumerate(self.all_bank_data):
            if existing_data.get('bank_name') == bank_name:
                existing_index = idx
                break
        
        if existing_index is not None:
            # Update existing bank data (for retry)
            self.all_bank_data[existing_index] = bank_data
            self.logger.info(f"  [INFO] Updated bank data for {bank_name}")
        else:
            # Add new bank data
            self.all_bank_data.append(bank_data)
        
        # Small delay before next bank
        time.sleep(1.0)
    
    def _retry_zero_value_banks_in_array(self):
        """
        Check all_bank_data array for zero values and retry those banks
        This is called before exporting to Excel to ensure we have complete data
        """
        try:
            banks_with_zero = []
            
            # Check each bank's data for zero values
            for bank_data in self.all_bank_data:
                bank_name = bank_data.get('bank_name', 'N/A')
                bank_type = bank_data.get('bank_type', 'konvensional')
                form1 = bank_data.get('form1') or {}
                form2 = bank_data.get('form2') or {}
                
                has_zero = False
                
                # Check form1 data (ASET, KREDIT/PIUTANG, DPK)
                if form1:
                    # Check ASET
                    aset = form1.get('ASET', {})
                    if isinstance(aset, dict):
                        if (aset.get('2025', 0) == 0 or aset.get('2024', 0) == 0):
                            has_zero = True
                    
                    # Check KREDIT (konvensional) or PIUTANG (syariah)
                    if bank_type == 'konvensional':
                        kredit = form1.get('KREDIT', {})
                        if isinstance(kredit, dict):
                            if (kredit.get('2025', 0) == 0 or kredit.get('2024', 0) == 0):
                                has_zero = True
                    else:  # syariah
                        piutang = form1.get('PIUTANG', {})
                        if isinstance(piutang, dict):
                            if (piutang.get('2025', 0) == 0 or piutang.get('2024', 0) == 0):
                                has_zero = True
                    
                    # Check DPK
                    dpk = form1.get('DPK', {})
                    if isinstance(dpk, dict):
                        if (dpk.get('2025', 0) == 0 or dpk.get('2024', 0) == 0):
                            has_zero = True
                
                # Check form2 data (LABA KOTOR, LABA BERSIH)
                if form2:
                    laba_kotor = form2.get('LABA KOTOR', {})
                    if isinstance(laba_kotor, dict):
                        if (laba_kotor.get('2025', 0) == 0 or laba_kotor.get('2024', 0) == 0):
                            has_zero = True
                    
                    laba_bersih = form2.get('LABA BERSIH', {})
                    if isinstance(laba_bersih, dict):
                        if (laba_bersih.get('2025', 0) == 0 or laba_bersih.get('2024', 0) == 0):
                            has_zero = True
                
                if has_zero:
                    banks_with_zero.append(bank_name)
            
            if not banks_with_zero:
                self.logger.info("[INFO] No banks with zero values found in stored data")
                return
            
            self.logger.info(f"[INFO] Found {len(banks_with_zero)} banks with zero values to retry")
            for bank_name in banks_with_zero:
                self.logger.info(f"  - {bank_name}")
            
            # Retry each bank
            for i, bank_name in enumerate(banks_with_zero, 1):
                self.logger.info("")
                self.logger.info("=" * 70)
                self.logger.info(f"[RETRY {i}/{len(banks_with_zero)}] Retrying: {bank_name}")
                self.logger.info("=" * 70)
                
                # Determine bank type
                bank_type = self._determine_bank_type(bank_name)
                url_type = "BPR Syariah" if bank_type == 'syariah' else "BPR Konvensional"
                
                # Process bank again (this will update all_bank_data)
                self.process_bank(bank_name, url_type)
                
                # Small delay
                time.sleep(1.0)
            
            self.logger.info("")
            self.logger.info(f"[OK] Retry completed for {len(banks_with_zero)} banks")
            
        except Exception as e:
            self.logger.error(f"[ERROR] Error retrying zero value banks: {e}")
            import traceback
            self.logger.error(traceback.format_exc())
    
    def _determine_bank_type(self, bank_name: str) -> str:
        """
        Determine which URL to check based on bank name
        
        Args:
            bank_name: Bank name to check
            
        Returns:
            'syariah' if bank contains BPRS or Syariah, 'konvensional' otherwise
        """
        bank_upper = bank_name.upper()
        
        # Check for BPRS or Syariah indicators
        if 'BPRS' in bank_upper or 'SYARIAH' in bank_upper:
            return 'syariah'
        else:
            return 'konvensional'
    
    def _copy_excel_to_destination_paths(self, filepath: Path, file_type: str = "sindikasi"):
        """
        Copy Excel file to destination paths
        
        Args:
            filepath: Path to the source Excel file
            file_type: Type of file - "publikasi" or "sindikasi"
        """
        if not filepath.exists():
            self.logger.warning(f"  [WARNING] Source file does not exist: {filepath}")
            return
        
        # Define destination paths
        destination_paths = []
        if file_type == "publikasi":
            destination_paths = [
                Path(r"D:\APP\OSS\client\assets-no_backup\publikasi"),
                Path(r"C:\Users\MSI\Desktop\OSS\client\assets-no_backup\publikasi")
            ]
        elif file_type == "sindikasi":
            destination_paths = [
                Path(r"D:\APP\OSS\client\assets-no_backup\sindikasi"),
                Path(r"C:\Users\MSI\Desktop\OSS\client\assets-no_backup\sindikasi")
            ]
        else:
            self.logger.warning(f"  [WARNING] Unknown file type: {file_type}")
            return
        
        # Copy to each destination
        for dest_dir in destination_paths:
            try:
                # Create destination directory if it doesn't exist
                dest_dir.mkdir(parents=True, exist_ok=True)
                
                # Copy file
                dest_file = dest_dir / filepath.name
                shutil.copy2(filepath, dest_file)
                self.logger.info(f"  [OK] Copied {filepath.name} to {dest_dir}")
            except Exception as e:
                self.logger.warning(f"  [WARNING] Failed to copy to {dest_dir}: {e}")
    
    def _extract_date_from_filename(self, filename: str) -> tuple:
        """
        Extract date from filename pattern: sindikasi_NAME_DD_MM_YYYY.txt
        
        Args:
            filename: Filename (e.g., "sindikasi_TestName_28_11_2025.txt")
            
        Returns:
            Tuple of (day, month, year) as strings, or (None, None, None) if not found
        """
        # Pattern: sindikasi_NAME_DD_MM_YYYY.txt
        match = re.search(r'sindikasi_.+_(\d{2})_(\d{2})_(\d{4})\.txt$', filename)
        if match:
            return match.group(1), match.group(2), match.group(3)  # DD, MM, YYYY
        
        return None, None, None
    
    def _calculate_peringkat(self, ratio_name: str, value: float) -> int:
        """
        Calculate Peringkat (Rating) for a given ratio value.
        
        All ratios use the % column value (the ratio value scraped).
        BPRS and BPR use the same logic, just different naming.
        
        Args:
            ratio_name: Name of the ratio (e.g., "ROA", "BOPO", "NIM", "NI", "KPMM", "Cash Ratio", "LDR", "FDR", "NPL", "NPF Neto")
            value: The ratio value from the % column
            
        Returns:
            int: Peringkat value (1-5)
        """
        if value is None:
            return 5  # Default to worst rating if value is missing
        
        # ROA
        if ratio_name == "ROA":
            if value >= 2:
                return 1
            elif 1.5 <= value < 2:
                return 2
            elif 1 <= value < 1.5:
                return 3
            elif 0.5 <= value < 1:
                return 4
            else:  # value < 0.5
                return 5
        
        # BOPO
        elif ratio_name == "BOPO":
            if value <= 85:
                return 1
            elif 85 < value <= 90:
                return 2
            elif 90 < value <= 95:
                return 3
            elif 95 < value <= 100:
                return 4
            else:  # value > 100
                return 5
        
        # NIM / NI (same logic, different naming)
        elif ratio_name in ["NIM", "NI"]:
            if value >= 10:
                return 1
            elif 8 <= value < 10:
                return 2
            elif 6 <= value < 8:
                return 3
            elif 4 <= value < 6:
                return 4
            else:  # value < 4
                return 5
        
        # KPMM
        elif ratio_name == "KPMM":
            if value >= 15:
                return 1
            elif 13 <= value < 15:
                return 2
            elif 12 <= value < 13:
                return 3
            elif 8 <= value < 12:
                return 4
            else:  # value < 8
                return 5
        
        # CR (Cash Ratio)
        elif ratio_name == "Cash Ratio":
            if value >= 20:
                return 1
            elif 15 <= value < 20:
                return 2
            elif 10 <= value < 15:
                return 3
            elif 5 <= value < 10:
                return 4
            else:  # value < 5
                return 5
        
        # LDR / FDR (same logic, different naming)
        elif ratio_name in ["LDR", "FDR"]:
            if value <= 90:
                return 1
            elif 90 < value <= 92.5:
                return 2
            elif 92.5 < value <= 95:
                return 3
            elif 95 < value <= 97.5:
                return 4
            else:  # value > 97.5
                return 5
        
        # NPL / NPF Neto (same logic, different naming)
        elif ratio_name in ["NPL", "NPF Neto"]:
            if value <= 5:
                return 1
            elif 5 < value <= 8:
                return 2
            elif 8 < value <= 11:
                return 3
            elif 11 < value <= 14:
                return 4
            else:  # value > 14
                return 5
        
        # Default: return worst rating for unknown ratios
        else:
            self.logger.warning(f"  [WARNING] Unknown ratio name for peringkat calculation: {ratio_name}, defaulting to 5")
            return 5
    
    def _create_excel_file(self, month: str, year: str, name: str = None, day: str = None, filename_month: str = None, filename_year: str = None):
        """
        Create Excel file with all bank data
        
        Structure:
        - Main Data Section: Nama BPR | 12 2025 | 12 2024 | YOY | % YOY
          Rows: ASET, PIUTANG/KREDIT, DPK, LABA KOTOR, LABA BERSIH
        - Rasio Section: RASIO | % | PERINGKAT
          Rows: All ratios from form 3
        
        Args:
            month: Month name (e.g., "Desember")
            year: Year (e.g., "2025")
        """
        if not Workbook:
            self.logger.error("  [ERROR] openpyxl not installed. Cannot create Excel file.")
            return
        
        if not self.all_bank_data:
            self.logger.warning("  [WARNING] No bank data to export")
            return
        
        try:
            # Sort banks by name
            sorted_banks = sorted(self.all_bank_data, key=lambda x: x['bank_name'])
            
            # Get month number
            month_num = self._month_name_to_number(month)
            prev_year = str(int(year) - 1)
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Sindikasi Report"
            
            # Define styles
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            center_align = Alignment(horizontal="center", vertical="center")
            
            # Number format to prevent scientific notation (use #,##0 for integers)
            number_format = '#,##0'
            
            row = 1
            
            # Column headers
            headers = ["Nama BPR", "", f"{month_num} {year}", f"{month_num} {prev_year}", "YOY", "% YOY"]
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = border
            row += 1
            
            # Main Data Rows
            for bank_data in sorted_banks:
                bank_name = bank_data.get('bank_name', 'N/A')
                bank_type = bank_data.get('bank_type', 'konvensional')
                form1 = bank_data.get('form1') or {}
                form2 = bank_data.get('form2') or {}
                
                # Determine category name (PIUTANG for Syariah, KREDIT for Konvensional)
                category_name = "PIUTANG" if bank_type == "syariah" else "KREDIT"
                
                # Safely get values from form1 and form2
                aset_values = form1.get('ASET', {}) if isinstance(form1, dict) else {}
                category_values = form1.get(category_name, {}) if isinstance(form1, dict) else {}
                dpk_values = form1.get('DPK', {}) if isinstance(form1, dict) else {}
                laba_kotor_values = form2.get('LABA KOTOR', {}) if isinstance(form2, dict) else {}
                laba_bersih_values = form2.get('LABA BERSIH', {}) if isinstance(form2, dict) else {}
                
                # Data rows for this bank
                data_rows = [
                    ("ASET", aset_values),
                    (category_name, category_values),
                    ("DPK", dpk_values),
                    ("LABA KOTOR", laba_kotor_values),
                    ("LABA BERSIH", laba_bersih_values)
                ]
                
                for label, values in data_rows:
                    if not isinstance(values, dict):
                        values = {}
                    val_2025 = values.get('2025', 0.0)
                    val_2024 = values.get('2024', 0.0)
                    yoy = val_2025 - val_2024
                    pct_yoy = (yoy / abs(val_2024) * 100) if val_2024 != 0 else 0.0
                    
                    ws.cell(row=row, column=1, value=bank_name).border = border
                    ws.cell(row=row, column=2, value=label).border = border
                    
                    # Format numeric cells to prevent scientific notation
                    cell_2025 = ws.cell(row=row, column=3, value=val_2025)
                    cell_2025.border = border
                    cell_2025.number_format = number_format
                    
                    cell_2024 = ws.cell(row=row, column=4, value=val_2024)
                    cell_2024.border = border
                    cell_2024.number_format = number_format
                    
                    cell_yoy = ws.cell(row=row, column=5, value=yoy)
                    cell_yoy.border = border
                    cell_yoy.number_format = number_format
                    
                    cell_pct = ws.cell(row=row, column=6, value=pct_yoy)
                    cell_pct.border = border
                    cell_pct.number_format = '0.00'  # Percentage with 2 decimals
                    row += 1
                
                # Empty row between banks
                row += 1
            
            # Rasio Section Header
            row += 2
            ws.merge_cells(f'A{row}:C{row}')
            rasio_header = ws.cell(row=row, column=1, value=f"TINGKAT KESEHATAN PERIODE {month_num} {year}")
            rasio_header.fill = header_fill
            rasio_header.font = header_font
            rasio_header.alignment = center_align
            rasio_header.border = border
            row += 1
            
            # Rasio Column headers
            rasio_headers = ["Nama BPR", "RASIO", "%", "PERINGKAT"]
            for col_idx, header in enumerate(rasio_headers, 1):
                cell = ws.cell(row=row, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = border
            row += 1
            
            # Rasio Rows
            for bank_data in sorted_banks:
                bank_name = bank_data.get('bank_name', 'N/A')
                form3 = bank_data.get('form3') or {}
                
                if not isinstance(form3, dict) or not form3:
                    continue
                
                # Get ratio order based on bank type
                bank_type = bank_data.get('bank_type', 'konvensional')
                if bank_type == "syariah":
                    ratio_order = ["ROA", "BOPO", "NI", "KPMM", "Cash Ratio", "FDR", "NPF Neto"]
                else:
                    ratio_order = ["ROA", "BOPO", "NIM", "KPMM", "Cash Ratio", "LDR", "NPL"]
                
                for ratio_name in ratio_order:
                    if ratio_name in form3:
                        values = form3[ratio_name]
                        if isinstance(values, dict):
                            val_2025 = values.get('2025', 0.0)
                        else:
                            val_2025 = 0.0
                        
                        # Calculate peringkat based on ratio value
                        peringkat = self._calculate_peringkat(ratio_name, val_2025)
                        
                        ws.cell(row=row, column=1, value=bank_name).border = border
                        ws.cell(row=row, column=2, value=ratio_name).border = border
                        
                        # Format ratio value with 2 decimal places
                        cell_ratio = ws.cell(row=row, column=3, value=val_2025)
                        cell_ratio.border = border
                        cell_ratio.number_format = '0.00'  # Number format with 2 decimal places
                        
                        # Set peringkat value
                        cell_peringkat = ws.cell(row=row, column=4, value=peringkat)
                        cell_peringkat.border = border
                        cell_peringkat.alignment = center_align
                        row += 1
                
                # Empty row between banks
                row += 1
            
            # Auto-adjust column widths
            for col in range(1, 7):
                max_length = 0
                column_letter = get_column_letter(col)
                for cell in ws[column_letter]:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
            
            # Save file to output/sindikasi subdirectory
            output_dir = Path(__file__).parent.parent / "output" / "sindikasi"
            output_dir.mkdir(parents=True, exist_ok=True)
            
            # Generate filename based on whether we have name and date components
            if name and day and filename_month and filename_year:
                # New format: Sindikasi_NAME_DD_MM_YYYY.xlsx
                # filename_month is MM from filename (list_DD_MM_YYYY)
                filename = f"Sindikasi_{name}_{day}_{filename_month}_{filename_year}.xlsx"
            else:
                # Fallback to old format: Sindikasi_MM_YYYY.xlsx
                # month_num here is from _month_name_to_number() (converted from month name)
                month_str = f"{month_num:02d}"
                filename = f"Sindikasi_{month_str}_{year}.xlsx"
            
            filepath = output_dir / filename
            
            # Check if file already exists and will be replaced
            if filepath.exists():
                self.logger.info(f"  [INFO] File already exists, replacing: {filepath.name}")
            
            wb.save(filepath)
            self.logger.info(f"  [OK] Excel file created: {filepath}")
            
            # Copy to destination paths
            self._copy_excel_to_destination_paths(filepath, "sindikasi")
            
        except Exception as e:
            self.logger.error(f"  [ERROR] Error creating Excel file: {e}")
            import traceback
            self.logger.debug(traceback.format_exc())
    
    def find_all_banks(self, list_file_path: Path):
        """
        Main orchestrator: Find all banks from list in appropriate URL
        
        Args:
            list_file_path: Path to the list file (should be named sindikasi_NAME_DD_MM_YYYY.txt)
        """
        # Clear previous data to ensure fresh start for each file
        self.all_bank_data = []
        
        # Extract date from filename
        filename = list_file_path.name
        day, month_num, year_num = self._extract_date_from_filename(filename)
        
        if not all([day, month_num, year_num]):
            self.logger.warning(f"Could not extract date from filename: {filename}. Expected format: sindikasi_NAME_DD_MM_YYYY.txt")
            # Fallback: use current date or default
            day, month_num, year_num = None, None, None
        
        # Read list file
        list_data = self.read_list_file(list_file_path)
        
        if not list_data['scrape']:
            self.logger.info("SCRAPE = FALSE, skipping bank search")
            return
        
        if not list_data['banks']:
            self.logger.warning("No banks found in list file")
            return
        
        # Get target month and year
        target_month, target_year = self._get_target_month_year()
        self.logger.info(f"Target period: {target_month} {target_year}")
        
        self.logger.info("=" * 70)
        self.logger.info("Starting bank search...")
        self.logger.info("=" * 70)
        
        # Initialize browser
        self.initialize()
        
        # Statistics
        found_konvensional = []  # List of (search_name, [matched_banks])
        found_syariah = []  # List of (search_name, [matched_banks])
        not_found = []
        
        try:
            # Process each bank
            for i, bank_name in enumerate(list_data['banks'], 1):
                self.logger.info("")
                self.logger.info("=" * 70)
                self.logger.info(f"[{i}/{len(list_data['banks'])}] Searching for: {bank_name}")
                self.logger.info("=" * 70)
                
                # Determine which URL type based on bank name
                bank_type = self._determine_bank_type(bank_name)
                
                if bank_type == 'syariah':
                    self.logger.info(f"  Bank type: BPR Syariah (contains BPRS or Syariah)")
                    url_type = "BPR Syariah"
                else:
                    self.logger.info(f"  Bank type: BPR Konvensional")
                    url_type = "BPR Konvensional"
                
                # Process bank directly using URL construction
                self.process_bank(bank_name, url_type)
                
                # Track for summary
                if bank_type == 'syariah':
                    found_syariah.append(bank_name)
                else:
                    found_konvensional.append(bank_name)
            
            # Print summary
            self.logger.info("")
            self.logger.info("=" * 70)
            self.logger.info("SEARCH SUMMARY")
            self.logger.info("=" * 70)
            self.logger.info(f"Target period: {target_month} {target_year}")
            self.logger.info(f"Total banks searched: {len(list_data['banks'])}")
            self.logger.info(f"Found in BPR Konvensional: {len(found_konvensional)}")
            self.logger.info(f"Found in BPR Syariah: {len(found_syariah)}")
            self.logger.info(f"Not found: {len(not_found)}")
            
            if found_konvensional:
                self.logger.info("")
                self.logger.info("Banks processed in BPR Konvensional:")
                for bank_name in found_konvensional:
                    self.logger.info(f"  - {bank_name}")
            
            if found_syariah:
                self.logger.info("")
                self.logger.info("Banks processed in BPR Syariah:")
                for bank_name in found_syariah:
                    self.logger.info(f"  - {bank_name}")
            
            if not_found:
                self.logger.info("")
                self.logger.warning("Banks NOT found in their respective URLs:")
                for bank in not_found:
                    self.logger.warning(f"  - {bank}")
            
            self.logger.info("=" * 70)
            
            # Check for zero values in stored array and retry before exporting to Excel
            if self.all_bank_data:
                self.logger.info("")
                self.logger.info("=" * 70)
                self.logger.info("Checking for zero values in stored data...")
                self.logger.info("=" * 70)
                self._retry_zero_value_banks_in_array()
            
            # Export to Excel
            if self.all_bank_data:
                self.logger.info("")
                self.logger.info("=" * 70)
                self.logger.info("Exporting data to Excel...")
                self.logger.info("=" * 70)
                # Pass NAME and date components if available
                self._create_excel_file(
                    target_month, 
                    target_year, 
                    name=list_data.get('name'),
                    day=day,
                    filename_month=month_num,
                    filename_year=year_num
                )
            
        except Exception as e:
            self.logger.error(f"Error during bank search: {e}")
            import traceback
            self.logger.error(traceback.format_exc())
        
        finally:
            self.cleanup()

