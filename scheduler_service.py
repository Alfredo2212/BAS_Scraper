"""
OJK Publikasi Scraper - Continuous Scheduler Service
Long-running service that runs the scraper automatically on schedule
Uses APScheduler to run every Tuesday and Thursday at 15:00 (GMT+7)
Also checks sindikasi queue every 10 minutes
"""

import sys
import logging
import re
import threading
import time
from pathlib import Path
from datetime import datetime
from apscheduler.schedulers.blocking import BlockingScheduler
from apscheduler.triggers.cron import CronTrigger
from apscheduler.triggers.interval import IntervalTrigger
import pytz

# Add the parent directory to path to import the module
sys.path.insert(0, str(Path(__file__).parent))

# Import using importlib to handle directory name with spaces
import importlib.util
module_path = Path(__file__).parent / "Laporan Publikasi BPR Konvensional" / "scraper.py"
spec = importlib.util.spec_from_file_location("scraper_module", module_path)
scraper_module = importlib.util.module_from_spec(spec)
spec.loader.exec_module(scraper_module)
OJKExtJSScraper = scraper_module.OJKExtJSScraper

# Import shared execution function
from scraper_runner import run_scraper_execution

# Import Sindikasi scraper
sindikasi_module_path = Path(__file__).parent / "Laporan Publikasi Sindikasi" / "scraper.py"
sindikasi_spec = importlib.util.spec_from_file_location("sindikasi_scraper_module", sindikasi_module_path)
sindikasi_scraper_module = importlib.util.module_from_spec(sindikasi_spec)
sindikasi_spec.loader.exec_module(sindikasi_scraper_module)
SindikasiScraper = sindikasi_scraper_module.SindikasiScraper

# Import IBPRS scraper
ibprs_module_path = Path(__file__).parent / "Laporan Bulanan IBPRS" / "scraper.py"
ibprs_spec = importlib.util.spec_from_file_location("ibprs_scraper_module", ibprs_module_path)
ibprs_scraper_module = importlib.util.module_from_spec(ibprs_spec)
ibprs_spec.loader.exec_module(ibprs_scraper_module)
IBPRSScraper = ibprs_scraper_module.IBPRSScraper

### GLOBAL STATE ###
# Track if scrapers are running
publikasi_running = False
sindikasi_running = False
ibprs_running = False
lock = threading.Lock()

### LOGGING SETUP ###
log_dir = Path(__file__).parent / "logs"
log_dir.mkdir(exist_ok=True)
log_file = log_dir / "scheduler.log"

# Configure logging to append to single log file
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_file, encoding='utf-8', mode='a'),
        logging.StreamHandler(sys.stdout)
    ]
)

logger = logging.getLogger(__name__)

### JOB FUNCTIONS ###
def run_scraper_job():
    """
    Job function that runs the publikasi scraper
    Called by APScheduler at scheduled times
    Uses shared execution function to ensure identical behavior with manual runner
    """
    global publikasi_running
    
    with lock:
        if publikasi_running:
            logger.warning("[WARNING] Publikasi scraper is already running, skipping...")
            return
        publikasi_running = True
    
    try:
        target_output_dir = Path(r"D:\APP\OSS\client\assets\publikasi")
        
        # Use shared execution function - ensures identical behavior with manual runner
        run_scraper_execution(
            scraper_class=OJKExtJSScraper,
            target_output_dir=target_output_dir,
            run_type="Scheduled Job"
        )
    finally:
        with lock:
            publikasi_running = False

def check_sindikasi_queue():
    """
    Check sindikasi queue folder for .txt files with SCRAPE = TRUE
    Runs every 10 minutes
    """
    global publikasi_running, sindikasi_running
    
    # Skip if publikasi or sindikasi is already running
    with lock:
        if publikasi_running:
            logger.debug("[DEBUG] Publikasi scraper is running, skipping sindikasi queue check")
            return
        if sindikasi_running:
            logger.debug("[DEBUG] Sindikasi scraper is already running, skipping queue check")
            return
        sindikasi_running = True
    
    try:
        queue_dir = Path(r"C:\Users\MSI\Desktop\OSS\client\assets-no_backup\sindikasi\queue")
        
        if not queue_dir.exists():
            logger.debug(f"[DEBUG] Queue directory does not exist: {queue_dir}")
            return
        
        # Find .txt files matching pattern: sindikasi_NAME_DD_MM_YYYY.txt
        txt_files = []
        for file in queue_dir.glob("sindikasi_*.txt"):
            if re.match(r'sindikasi_.+_\d{2}_\d{2}_\d{4}\.txt$', file.name):
                txt_files.append(file)
        
        if not txt_files:
            logger.debug("[DEBUG] No sindikasi queue files found")
            return
        
        # Process each file
        for txt_file in sorted(txt_files):
            should_set_false = False
            try:
                # Read file to check SCRAPE flag
                with open(txt_file, 'r', encoding='utf-8') as f:
                    content = f.read()
                
                # Check if SCRAPE = TRUE
                scrape_match = re.search(r'SCRAPE\s*=\s*(TRUE|FALSE)', content, re.IGNORECASE)
                if not scrape_match or scrape_match.group(1).upper() != 'TRUE':
                    logger.debug(f"[DEBUG] File {txt_file.name} has SCRAPE = FALSE, skipping")
                    continue
                
                # Mark that we should set FALSE after processing
                should_set_false = True
                
                logger.info("=" * 70)
                logger.info(f"Found sindikasi queue file with SCRAPE = TRUE: {txt_file.name}")
                logger.info("=" * 70)
                
                # Extract NAME and date from filename
                # Pattern: sindikasi_NAME_DD_MM_YYYY.txt
                filename_match = re.match(r'sindikasi_(.+)_(\d{2})_(\d{2})_(\d{4})\.txt$', txt_file.name)
                if not filename_match:
                    logger.warning(f"[WARNING] Could not parse filename: {txt_file.name}")
                    continue
                
                name_from_file = filename_match.group(1)
                day = filename_match.group(2)
                month = filename_match.group(3)
                year = filename_match.group(4)
                
                logger.info(f"Processing: NAME={name_from_file}, Date={day}/{month}/{year}")
                
                # Run sindikasi scraper
                run_sindikasi_scraper(txt_file, name_from_file, day, month, year)
                
                logger.info(f"[OK] Completed processing {txt_file.name}")
                
            except Exception as e:
                logger.error(f"[ERROR] Error processing {txt_file.name}: {e}")
                import traceback
                logger.error(traceback.format_exc())
            finally:
                # Always set SCRAPE = FALSE after processing (success or failure)
                if should_set_false:
                    try:
                        update_scrape_flag(txt_file, False)
                        logger.info(f"[OK] Set SCRAPE = FALSE in {txt_file.name}")
                    except Exception as e:
                        logger.error(f"[ERROR] Failed to update SCRAPE flag in {txt_file.name}: {e}")
    
    finally:
        with lock:
            sindikasi_running = False

def run_ibprs_scraper_job():
    """
    Job function that runs the IBPRS scraper
    Called by APScheduler every Wednesday at 15:00
    Runs in headless mode
    """
    global ibprs_running
    
    with lock:
        if ibprs_running:
            logger.warning("[WARNING] IBPRS scraper is already running, skipping...")
            return
        ibprs_running = True
    
    scraper = None
    try:
        logger.info("=" * 70)
        logger.info("IBPRS Scraper - Scheduled Job Started")
        logger.info(f"Started at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        logger.info("=" * 70)
        
        # Initialize scraper in headless mode
        logger.info("[INFO] Initializing IBPRS scraper (headless mode)...")
        scraper = IBPRSScraper(headless=True)
        
        # Run the complete scrape and save workflow
        success = scraper.scrape_and_save("Kep. Riau")
        
        if success:
            logger.info("=" * 70)
            logger.info("IBPRS Scraper - Scheduled Job Completed Successfully")
            logger.info(f"Completed at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            logger.info("Files saved to: D:\\APP\\OSS\\client\\assets-no_backup\\ibprs")
            logger.info("=" * 70)
        else:
            logger.error("=" * 70)
            logger.error("IBPRS Scraper - Scheduled Job Failed")
            logger.error(f"Failed at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            logger.error("=" * 70)
            
    except Exception as e:
        logger.error(f"[ERROR] Error running IBPRS scraper: {e}")
        import traceback
        logger.error(traceback.format_exc())
    finally:
        if scraper:
            try:
                scraper.cleanup()
            except:
                pass
        with lock:
            ibprs_running = False

def run_sindikasi_scraper(list_file: Path, name: str, day: str, month: str, year: str):
    """
    Run sindikasi scraper and retry for zero values
    
    Args:
        list_file: Path to the list file
        name: Name extracted from filename
        day: Day from filename
        month: Month from filename
        year: Year from filename
    """
    scraper = None
    try:
        logger.info("[INFO] Initializing sindikasi scraper...")
        scraper = SindikasiScraper(headless=True)
        
        # Initialize and run scraper
        scraper.initialize()
        scraper.find_all_banks(list_file)
        
        # After scraping, check for zero values and retry
        logger.info("[INFO] Checking for zero values and retrying...")
        retry_zero_value_banks_sindikasi(scraper, name, day, month, year)
        
    except Exception as e:
        logger.error(f"[ERROR] Error running sindikasi scraper: {e}")
        import traceback
        logger.error(traceback.format_exc())
    finally:
        if scraper:
            try:
                scraper.cleanup()
            except:
                pass

def retry_zero_value_banks_sindikasi(scraper, name: str, day: str, month: str, year: str):
    """
    Check sindikasi Excel for zero values and retry those banks
    
    Args:
        scraper: SindikasiScraper instance
        name: Name from filename
        day: Day from filename
        month: Month from filename
        year: Year from filename
    """
    try:
        from openpyxl import load_workbook
        
        # Find the Excel file
        output_dir = Path(__file__).parent / "output" / "sindikasi"
        filename = f"Sindikasi_{name}_{day}_{month}_{year}.xlsx"
        filepath = output_dir / filename
        
        if not filepath.exists():
            logger.warning(f"[WARNING] Excel file not found: {filepath}")
            return
        
        logger.info(f"[INFO] Reading Excel file: {filename}")
        wb = load_workbook(filepath)
        ws = wb.active
        
        # Find banks with zero values
        # Excel structure: Each bank has 5 rows (ASET, KREDIT/PIUTANG, DPK, LABA KOTOR, LABA BERSIH)
        # Columns: A = Nama BPR, B = Label, C = 2025 value, D = 2024 value
        banks_with_zero = set()  # Use set to avoid duplicates
        
        current_bank = None
        for row_num in range(2, ws.max_row + 1):
            bank_name_cell = ws.cell(row=row_num, column=1).value
            label_cell = ws.cell(row=row_num, column=2).value
            val_2025 = ws.cell(row=row_num, column=3).value
            val_2024 = ws.cell(row=row_num, column=4).value
            
            # Skip empty rows
            if not bank_name_cell and not label_cell:
                current_bank = None
                continue
            
            # If we have a bank name in column A, it's a data row for that bank
            if bank_name_cell:
                current_bank = str(bank_name_cell).strip()
            
            # Check if any value is zero or None (only check if we have a label, meaning it's a data row)
            if current_bank and label_cell and ((val_2025 == 0 or val_2025 == 0.0 or val_2025 is None) or \
                                                (val_2024 == 0 or val_2024 == 0.0 or val_2024 is None)):
                banks_with_zero.add(current_bank)
        
        wb.close()
        
        if not banks_with_zero:
            logger.info("[INFO] No banks with zero values found")
            return
        
        banks_with_zero = list(banks_with_zero)
        logger.info(f"[INFO] Found {len(banks_with_zero)} banks with zero values to retry")
        
        # Re-initialize browser if needed
        if scraper.driver is None:
            scraper.initialize()
        
        # Retry each bank
        for i, bank_name in enumerate(banks_with_zero, 1):
            logger.info(f"[{i}/{len(banks_with_zero)}] Retrying: {bank_name}")
            
            # Determine bank type
            bank_type = scraper._determine_bank_type(bank_name)
            url_type = "BPR Syariah" if bank_type == 'syariah' else "BPR Konvensional"
            
            # Process bank again (this will update all_bank_data)
            scraper.process_bank(bank_name, url_type)
            
            # Small delay
            time.sleep(1.0)
        
        # Re-create Excel file with updated data
        logger.info("[INFO] Regenerating Excel file with retry data...")
        target_month, target_year = scraper._get_target_month_year()
        scraper._create_excel_file(
            target_month,
            target_year,
            name=name,
            day=day,
            filename_month=month,
            filename_year=year
        )
        
        logger.info(f"[OK] Retry completed for {len(banks_with_zero)} banks")
        
    except Exception as e:
        logger.error(f"[ERROR] Error retrying zero value banks: {e}")
        import traceback
        logger.error(traceback.format_exc())

def update_scrape_flag(txt_file: Path, value: bool):
    """
    Update SCRAPE flag in txt file
    
    Args:
        txt_file: Path to the txt file
        value: True or False
    """
    try:
        with open(txt_file, 'r', encoding='utf-8') as f:
            lines = f.readlines()
        
        # Update SCRAPE line
        updated_lines = []
        for line in lines:
            if re.match(r'SCRAPE\s*=', line, re.IGNORECASE):
                updated_lines.append(f"SCRAPE = {'TRUE' if value else 'FALSE'}\n")
            else:
                updated_lines.append(line)
        
        # Write back
        with open(txt_file, 'w', encoding='utf-8') as f:
            f.writelines(updated_lines)
        
        logger.info(f"[OK] Updated SCRAPE = {'TRUE' if value else 'FALSE'} in {txt_file.name}")
        
    except Exception as e:
        logger.error(f"[ERROR] Error updating SCRAPE flag in {txt_file.name}: {e}")

### SCHEDULER SETUP ###
def get_next_run_times(scheduler):
    """Calculate and return next scheduled run times"""
    jobs = scheduler.get_jobs()
    next_runs = []
    tz = pytz.timezone('Asia/Jakarta')
    now = datetime.now(tz)
    
    for job in jobs:
        try:
            # Use trigger's get_next_fire_time to calculate next run
            # First parameter is previous fire time (None for first run), second is now
            next_run = job.trigger.get_next_fire_time(None, now)
            if next_run:
                # Convert to Jakarta timezone if needed
                if next_run.tzinfo is None:
                    next_run = tz.localize(next_run)
                next_runs.append(next_run.strftime('%Y-%m-%d %H:%M:%S'))
        except Exception as e:
            logger.debug(f"Could not get next run time for job {job.id}: {e}")
    
    return sorted(next_runs) if next_runs else []

def main():
    """Main function - starts the continuous scheduler service"""
    logger.info("=" * 70)
    logger.info("OJK Publikasi Scraper - Scheduler Service Starting")
    logger.info(f"Started at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    logger.info("=" * 70)
    
    # Create scheduler with GMT+7 timezone (Asia/Jakarta)
    scheduler = BlockingScheduler(timezone=pytz.timezone('Asia/Jakarta'))
    
    # Schedule job for Tuesday at 15:00
    scheduler.add_job(
        run_scraper_job,
        trigger=CronTrigger(day_of_week='tue', hour=15, minute=0),
        id='ojk_scraper_tuesday',
        name='OJK Scraper - Tuesday 15:00',
        replace_existing=True
    )
    
    # Schedule job for Thursday at 15:00
    scheduler.add_job(
        run_scraper_job,
        trigger=CronTrigger(day_of_week='thu', hour=15, minute=0),
        id='ojk_scraper_thursday',
        name='OJK Scraper - Thursday 15:00',
        replace_existing=True
    )
    
    # Schedule sindikasi queue checker every 10 minutes
    scheduler.add_job(
        check_sindikasi_queue,
        trigger=IntervalTrigger(minutes=10),
        id='sindikasi_queue_checker',
        name='Sindikasi Queue Checker - Every 10 minutes',
        replace_existing=True
    )
    
    # Schedule IBPRS scraper for Wednesday at 15:00
    scheduler.add_job(
        run_ibprs_scraper_job,
        trigger=CronTrigger(day_of_week='wed', hour=15, minute=0),
        id='ibprs_scraper_wednesday',
        name='IBPRS Scraper - Wednesday 15:00',
        replace_existing=True
    )
    
    # Get and display next run times
    try:
        next_runs = get_next_run_times(scheduler)
        logger.info("Scheduler Started - waiting for Tue/Thu 15:00")
        if next_runs:
            logger.info("Next scheduled runs:")
            for next_run in next_runs:
                logger.info(f"  - {next_run}")
        else:
            logger.info("Next scheduled runs: (calculating...)")
    except Exception as e:
        logger.warning(f"Could not calculate next run times: {e}")
        logger.info("Scheduler Started - waiting for Tue/Thu 15:00")
    
    logger.info("=" * 70)
    logger.info("Scheduler is running:")
    logger.info("  - Publikasi scraper: Tuesday and Thursday at 15:00")
    logger.info("  - IBPRS scraper: Wednesday at 15:00 (headless)")
    logger.info("  - Sindikasi queue checker: Every 10 minutes")
    logger.info("Press Ctrl+C to stop.")
    logger.info("=" * 70)
    
    try:
        # Start the scheduler (this will block and run forever)
        scheduler.start()
    except (KeyboardInterrupt, SystemExit):
        logger.info("=" * 70)
        logger.info("Scheduler service stopped by user")
        logger.info(f"Stopped at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        logger.info("=" * 70)
        scheduler.shutdown()
    except Exception as e:
        logger.error(f"[ERROR] Scheduler service error: {e}")
        import traceback
        logger.error(traceback.format_exc())
        scheduler.shutdown()
        sys.exit(1)

if __name__ == "__main__":
    main()

